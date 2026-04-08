# version: 1.0
# Moduł: gui_settings
# ⏹ KONIEC WSTĘPU

from __future__ import annotations

import copy
import datetime
import glob
import io
import json
import os
import shutil
import subprocess
import sys
import threading
import re
import tkinter as tk
from pathlib import Path
from typing import Any, Callable, Dict
from tkinter import colorchooser
from tkinter import filedialog, messagebox, scrolledtext, ttk

import importlib.util

if importlib.util.find_spec("ui_scrolled_frame"):
    from ui_scrolled_frame import ScrolledFrame
else:
    # Prosty, niezawodny ScrolledFrame jako fallback
    class ScrolledFrame(ttk.Frame):
        def __init__(self, parent, **kwargs):
            super().__init__(parent, **kwargs)
            canvas = tk.Canvas(self, highlightthickness=0, bg="#111214")
            scrollbar = ttk.Scrollbar(self, orient="vertical", command=canvas.yview)
            self.inner = ttk.Frame(canvas)

            self.inner.bind(
                "<Configure>",
                lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
            )

            canvas.create_window((0, 0), window=self.inner, anchor="nw")
            canvas.configure(yscrollcommand=scrollbar.set)

            canvas.pack(side="left", fill="both", expand=True)
            scrollbar.pack(side="right", fill="y")
from logging import getLogger

from ui_theme import ensure_theme_applied, get_theme_color
from config_manager import ConfigError, ConfigManager, PATH_MAP, get_root, resolve_rel
from core.logging_config import init_logging
from core.crash_handler import (
    clear_crash_log,
    get_crash_log_path,
    get_crash_log_stats,
    mark_crash_log_read,
)
from tools_config_loader import load_config as load_tools_config
from utils_json import ensure_json
from config.paths import (
    bind_settings,
    data_path,
    ensure_core_tree,
    p_config,
    p_settings_schema,
    p_tools_defs,
    resolve,
)
from gui_uzytkownicy import panel_uzytkownicy
from gui_settings_users_tab import create_users_tab

logger = getLogger(__name__)


TAB_OGOLNE = "Ogólne"
TAB_USERS = "Użytkownicy"
TAB_TOOLS = "Narzędzia"
TAB_PATHS = "Ścieżki"
TAB_BACKUP = "Kopia zapasowa"
TAB_UI = "UI i wygląd"
TAB_MODULES = "Moduły"
TAB_JARVIS = "Jarvis / powiadomienia"
TAB_ADVANCED = "Zaawansowane / Dev"


PATH_SAVE_INFO = (
    "Po wybraniu katalogu przyciskiem \"Wybierz…\" ścieżka Folderu WM "
    "(root) zapisywana jest automatycznie w pliku config.json w tym katalogu. "
    "Aplikacja utworzy też katalogi logs/ i backup/, jeśli ich brakuje."
)

PATH_EXAMPLES_TEXT = (
    "Przykładowe ścieżki:\n"
    "  <root>/maszyny.json\n"
    "  <root>/narzedzia/typy_narzedzi.json\n"
    "  <root>/narzedzia/001.json\n"
    "  <root>/zlecenia/zlecenia.json\n"
    "  <root>/logs/"
)

try:
    from PIL import Image  # type: ignore
except Exception:  # pragma: no cover - PIL opcjonalne
    Image = None

# >>> PATCH START: import narzędzi z Excela
try:
    from openpyxl import Workbook, load_workbook  # type: ignore
    from openpyxl.utils import get_column_letter  # type: ignore
except Exception:  # pragma: no cover - zależność opcjonalna
    load_workbook = None
    Workbook = None  # type: ignore
    get_column_letter = None  # type: ignore
# <<< PATCH END: import narzędzi z Excela

def _default_tools_definitions_path() -> str:
    """Return the preferred path for ``zadania_narzedzia.json``."""

    try:
        return str(p_tools_defs(ConfigManager()))
    except Exception:
        return str(Path("zadania_narzedzia.json").resolve())


def _wm_read_textwidget(widget) -> str:
    """Zwraca cały tekst z Text/ScrolledText, jeżeli to taki widżet; w innym wypadku pusty string."""

    try:
        return widget.get("1.0", "end").strip()
    except Exception:
        return ""


def _wm_copy_to_clipboard(owner, text: str) -> None:
    """Kopiuje tekst do schowka systemowego + krótka notyfikacja/log."""

    try:
        owner.clipboard_clear()
        owner.clipboard_append(text)
        owner.update()
        size = len(text.encode("utf-8")) if isinstance(text, str) else 0
        logger.info("[AUDIT][COPY] Skopiowano do schowka (%s bajtów).", size)
        messagebox.showinfo("Kopiuj raport", "Raport skopiowany do schowka.")
    except Exception as e:
        logger.exception("[AUDIT][COPY] Błąd kopiowania do schowka")
        messagebox.showerror("Kopiuj raport", f"Nie udało się skopiować:\n{e}")


def _format_diag_path(value: str, root: str, *, warn: bool) -> tuple[str, str]:
    """Return formatted diagnostic text and colour for path preview."""

    text = value or ""
    default_color = "#9ca3af"
    if not text or not warn:
        return text, default_color

    try:
        resolved_value = os.path.normcase(os.path.abspath(text))
        resolved_root = os.path.normcase(os.path.abspath(root))
    except Exception:
        return f"{text}  ⚠ poza <root>", "#dc2626"

    try:
        common = os.path.commonpath([resolved_value, resolved_root])
    except Exception:
        return f"{text}  ⚠ poza <root>", "#dc2626"

    if common != resolved_root:
        return f"{text}  ⚠ poza <root>", "#dc2626"
    return text, default_color


def _wm_read_latest_audit_from_disk(cfg_manager) -> str:
    """Szuka najnowszego pliku audytu w katalogu logs_dir (paths.logs_dir) i zwraca jego treść."""

    try:
        if cfg_manager is None:
            logger.warning("[AUDIT][COPY] Brak menedżera konfiguracji.")
            return ""
        cfg = {}
        if hasattr(cfg_manager, "load") and callable(getattr(cfg_manager, "load")):
            cfg = cfg_manager.load() or {}
        elif hasattr(cfg_manager, "merged"):
            cfg = getattr(cfg_manager, "merged", {}) or {}
        if not isinstance(cfg, dict):
            cfg = {}
        root = get_root(cfg)
        logs_dir = resolve_rel(cfg, "root.logs")
        if not logs_dir:
            paths_cfg = (cfg.get("paths") or {})
            legacy = (paths_cfg.get("logs_dir") or "").strip()
            if legacy:
                logs_dir = legacy
            elif root:
                rel = PATH_MAP.get("root.logs", "logs")
                logs_dir = os.path.join(root, rel)
        if not logs_dir or not os.path.isdir(logs_dir):
            logger.warning("[AUDIT][COPY] Brak/niepoprawny logs_dir: %r", logs_dir)
            return ""
        pattern = os.path.join(logs_dir, "audyt_wm-*.txt")
        files = sorted(glob.glob(pattern), key=os.path.getmtime, reverse=True)
        if not files:
            logger.info("[AUDIT][COPY] Nie znaleziono plików audytu (%s).", pattern)
            return ""
        latest = files[0]
        with io.open(latest, "r", encoding="utf-8") as f:
            content = f.read().strip()
        logger.info(
            "[AUDIT][COPY] Wczytano najnowszy raport: %s (%s B)", latest, len(content)
        )
        return content
    except Exception:
        logger.exception("[AUDIT][COPY] Wyjątek przy odczycie raportu z dysku")
        return ""


def _wm_copy_audit_report(owner, cfg_manager, text_widget=None):
    """Główna akcja przycisku. Najpierw próbuje z widżetu tekstowego; jak pusty → najnowszy plik z logs_dir."""

    text = ""
    if text_widget is not None:
        text = _wm_read_textwidget(text_widget)
    if not text:
        text = _wm_read_latest_audit_from_disk(cfg_manager)
    if not text:
        messagebox.showwarning("Kopiuj raport", "Brak treści raportu do skopiowania.")
        return
    _wm_copy_to_clipboard(owner, text)


# >>> PATCH START: import narzędzi z Excela
def _normalize_excel_cell(value: Any) -> str:
    """Convert Excel cell value to a trimmed string."""

    if value is None:
        return ""
    if isinstance(value, datetime.datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, datetime.date):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    return str(value).strip()


def _extract_tools_records_from_excel(excel_path: str) -> tuple[list[dict[str, Any]], int, int]:
    """Read Excel file and return tool records with counters."""

    if load_workbook is None:
        raise RuntimeError("Brak biblioteki openpyxl – zainstaluj pakiet openpyxl.")
    if not os.path.exists(excel_path):
        raise FileNotFoundError(excel_path)

    workbook = load_workbook(excel_path, data_only=True, read_only=True)
    try:
        sheet = None
        for name in ("Narzędzia", "narzędzia", "Narzedzia", "narzedzia"):
            if name in workbook.sheetnames:
                sheet = workbook[name]
                break
        if sheet is None:
            sheet = workbook[workbook.sheetnames[0]]

        rows = sheet.iter_rows(values_only=True)
        headers: list[str] = []
        for header_row in rows:
            header_values = [_normalize_excel_cell(cell) for cell in header_row]
            if any(header_values):
                headers = header_values
                break
        if not headers:
            return [], 0, 0

        normalized_headers = [header.lower() for header in headers]
        records: list[dict[str, Any]] = []
        total_rows = 0
        skipped_rows = 0

        for raw_row in rows:
            if raw_row is None:
                continue
            values = list(raw_row)
            if not any(_normalize_excel_cell(cell) for cell in values):
                continue
            total_rows += 1

            row_map: dict[str, str] = {}
            for idx, header in enumerate(normalized_headers):
                if not header:
                    continue
                cell_value = values[idx] if idx < len(values) else None
                row_map[header] = _normalize_excel_cell(cell_value)

            tool_id = (
                row_map.get("id")
                or row_map.get("nr")
                or row_map.get("numer")
                or ""
            ).strip()
            if not tool_id:
                skipped_rows += 1
                continue

            name = row_map.get("nazwa", "").strip() or tool_id
            tool_type = row_map.get("typ", "").strip()
            status = row_map.get("status", "").strip() or "OK"

            entry: dict[str, Any] = {
                "id": tool_id,
                "nr": tool_id,
                "nazwa": name,
                "typ": tool_type,
                "status": status,
                "zadania": [],
            }

            description = row_map.get("opis", "").strip()
            if description:
                entry["opis"] = description
            location = row_map.get("lokalizacja", "").strip()
            if location:
                entry["lokalizacja"] = location
            notes = row_map.get("uwagi", "").strip()
            if notes:
                entry["uwagi"] = notes

            records.append(entry)

        return records, total_rows, skipped_rows
    finally:
        try:
            workbook.close()
        except Exception:
            pass


def _sanitize_tool_id(raw: str) -> str:
    """Return filesystem-safe identifier for tool file name."""

    cleaned = re.sub(r"[\\/:*?\"<>|]", "_", str(raw or "").strip())
    cleaned = cleaned.replace(" ", "_")
    return cleaned or "narzedzie"


def _resolve_tools_dir_from_cfg(cfg_manager: ConfigManager | None) -> str | None:
    """Resolve tools directory using configuration manager."""

    manager = cfg_manager
    if manager is None:
        try:
            manager = ConfigManager()
        except Exception:
            manager = None

    cfg_snapshot: dict[str, Any] = {}
    if manager is not None:
        try:
            cfg_snapshot = manager.load() or {}
        except Exception:
            cfg_snapshot = getattr(manager, "merged", {}) or {}

    for key in ("tools.dir", "tools_dir", "paths.tools_dir"):
        try:
            candidate = resolve_rel(cfg_snapshot, key)
        except Exception:
            candidate = None
        if candidate:
            return os.path.normpath(candidate)

    root_dir = get_root(cfg_snapshot)
    if root_dir:
        rel = PATH_MAP.get("tools.dir", "narzedzia")
        return os.path.normpath(os.path.join(root_dir, rel))

    if manager is not None:
        try:
            root_path = manager.path_root()
        except Exception:
            root_path = ""
        if root_path:
            rel = PATH_MAP.get("tools.dir", "narzedzia")
            return os.path.normpath(os.path.join(root_path, rel))

    return None


def _import_tools_from_excel_file(
    excel_path: str, output_dir: str
) -> tuple[int, int, int]:
    """Import tools from Excel file into JSON files within ``output_dir``."""

    records, total_rows, skipped_rows = _extract_tools_records_from_excel(excel_path)
    os.makedirs(output_dir, exist_ok=True)

    saved = 0
    for entry in records:
        identifier = _sanitize_tool_id(entry.get("id", ""))
        file_path = os.path.join(output_dir, f"{identifier}.json")
        with open(file_path, "w", encoding="utf-8") as handle:
            json.dump(entry, handle, ensure_ascii=False, indent=2)
        saved += 1
    return total_rows, saved, skipped_rows


def export_tools_to_excel(output_path: str, tools_dir: str) -> int:
    """Export existing tools (JSON files) to an Excel workbook (.xlsx).

    Returns number of exported tools.
    """

    if Workbook is None:
        raise RuntimeError("Brak biblioteki openpyxl – nie można eksportować do Excela.")

    if not tools_dir or not os.path.isdir(tools_dir):
        raise FileNotFoundError(f"Katalog narzędzi nie istnieje: {tools_dir}")

    tools: list[dict[str, Any]] = []
    for file_path in sorted(glob.glob(os.path.join(tools_dir, "*.json"))):
        try:
            with open(file_path, "r", encoding="utf-8") as handle:
                obj = json.load(handle)
        except Exception:
            continue

        if not isinstance(obj, dict):
            continue

        if not obj.get("id") and not obj.get("nazwa"):
            continue

        tools.append(obj)

    wb = Workbook()
    ws = wb.active
    ws.title = "narzedzia"

    headers = [
        "id",
        "nazwa",
        "typ",
        "nn_sn",
        "status",
        "lokalizacja",
        "opis",
        "aktywny",
    ]
    ws.append(headers)

    for tool in tools:
        ws.append(
            [
                tool.get("id", ""),
                tool.get("nazwa", ""),
                tool.get("typ", ""),
                tool.get("nn_sn", ""),
                tool.get("status", ""),
                tool.get("lokalizacja", ""),
                tool.get("opis", ""),
                tool.get("active", True),
            ]
        )

    if get_column_letter is not None:
        for col in ws.columns:
            try:
                max_len = 0
                for cell in col:
                    val = cell.value
                    if val is None:
                        continue
                    max_len = max(max_len, len(str(val)))
                ws.column_dimensions[get_column_letter(col[0].column)].width = min(
                    max_len + 2, 40
                )
            except Exception:
                pass

    wb.save(output_path)
    return len(tools)


# <<< PATCH END: import narzędzi z Excela

TabHandler = Callable[[tk.Widget, dict[str, Any], tuple[str, ...]], tuple[int, int] | None]


def _safe_pick_json(
    owner: tk.Misc | None,
    reason: str = "",
    *,
    title: str | None = None,
    filetypes: list[tuple[str, str]] | None = None,
) -> str | None:
    """Return file path picked via dialog unless bootstrap is active."""

    try:
        from start import BOOTSTRAP_ACTIVE
    except Exception:
        BOOTSTRAP_ACTIVE = False

    if BOOTSTRAP_ACTIVE:
        logger.info(
            "[FILEDIALOG] Zablokowano dialog podczas bootstrapa (powód: %s)",
            reason or "nie podano",
        )
        return None

    kwargs: dict[str, Any] = {}
    if owner is not None:
        kwargs["parent"] = owner
    kwargs["title"] = title or "Wybierz plik JSON"
    if filetypes:
        kwargs["filetypes"] = filetypes
    else:
        kwargs["filetypes"] = [("Plik JSON", "*.json")]

    path = filedialog.askopenfilename(**kwargs)
    return path or None


def _safe_save_json(
    owner: tk.Misc | None,
    reason: str = "",
    *,
    title: str | None = None,
    defaultextension: str = ".json",
    initialfile: str | None = None,
    filetypes: list[tuple[str, str]] | None = None,
) -> str | None:
    """Return save path picked via dialog unless bootstrap is active."""

    try:
        from start import BOOTSTRAP_ACTIVE
    except Exception:
        BOOTSTRAP_ACTIVE = False

    if BOOTSTRAP_ACTIVE:
        logger.info(
            "[FILEDIALOG] Zablokowano dialog zapisu podczas bootstrapa (powód: %s)",
            reason or "nie podano",
        )
        return None

    kwargs: dict[str, Any] = {
        "defaultextension": defaultextension,
    }
    if owner is not None:
        kwargs["parent"] = owner
    if title:
        kwargs["title"] = title
    if initialfile:
        kwargs["initialfile"] = initialfile
    if filetypes:
        kwargs["filetypes"] = filetypes
    else:
        kwargs["filetypes"] = [("Plik JSON", "*.json")]

    parent = owner if owner is not None else None
    with TopMost(parent, grab=False):
        try:
            _ensure_topmost(parent)
        except Exception:
            pass
        path = filedialog.asksaveasfilename(**kwargs)
    return path or None


# =========================
# R-07: System → Folder <root> + status ścieżek
# =========================
def _exists(path: str) -> bool:
    try:
        return os.path.exists(path)
    except Exception:
        return False


def _mk_status(lbl_parent, name: str, ok: bool, path: str) -> None:
    success_color = get_theme_color("success", fallback="#22c55e")
    danger_color = get_theme_color("error", fallback="#ef4444")
    color = success_color if ok else danger_color
    row = ttk.Frame(lbl_parent)
    ttk.Label(row, text=name, width=26).pack(side="left")
    tag = ttk.Label(row, text=("OK" if ok else "BRAK"), foreground=color)
    tag.pack(side="left", padx=6)
    muted = get_theme_color("fg_dim", fallback="#9ca3af")
    ttk.Label(row, text=path, foreground=muted).pack(side="left")
    row.pack(fill="x", padx=8, pady=2)


def _build_root_section(
    parent, cm: ConfigManager | None, *, owner: "SettingsPanel" | None = None, on_root_change=None
) -> None:
    cfg: dict[str, Any]
    if cm is None:
        try:
            cm = ConfigManager()
        except Exception:
            cm = None

    try:
        cfg = cm.load() if cm is not None else {}
    except Exception:
        cfg = {}

    if not isinstance(cfg, dict):
        cfg = {}

    root = get_root(cfg)
    box = ttk.Labelframe(parent, text="Folder WM (<root>) i status plików")

    row_top = ttk.Frame(box)
    var_path = tk.StringVar(value=root or "")
    entry = ttk.Entry(row_top, textvariable=var_path, width=60)
    entry.pack(side="left", padx=6, pady=6)

    def _pick_dir(key: str = "<root>") -> None:
        title = (
            "Wybierz katalog <root>"
            if key in {"<root>", "root"}
            else "Wybierz katalog"
        )
        selected = filedialog.askdirectory(title=title)
        if not selected:
            return

        cfg_key = "paths.data_root" if key in {"<root>", "root", "paths.<root>"} else key

        try:
            normalized = os.path.normpath(selected)
        except Exception:
            normalized = selected

        try:
            os.makedirs(normalized, exist_ok=True)
        except Exception:
            logger.exception(
                "[SETTINGS] Nie udało się utworzyć katalogu '%s' dla klucza %s.",
                normalized,
                key,
            )
            return

        manager = cm
        if manager is None:
            try:
                manager = ConfigManager()
            except Exception:
                manager = None

        if manager is None:
            logger.error("[SETTINGS] Brak menedżera konfiguracji – nie zapisano %s.", key)
            return

        try:
            manager.set(cfg_key, normalized)
            if cfg_key == "paths.data_root":
                try:
                    manager.update_root_paths(normalized)
                except Exception:
                    logger.exception("[SETTINGS] Aktualizacja katalogu root nieudana")
                backup_dir = manager.get("paths.backup_dir") or os.path.join(
                    manager.path_root(), "backup"
                )
                logs_dir = manager.get("paths.logs_dir") or os.path.join(
                    manager.path_root(), "logs"
                )
                assets_dir = manager.get("paths.assets_dir") or manager.path_assets()
                for directory in (backup_dir, logs_dir, assets_dir):
                    try:
                        os.makedirs(directory, exist_ok=True)
                    except Exception:
                        logger.exception(
                            "[SETTINGS] Nie udało się utworzyć katalogu %s", directory
                        )
            manager.save_all()
            var_path.set(normalized)
            logger.info("[SETTINGS] Ustawiono %s = %s", key, normalized)
            if callable(on_root_change):
                try:
                    on_root_change(normalized)
                except Exception:
                    logger.exception("[SETTINGS] on_root_change callback failed")
        except Exception:
            logger.exception(
                "[SETTINGS] Nie udało się zapisać ścieżki %s.",
                key,
            )

    ttk.Button(row_top, text="Wybierz…", command=_pick_dir).pack(side="left", padx=6)
    row_top.pack(fill="x")

    actions_row = ttk.Frame(box)
    actions_row.pack(fill="x", padx=6, pady=(0, 4))
    ttk.Button(
        actions_row,
        text="Wyczyść stare ścieżki plików (legacy)",
        command=_reset_legacy_file_overrides,
    ).pack(side="left")

    _add_readonly_info(box, PATH_SAVE_INFO, label="Informacja o zapisie")

    ttk.Label(
        box,
        text=PATH_EXAMPLES_TEXT,
        justify="left",
    ).pack(fill="x", padx=8, pady=(0, 6))

    paths = {
        "Maszyny": resolve_rel(cfg, "machines"),
        "Tło hali": resolve_rel(cfg, "machines_bg"),
        "Narzędzia.idx": resolve_rel(cfg, "tools_index"),
        "Magazyn": resolve_rel(cfg, "warehouse_stock"),
        "Zlecenia": resolve_rel(cfg, "orders"),
        "BOM": resolve_rel(cfg, "bom"),
    }

    for name, path in paths.items():
        _mk_status(box, name, _exists(path), path or "—")

    root_dir = get_root(cfg) if cfg else ""
    required_dirs = {
        "Katalog data/": os.path.join(root_dir, "data") if root_dir else "",
        "Katalog assets/": os.path.join(root_dir, "assets") if root_dir else "",
        "Katalog logs/": os.path.join(root_dir, "logs") if root_dir else "",
    }
    for name, path in required_dirs.items():
        exists = bool(path and os.path.isdir(path))
        _mk_status(box, name, exists, path or "<brak root>")

    diag_labels: dict[str, ttk.Label] = {}
    diag_frame = ttk.LabelFrame(box, text="Diagnostyka ścieżek")
    diag_frame.pack(fill="x", padx=8, pady=(4, 8))

    def _current_cfg_manager() -> ConfigManager | None:
        if owner is not None and hasattr(owner, "cfg"):
            return getattr(owner, "cfg", None)
        return cm

    manager_obj = _current_cfg_manager()
    paths_preview = {
        "root": "",
        "config": "",
        "data": "",
        "backup": "",
        "logs": "",
        "assets": "",
    }
    if manager_obj is not None:
        try:
            paths_preview["root"] = manager_obj.path_root()
        except Exception:
            pass
        try:
            paths_preview["config"] = manager_obj.get_config_path()
        except Exception:
            pass
        try:
            paths_preview["data"] = manager_obj.path_data()
        except Exception:
            pass
        try:
            paths_preview["backup"] = manager_obj.path_backup()
        except Exception:
            pass
        try:
            paths_preview["logs"] = manager_obj.path_logs()
        except Exception:
            pass
        try:
            paths_preview["assets"] = manager_obj.path_assets()
        except Exception:
            pass

    labels_map = {
        "root": "Folder WM",
        "config": "config.json",
        "data": "katalog danych",
        "backup": "backup/",
        "logs": "logs/",
        "assets": "assets/",
    }

    warning_keys = {"data", "backup", "logs", "assets"}

    root_value = paths_preview.get("root", "")

    for key, label_text in labels_map.items():
        row = ttk.Frame(diag_frame)
        row.pack(fill="x", padx=6, pady=1)
        ttk.Label(row, text=f"{label_text}:", width=16).pack(side="left")
        text, color = _format_diag_path(
            paths_preview.get(key, ""),
            root_value,
            warn=bool(root_value) and key in warning_keys,
        )
        value_label = ttk.Label(row, text=text, foreground=color)
        value_label.pack(side="left", fill="x", expand=True)
        diag_labels[key] = value_label

    if owner is not None:
        setattr(owner, "_root_paths_labels", diag_labels)

    box.pack(fill="x", padx=8, pady=8)


def _add_readonly_info(parent: tk.Widget, text: str, *, label: str | None = None) -> ttk.Frame:
    """Adds an informational, read-only note to the given parent widget."""

    frame = ttk.Frame(parent)
    frame.pack(fill="x", padx=5, pady=(6, 0))
    if label:
        ttk.Label(frame, text=label).pack(anchor="w")
    ttk.Label(frame, text=text, wraplength=520, justify="left").pack(anchor="w")
    return frame


def _add_machines_bg_group(parent, cfg_manager):
    frm = ttk.LabelFrame(parent, text="Tło hali (renderer)")
    frm.pack(fill="x", padx=8, pady=8)

    machines_cfg = {}
    try:
        machines_cfg = cfg_manager.get("machines", {}) if cfg_manager else {}
    except Exception:
        machines_cfg = {}

    bg_var = tk.StringVar(value=(machines_cfg.get("background_image") or ""))
    w_var = tk.IntVar(value=int(machines_cfg.get("bg_required_w", 1920) or 0))
    h_var = tk.IntVar(value=int(machines_cfg.get("bg_required_h", 1080) or 0))
    ext_default = machines_cfg.get("bg_allowed_ext", ".jpg,.png") or ".jpg,.png"
    ext_var = tk.StringVar(value=str(ext_default))

    def _browse_bg():
        picked = filedialog.askopenfilename(
            parent=parent,
            title="Wybierz tło hali",
            filetypes=[
                ("Obrazy", "*.jpg;*.jpeg;*.png;*.bmp;*.gif"),
                ("Wszystkie pliki", "*.*"),
            ],
        )
        if picked:
            bg_var.set(picked)

    def _validate_and_save():
        if cfg_manager is None:
            messagebox.showerror("Tło hali", "Brak menedżera konfiguracji.", parent=parent)
            return

        path = (bg_var.get() or "").strip()
        req_w = int(w_var.get() or 0)
        req_h = int(h_var.get() or 0)
        allowed = [
            ext.strip().lower()
            for ext in (ext_var.get() or ".jpg,.png").split(",")
            if ext.strip()
        ]

        if path:
            ext = os.path.splitext(path)[1].lower()
            if allowed and ext and ext not in allowed:
                messagebox.showerror(
                    "Tło – błąd",
                    f"Nieakceptowalne rozszerzenie: {ext}\nDozwolone: {', '.join(allowed)}",
                    parent=parent,
                )
                return

            if Image is not None and os.path.exists(path):
                try:
                    with Image.open(path) as im:
                        w_img, h_img = im.size
                    if (req_w and w_img != req_w) or (req_h and h_img != req_h):
                        if not messagebox.askyesno(
                            "Tło – rozdzielczość",
                            (
                                "Wybrane tło ma "
                                f"{w_img}×{h_img}px, wymagane {req_w}×{req_h}px.\nZapisać mimo to?"
                            ),
                            parent=parent,
                        ):
                            return
                except Exception:
                    logger.exception("[SETTINGS] Nie udało się odczytać rozdzielczości tła (PIL)")
            elif Image is None:
                logger.warning("[SETTINGS] PIL niedostępny – pomijam sprawdzenie rozdzielczości obrazu.")

        try:
            cfg_manager.set("machines.background_image", path)
            cfg_manager.set("machines.bg_required_w", req_w)
            cfg_manager.set("machines.bg_required_h", req_h)
            cfg_manager.set("machines.bg_allowed_ext", ",".join(allowed))
            cfg_manager.save_all()
        except Exception:
            logger.exception("[SETTINGS] Nie udało się zapisać ustawień tła hali")
            messagebox.showerror(
                "Tło hali",
                "Nie udało się zapisać ustawień tła.",
                parent=parent,
            )
            return

        messagebox.showinfo("Tło hali", "Zapisano ustawienia tła i walidacji.", parent=parent)

    row = 0
    ttk.Label(frm, text="Plik tła:").grid(row=row, column=0, sticky="w")
    ttk.Entry(frm, textvariable=bg_var, width=60).grid(row=row, column=1, sticky="we", padx=6)
    ttk.Button(frm, text="Przeglądaj…", command=_browse_bg).grid(row=row, column=2, sticky="e")
    row += 1

    ttk.Label(frm, text="Wymagana szerokość (px):").grid(row=row, column=0, sticky="w", pady=(6, 0))
    ttk.Spinbox(frm, from_=0, to=10000, textvariable=w_var, width=8).grid(
        row=row, column=1, sticky="w", pady=(6, 0)
    )
    row += 1

    ttk.Label(frm, text="Wymagana wysokość (px):").grid(row=row, column=0, sticky="w")
    ttk.Spinbox(frm, from_=0, to=10000, textvariable=h_var, width=8).grid(row=row, column=1, sticky="w")
    row += 1

    ttk.Label(frm, text="Dozwolone rozszerzenia (CSV):").grid(row=row, column=0, sticky="w")
    ttk.Entry(frm, textvariable=ext_var, width=30).grid(row=row, column=1, sticky="w")

    ttk.Button(frm, text="Zapisz tło i walidację", command=_validate_and_save).grid(
        row=row, column=2, sticky="e", padx=6
    )

    for col in range(3):
        frm.columnconfigure(col, weight=1 if col == 1 else 0)

    return frm


def _add_machines_map_group(parent, cfg_manager):
    frm = ttk.LabelFrame(parent, text="Mapa hali – etykieta i rozmiar kropki")
    frm.pack(fill="x", padx=8, pady=8)

    machines_cfg = {}
    try:
        machines_cfg = cfg_manager.get("machines", {}) if cfg_manager else {}
    except Exception:
        logger.exception("[SETTINGS] Nie udało się wczytać konfiguracji mapy")
        machines_cfg = {}

    label_mode = tk.StringVar(value=(machines_cfg.get("map_label") or "id"))
    dot_radius = tk.IntVar(value=int(machines_cfg.get("map_dot_radius") or 18))

    ttk.Label(frm, text="Etykieta w kropce:").grid(row=0, column=0, sticky="w")
    ttk.Combobox(
        frm,
        state="readonly",
        width=12,
        values=["id", "typ", "nazwa"],
        textvariable=label_mode,
    ).grid(row=0, column=1, sticky="w", padx=6)

    ttk.Label(frm, text="Promień kropki (px):").grid(row=1, column=0, sticky="w", pady=(6, 0))
    ttk.Spinbox(frm, from_=10, to=40, width=6, textvariable=dot_radius).grid(
        row=1, column=1, sticky="w", padx=6, pady=(6, 0)
    )

    def _save():
        if cfg_manager is None:
            messagebox.showerror("Mapa hali", "Brak menedżera konfiguracji.", parent=parent)
            return
        try:
            cfg_manager.set("machines.map_label", label_mode.get())
            cfg_manager.set("machines.map_dot_radius", int(dot_radius.get() or 18))
            cfg_manager.save_all()
            messagebox.showinfo("Mapa hali", "Zapisano ustawienia mapy.")
        except Exception:
            logger.exception("[SETTINGS] Nie udało się zapisać ustawień mapy")
            messagebox.showerror("Mapa hali", "Nie udało się zapisać ustawień mapy.")

    ttk.Button(frm, text="Zapisz", command=_save).grid(row=0, column=2, rowspan=2, sticky="e", padx=6)

    for col in range(3):
        frm.columnconfigure(col, weight=1 if col == 1 else 0)

    return frm


from gui.settings_action_handlers import (
    bind as settings_actions_bind,
    execute as settings_action_exec,
)

try:
    from wm_log import (
        bind_settings_getter as wm_bind_settings_getter,
        dbg as wm_dbg,
        err as wm_err,
        info as wm_info,
    )
except ImportError:  # pragma: no cover - fallback for environments without wm_log
    def wm_bind_settings_getter(_getter):
        return None


    def wm_dbg(*_args, **_kwargs):
        return None


    def wm_err(*_args, **_kwargs):
        return None


    def wm_info(*_args, **_kwargs):
        return None

def _theme_frame_background(widget: tk.Misc | None = None) -> str:
    """Return the active theme background colour for frame-like widgets."""

    try:
        style = ttk.Style(widget)
        for style_name in ("WM.Container.TFrame", "TFrame", "."):
            value = style.lookup(style_name, "background")
            if value:
                return str(value)
    except Exception:
        pass
    return "#111214"


class ScrollableFrame(ttk.Frame):
    """Generic vertically scrollable frame with mouse wheel support."""

    _SCROLL_LINES = 3

    def __init__(self, parent: tk.Misc, *args: object, **kwargs: object) -> None:
        super().__init__(parent, *args, **kwargs)
        self.canvas = tk.Canvas(self, highlightthickness=0, borderwidth=0)
        self._canvas_alive = True
        self.canvas.bind("<Destroy>", lambda e: setattr(self, "_canvas_alive", False))
        self.vsb = ttk.Scrollbar(self, orient="vertical", command=self.canvas.yview)
        self.canvas.configure(yscrollcommand=self.vsb.set)

        frame_bg = _theme_frame_background(self)
        self.inner = tk.Frame(self.canvas, bg=frame_bg)
        self._window = self.canvas.create_window((0, 0), window=self.inner, anchor="nw")

        self.inner.bind("<Configure>", self._on_inner_configure)
        self.canvas.bind("<Configure>", self._on_canvas_configure)

        self.canvas.grid(row=0, column=0, sticky="nsew")
        self.vsb.grid(row=0, column=1, sticky="ns")
        self.grid_rowconfigure(0, weight=1)
        self.grid_columnconfigure(0, weight=1)

        self.canvas.bind_all("<MouseWheel>", self._on_mousewheel)
        self.canvas.bind_all("<Button-4>", lambda event: self._scroll(-self._SCROLL_LINES))
        self.canvas.bind_all("<Button-5>", lambda event: self._scroll(self._SCROLL_LINES))
        top = self.winfo_toplevel()
        top.bind("<Destroy>", self._on_toplevel_destroy, add="+")

    def _on_canvas_configure(self, event: tk.Event) -> None:
        """Ensure the inner frame matches the canvas width."""

        if not self._canvas_alive:
            return
        try:
            if self.canvas.winfo_exists():
                self.canvas.itemconfigure(self._window, width=event.width)
        except tk.TclError:
            self._canvas_alive = False
            print("[WM-DBG][SETTINGS] Ignoruję configure po zniszczeniu canvas (TclError)")

    def _on_inner_configure(self, _event: tk.Event) -> None:
        """Update the scrollregion when the inner frame changes size."""

        if not self._canvas_alive:
            return
        try:
            if self.canvas.winfo_exists():
                self.canvas.configure(scrollregion=self.canvas.bbox("all"))
        except tk.TclError:
            self._canvas_alive = False
            print("[WM-DBG][SETTINGS] Ignoruję update scrollregion po zniszczeniu canvas (TclError)")

    def _on_toplevel_destroy(self, _event: tk.Event) -> None:
        """Unbind global scroll handlers when the settings window is closed."""

        try:
            self.canvas.unbind_all("<MouseWheel>")
            self.canvas.unbind_all("<Button-4>")
            self.canvas.unbind_all("<Button-5>")
        except tk.TclError:
            pass

    def _on_mousewheel(self, event: tk.Event) -> None:
        """Scroll canvas when mouse wheel is used (Windows/Linux)."""

        delta = getattr(event, "delta", 0) or 0
        if not delta:
            return

        lines_float = (delta / 120) * self._SCROLL_LINES
        lines = -int(round(lines_float))

        if lines == 0:
            lines = -1 if delta > 0 else 1

        self._scroll(lines)

    def _scroll(self, units: int) -> None:
        """Perform vertical scrolling by the given unit delta, handling widget teardown."""

        c = getattr(self, "canvas", None)
        if not c or not self._canvas_alive:
            print("[WM-DBG][SETTINGS] Scroll przerwany: canvas nie istnieje/already destroyed")
            return
        try:
            if c.winfo_exists():
                c.yview_scroll(units, "units")
            else:
                print("[WM-DBG][SETTINGS] Scroll przerwany: winfo_exists=False")
        except tk.TclError:
            # Canvas został zniszczony w trakcie callbacku – ignorujemy
            self._canvas_alive = False
            print("[WM-DBG][SETTINGS] Ignoruję scroll po zniszczeniu canvas (TclError)")

# A-2e: alias do edytora advanced (wyszukiwarka, limity, kolekcje NN/SN)
try:
    from gui_tools_config import ToolsConfigDialog  # preferuje advanced, fallback prosty
except Exception:  # pragma: no cover - środowisko bez gui_tools_config
    ToolsConfigDialog = None

import config_manager as cm
from config_manager import ConfigManager, set_path
from gui_products import ProductsMaterialsTab
from ustawienia_magazyn import MagazynSettingsFrame
import ustawienia_produkty_bom
from ui_utils import TopMost, _ensure_topmost
import logika_zadan as LZ
from profile_utils import SIDEBAR_MODULES
from services import profile_service
from logger import log_akcja

from zlecenia_utils import DEFAULT_ORDER_TYPES


def _mag_dict_path() -> Path:
    try:
        cfg = ConfigManager()
        return Path(cfg.path_data("magazyn/slowniki.json"))
    except Exception:
        return Path("data") / "magazyn" / "slowniki.json"  # [ROOT] fallback awaryjny


MAG_DICT_PATH = _mag_dict_path()


_SYSTEM_LEGACY_KEYS = {
    "hall.machines_file",
    "tools.file",
    "orders.file",
    "warehouse.file",
    "profiles.file",
    "warehouse.stock_source",
    "warehouse.reservations_file",
    "tools.types_file",
    "tools.statuses_file",
    "tools.task_templates_file",
    "bom.file",
}


def _is_legacy_system_field(field_def: dict[str, Any]) -> bool:
    key = field_def.get("key")
    if not key:
        return False
    if key in _SYSTEM_LEGACY_KEYS:
        return True
    if key.endswith("_pick") and key[:-5] in _SYSTEM_LEGACY_KEYS:
        return True
    return False


def _reset_legacy_file_overrides() -> None:
    """Usuwa legacy-ścieżki plików tak, aby użyć Folderu WM."""

    for legacy_key in _SYSTEM_LEGACY_KEYS:
        try:
            set_path(legacy_key, "", who="settings.ui")
        except Exception:
            pass
    messagebox.showinfo(
        "WM – Ścieżki",
        "Wyczyszczono indywidualne ścieżki plików.\n"
        "Moduły będą używać lokalizacji względem Folderu WM.",
    )


def _make_system_tab(
    parent: tk.Misc, cfg_manager, *, owner: "SettingsPanel" | None = None, on_root_change=None
) -> tk.Misc:
    status_refresh: dict[str, Any] = {}

    def _handle_root_change(new_root: str) -> None:
        if callable(on_root_change):
            try:
                on_root_change(new_root)
            except Exception:
                logger.exception("[SETTINGS] root change handler failed")
        refresh_cb = status_refresh.get("fn")
        if callable(refresh_cb):
            try:
                refresh_cb()
            except Exception:
                logger.exception("[SETTINGS] root status refresh failed")

    _build_root_section(parent, cfg_manager, owner=owner, on_root_change=_handle_root_change)

    status_holder = ttk.Frame(parent)
    status_holder.pack(fill="x", expand=True, padx=4, pady=(8, 4))
    parent._root_status_widget = None  # type: ignore[attr-defined]

    def rebuild_status() -> None:
        for widget in status_holder.winfo_children():
            widget.destroy()
        table = _build_root_status(status_holder, cfg_manager)
        parent._root_status_widget = table  # type: ignore[attr-defined]

    status_refresh["fn"] = rebuild_status
    if owner is not None:
        setattr(owner, "_root_status_refresh", rebuild_status)

    rebuild_status()

    btns = ttk.Frame(parent)
    btns.pack(fill="x", expand=False, padx=4, pady=(4, 8))

    ttk.Button(
        btns,
        text="Utwórz brakujące pliki teraz",
        command=lambda: _init_root_resources(
            parent, cfg_manager, rebuild_status_cb=rebuild_status
        ),
    ).pack(side="left")

    return parent


def _root_status_rows(cfg: dict):
    rows = []
    what = [
        ("Maszyny", "machines"),
        ("Magazyn", "warehouse"),
        ("BOM", "bom"),
        ("Profile", "profiles"),
        ("Narzędzia (katalog)", "tools_dir"),
        ("Zlecenia (katalog)", "orders_dir"),
        ("Def. narzędzi (plik)", "tools_defs"),
    ]
    for label, key in what:
        path = resolve_rel(cfg, key)
        exists = bool(path and os.path.exists(path))
        kind = (
            "plik"
            if key in {"machines", "warehouse", "bom", "profiles", "tools_defs"}
            else "katalog"
        )
        rows.append((label, kind, path or "", exists))
    return rows


def _build_root_status(parent, cfg_manager):
    manager = cfg_manager
    if manager is None:
        try:
            manager = ConfigManager()
        except Exception:
            manager = None

    cfg: dict[str, Any] = {}
    if manager is not None:
        try:
            cfg = manager.load() if hasattr(manager, "load") else {}
        except Exception:
            cfg = {}

    if not isinstance(cfg, dict):
        return None

    frame = ttk.LabelFrame(
        parent, text="Status plików/katalogów względem Folderu WM (root)"
    )
    tv = ttk.Treeview(
        frame,
        columns=("label", "typ", "sciezka", "ok"),
        show="headings",
        height=7,
    )
    tv.heading("label", text="Pozycja")
    tv.column("label", width=200, anchor="w")
    tv.heading("typ", text="Typ")
    tv.column("typ", width=90, anchor="w")
    tv.heading("sciezka", text="Ścieżka")
    tv.column("sciezka", width=420, anchor="w")
    tv.heading("ok", text="OK?")
    tv.column("ok", width=60, anchor="center")

    for label, kind, path, exists in _root_status_rows(cfg):
        tv.insert(
            "",
            "end",
            values=(label, kind, path, "✔" if exists else "✖"),
            tags=("ok" if exists else "bad",),
        )

    tv.tag_configure("ok", background="#1e3a1e")
    tv.tag_configure("bad", background="#3a1e1e")
    tv.pack(fill="x", expand=True, padx=6, pady=6)
    frame.pack(fill="x", expand=True, padx=4, pady=4)
    return frame


ROOT_DEFAULTS = {
    "machines": {"maszyny": []},
    "warehouse": {"items": []},
    "bom": {"produkty": []},
    "tools.types": {"types": []},
    "tools.statuses": {"statuses": []},
    "tools.tasks": {"tasks": []},
    "tools.zadania": {"zadania": []},
    "orders": {"zlecenia": []},
    "tools.dir": None,
    "root.logs": None,
    "root.backup": None,
}


def _init_root_resources(owner, cfg_manager, rebuild_status_cb=None):
    """
    Tworzy brakujące zasoby pod wybranym <root> wg PATH_MAP:
      • pliki JSON z minimalnymi szablonami (ROOT_DEFAULTS)
      • katalogi (narzedzia/, logs/, backup/)
    Po zakończeniu odświeża tabelę statusu, jeśli podano rebuild_status_cb().
    """

    manager = cfg_manager
    if manager is None:
        try:
            manager = ConfigManager()
        except Exception:
            manager = None

    try:
        if manager is not None and hasattr(manager, "load"):
            cfg = manager.load()
        elif manager is not None and hasattr(manager, "merged"):
            cfg = getattr(manager, "merged", {})
        else:
            cfg = {}
    except Exception:
        cfg = {}

    if not isinstance(cfg, dict):
        cfg = {}

    root = get_root(cfg)
    if not root:
        messagebox.showwarning(
            "Folder WM (root)",
            "Najpierw ustaw Folder WM (root) w zakładce System.",
            parent=owner,
        )
        return

    created: list[tuple[str, str]] = []
    skipped: list[tuple[str, str]] = []
    errors: list[tuple[str, str, str]] = []

    for key, tmpl in ROOT_DEFAULTS.items():
        abs_path = resolve_rel(cfg, key)
        rel = PATH_MAP.get(key, "")

        if not rel or not abs_path:
            skipped.append((key, rel))
            continue

        try:
            if tmpl is None:
                os.makedirs(abs_path, exist_ok=True)
                created.append((key, abs_path))
            else:
                os.makedirs(os.path.dirname(abs_path), exist_ok=True)
                if not os.path.exists(abs_path):
                    ensure_json(abs_path, default=tmpl)
                    created.append((key, abs_path))
                else:
                    skipped.append((key, abs_path))
        except Exception as exc:
            logger.exception(
                "[ROOT-INIT] Błąd tworzenia %s (%s): %s", key, abs_path, exc
            )
            errors.append((key, abs_path, str(exc)))

    msg: list[str] = []
    if created:
        msg.append(f"Utworzono: {len(created)}")
        for key, path in created[:10]:
            msg.append(f"  + {key}: {path}")
        if len(created) > 10:
            msg.append("  ...")

    if skipped:
        msg.append(f"Pominięto (istnieją): {len(skipped)}")

    if errors:
        msg.append(f"Błędy: {len(errors)}")
        for key, path, err in errors[:5]:
            msg.append(f"  ! {key}: {path} -> {err}")
        if len(errors) > 5:
            msg.append("  ...")

    messagebox.showinfo(
        "Folder WM (root)",
        "\n".join(msg) if msg else "Brak zmian.",
        parent=owner,
    )

    if callable(rebuild_status_cb):
        try:
            rebuild_status_cb()
        except Exception:
            pass


def _is_deprecated(node: dict) -> bool:
    """Return True if schema node is marked as deprecated."""

    return node.get("deprecated") is True


def _normalize_field_definition(option: dict[str, Any]) -> dict[str, Any]:
    """Return copy of ``option`` with canonicalized type metadata."""

    normalized = dict(option)
    raw_type = str(normalized.get("type") or "").lower()
    type_map = {
        "bool": "bool",
        "boolean": "bool",
        "int": "int",
        "integer": "int",
        "float": "float",
        "number": "float",
        "double": "float",
        "enum": "enum",
        "select": "enum",
        "choice": "enum",
        "array": "array",
        "list": "array",
        "dict": "dict",
        "object": "dict",
        "string": "string",
        "text": "string",
        "path": "path",
    }
    normalized_type = type_map.get(raw_type, raw_type or "string")
    normalized["type"] = normalized_type

    if normalized_type == "enum":
        allowed = (
            normalized.get("enum")
            or normalized.get("allowed")
            or normalized.get("values")
            or []
        )
        str_allowed: list[str] = []
        for item in allowed:
            str_allowed.append(str(item))
        options = normalized.get("options")
        labels_map: dict[str, str] = {}
        if isinstance(options, list):
            for opt in options:
                if not isinstance(opt, dict):
                    continue
                value = opt.get("value", opt.get("label"))
                if value is None:
                    continue
                value_str = str(value)
                label = opt.get("label")
                labels_map[value_str] = str(label) if label is not None else value_str
                if value_str not in str_allowed:
                    str_allowed.append(value_str)
        if str_allowed:
            normalized["enum"] = str_allowed
        normalized.pop("values", None)
        normalized.pop("allowed", None)
        if labels_map:
            normalized["enum_labels"] = labels_map

    if normalized_type == "array" and "value_type" not in normalized:
        items = normalized.get("items")
        if isinstance(items, dict):
            item_type = str(items.get("type") or "").lower()
            if item_type in {"float", "number", "double"}:
                normalized["value_type"] = "float"
            elif item_type in {"int", "integer"}:
                normalized["value_type"] = "int"
            elif item_type in {"bool", "boolean"}:
                normalized["value_type"] = "bool"
            elif item_type:
                normalized["value_type"] = "string"

    if normalized_type == "dict" and "value_type" not in normalized:
        additional = normalized.get("additionalProperties")
        if isinstance(additional, dict):
            value_type = str(additional.get("type") or "").lower()
            if value_type in {"float", "number", "double"}:
                normalized["value_type"] = "float"
            elif value_type in {"int", "integer"}:
                normalized["value_type"] = "int"
            elif value_type in {"bool", "boolean"}:
                normalized["value_type"] = "bool"
            elif value_type:
                normalized["value_type"] = "string"

    return normalized


def _coerce_value_for_type(value: Any, option: dict[str, Any]) -> Any:
    """Return ``value`` coerced according to ``option`` definition."""

    opt_type = str(option.get("type") or "").lower()

    if opt_type == "bool":
        if isinstance(value, str):
            return value.strip().lower() in {"1", "true", "yes", "y", "on"}
        if value is None:
            return False
        return bool(value)

    if opt_type == "int":
        try:
            return int(value)
        except (TypeError, ValueError):
            return 0

    if opt_type == "float":
        try:
            return float(value)
        except (TypeError, ValueError):
            return 0.0

    if opt_type == "enum":
        allowed = [str(item) for item in option.get("enum", [])]
        if not allowed:
            return "" if value is None else str(value)
        current = "" if value is None else str(value)
        if current not in allowed:
            return allowed[0]
        return current

    return value


def _create_widget(
    option: dict[str, Any], parent: tk.Widget
) -> tuple[ttk.Frame, tk.Variable]:
    """Return a frame containing label, widget and description for the option."""

    option = _normalize_field_definition(option)
    option = dict(option)
    option["default"] = _coerce_value_for_type(option.get("default"), option)

    ui_meta = option.get("ui") or {}
    column_weight = ui_meta.get("column_weight")
    fixed_width = bool(ui_meta.get("fixed_width"))
    if column_weight is None:
        column_weight = 0 if fixed_width else 1

    frame = ttk.Frame(parent)
    label_widget = ttk.Label(
        frame,
        text=option.get("label_pl")
        or option.get("label")
        or option.get("title")
        or option["key"],
    )
    label_widget.grid(row=0, column=0, sticky="nw", padx=5, pady=(5, 0))

    opt_type = str(option.get("type") or "").lower()
    widget_type = option.get("widget")
    default = option.get("default")
    unit_label = option.get("unit") or ui_meta.get("unit")

    if opt_type == "bool":
        var = tk.BooleanVar(value=bool(default))
        widget = ttk.Checkbutton(frame, variable=var)
    elif opt_type in {"int", "float"}:
        if opt_type == "int":
            try:
                value = int(default)
            except (TypeError, ValueError):
                value = 0
            var = tk.IntVar(value=value)
        else:
            try:
                value = float(default)
            except (TypeError, ValueError):
                value = 0.0
            var = tk.DoubleVar(value=value)
        spin_args: dict[str, Any] = {}
        if "min" in option:
            spin_args["from_"] = option["min"]
        if "max" in option:
            spin_args["to"] = option["max"]
        widget = ttk.Spinbox(frame, textvariable=var, **spin_args)
    elif opt_type == "enum":
        allowed = [str(item) for item in option.get("enum", [])]
        labels_map = option.get("enum_labels", {}) or {}
        if allowed and str(default) not in allowed:
            default = allowed[0]
        var = tk.StringVar(value=str(default) if default is not None else "")
        display_values = [labels_map.get(val, val) for val in allowed]
        display_var = tk.StringVar()

        def _sync_display(*_args: Any) -> None:
            value = var.get()
            display_var.set(labels_map.get(value, value))

        _sync_display()

        widget = ttk.Combobox(
            frame,
            textvariable=display_var,
            values=display_values or allowed,
            state="readonly" if (display_values or allowed) else "normal",
        )

        def _on_select(event=None) -> None:  # noqa: ARG001
            chosen = display_var.get()
            for value in allowed:
                label = labels_map.get(value, value)
                if label == chosen:
                    var.set(value)
                    break
            else:
                var.set(chosen)

        widget.bind("<<ComboboxSelected>>", _on_select)
        var.trace_add("write", _sync_display)
    elif opt_type == "path":
        var = tk.StringVar(value=default or "")
        sub = ttk.Frame(frame)
        entry = ttk.Entry(sub, textvariable=var)
        entry.pack(side="left", fill="x", expand=True)

        def browse() -> None:
            owner = frame.winfo_toplevel() if hasattr(frame, "winfo_toplevel") else None
            if widget_type == "dir":
                with TopMost(owner, grab=False):
                    try:
                        _ensure_topmost(owner)
                    except Exception:
                        pass
                    path = filedialog.askdirectory(parent=owner)
            else:
                reason = f"option:{option.get('key', 'path')}"
                path = _safe_pick_json(
                    owner,
                    reason,
                    title=option.get("dialog_title") or "Wybierz plik",
                )
            if path:
                var.set(path)

        ttk.Button(sub, text="Przeglądaj", command=browse).pack(
            side="left", padx=2
        )
        widget = sub
    elif opt_type == "array":
        default_list = option.get("default", []) or []
        lines = "\n".join(str(x) for x in default_list)
        item_type = option.get("value_type")
        if item_type in {"float", "int"} or (
            default_list and all(isinstance(x, (int, float)) for x in default_list)
        ):
            var: tk.StringVar = FloatListVar(value=lines)
        else:
            var = StrListVar(value=lines)
        min_rows = int(ui_meta.get("min_rows", 3))
        max_rows = int(ui_meta.get("max_rows", max(min_rows, 6)))
        if max_rows < min_rows:
            max_rows = min_rows
        initial_rows = max(len(default_list), len(lines.splitlines()), 1)
        dynamic_height = min(max(initial_rows, min_rows), max_rows)
        text = tk.Text(frame, height=dynamic_height, wrap=ui_meta.get("wrap", "word"))
        text.insert("1.0", lines)

        def update_var(*_args: Any) -> None:
            var.set(text.get("1.0", "end").strip())

        text.bind("<KeyRelease>", update_var)
        widget = text
    elif opt_type in {"dict", "object"}:
        default_dict: Dict[str, Any] = option.get("default", {}) or {}
        lines = "\n".join(f"{k} = {v}" for k, v in default_dict.items())
        if option.get("value_type") == "float":
            var = FloatDictVar(value=lines)
        else:
            var = StrDictVar(value=lines)
        min_rows = int(ui_meta.get("min_rows", 3))
        max_rows = int(ui_meta.get("max_rows", max(min_rows, 6)))
        if max_rows < min_rows:
            max_rows = min_rows
        initial_rows = max(len(lines.splitlines()), 1)
        dynamic_height = min(max(initial_rows, min_rows), max_rows)
        text = tk.Text(frame, height=dynamic_height, wrap=ui_meta.get("wrap", "word"))
        text.insert("1.0", lines)

        def update_dict(*_args: Any) -> None:
            var.set(text.get("1.0", "end").strip())

        text.bind("<KeyRelease>", update_dict)
        widget = text
    elif opt_type == "string" and widget_type == "color":
        var = tk.StringVar(value=default or "")
        sub = ttk.Frame(frame)
        entry = ttk.Entry(sub, textvariable=var, width=10)
        entry.pack(side="left", fill="x", expand=True)

        def pick_color() -> None:
            color = colorchooser.askcolor(var.get())[1]
            if color:
                var.set(color)

        ttk.Button(sub, text="Kolor", command=pick_color).pack(
            side="left", padx=2
        )
        widget = sub
    elif opt_type == "string" and widget_type == "password":
        var = tk.StringVar(value=default or "")
        widget = ttk.Entry(frame, textvariable=var, show="*")
    else:
        var = tk.StringVar(value=default or "")
        widget = ttk.Entry(frame, textvariable=var)

    width = ui_meta.get("width")
    try:
        if width and hasattr(widget, "configure"):
            widget.configure(width=width)
    except Exception:
        logger.debug(
            "[SETTINGS] Nie można ustawić width=%s dla %s", width, option.get("key")
        )

    sticky = ui_meta.get("sticky")
    if not sticky:
        if opt_type == "bool":
            sticky = "w"
        else:
            sticky = "ew" if column_weight else "w"

    widget.grid(row=0, column=1, sticky=sticky, padx=5, pady=(5, 0))

    if unit_label:
        ttk.Label(frame, text=str(unit_label)).grid(
            row=0, column=2, sticky="w", padx=(0, 5), pady=(5, 0)
        )

    tip = option.get("help_pl") or option.get("help") or ""
    if tip:
        _bind_tooltip(widget, tip)
        _bind_tooltip(label_widget, tip)

    desc = option.get("help_pl") or option.get("help") or option.get("description")
    if desc:
        ttk.Label(frame, text=desc, font=("", 8)).grid(
            row=1, column=0, columnspan=3, sticky="w", padx=5, pady=(0, 5)
        )

    frame.columnconfigure(1, weight=column_weight)
    frame.columnconfigure(2, weight=0)
    return frame, var


class FloatListVar(tk.StringVar):
    """StringVar that parses lines into a list of floats."""

    def get(self) -> list[float]:  # type: ignore[override]
        vals: list[float] = []
        for line in super().get().splitlines():
            line = line.strip()
            if not line:
                continue
            try:
                vals.append(float(line))
            except ValueError:
                continue
        return vals


class FloatDictVar(tk.StringVar):
    """StringVar that parses "key = value" lines into a float dictionary."""

    def get(self) -> Dict[str, float]:  # type: ignore[override]
        result: Dict[str, float] = {}
        for line in super().get().splitlines():
            if "=" not in line:
                continue
            key, val = line.split("=", 1)
            key = key.strip()
            val = val.strip()
            if not key:
                continue
            try:
                result[key] = float(val)
            except ValueError:
                continue
        return result


class StrListVar(tk.StringVar):
    """StringVar that returns non-empty lines as a list of strings."""

    def get(self) -> list[str]:  # type: ignore[override]
        return [ln.strip() for ln in super().get().splitlines() if ln.strip()]


class StrDictVar(tk.StringVar):
    """StringVar that parses "key = value" lines into a string dictionary."""

    def get(self) -> Dict[str, str]:  # type: ignore[override]
        result: Dict[str, str] = {}
        for line in super().get().splitlines():
            if "=" not in line:
                continue
            key, val = line.split("=", 1)
            key = key.strip()
            val = val.strip()
            if key and val:
                result[key] = val
        return result


class CSVListVar(tk.StringVar):
    """StringVar that normalizes comma-separated values into a list."""

    def _split(self, raw: str) -> list[str]:
        parts: list[str] = []
        for chunk in raw.replace(";", ",").replace("\n", ",").split(","):
            text = chunk.strip()
            if text:
                parts.append(text)
        return parts

    def get(self) -> list[str]:  # type: ignore[override]
        return self._split(super().get())

    def set(self, value: Any) -> None:  # type: ignore[override]
        if isinstance(value, str):
            tokens = self._split(value)
        elif isinstance(value, (list, tuple, set)):
            tokens = [str(item).strip() for item in value if str(item).strip()]
        else:
            tokens = [str(value).strip()] if value is not None else []
        super().set(", ".join(tokens))


class NestedListVar(tk.StringVar):
    """StringVar that parses "typ: status1, status2" lines into dict list."""

    def get(self) -> dict[str, list[str]]:  # type: ignore[override]
        result: dict[str, list[str]] = {}
        for raw_line in super().get().splitlines():
            line = raw_line.strip()
            if not line or ":" not in line:
                continue
            key, values = line.split(":", 1)
            key = key.strip()
            if not key:
                continue
            tokens = [
                token.strip()
                for token in re.split(r"[;,]", values)
                if token.strip()
            ]
            result[key] = tokens
        return result

    def set(self, value: Any) -> None:  # type: ignore[override]
        if isinstance(value, dict):
            lines: list[str] = []
            for key, items in value.items():
                key_str = str(key).strip()
                if not key_str:
                    continue
                joined = ", ".join(
                    str(item).strip()
                    for item in items
                    if str(item).strip()
                )
                lines.append(f"{key_str}: {joined}" if joined else key_str)
            super().set("\n".join(lines))
        else:
            super().set("" if value is None else str(value))


def _bind_tooltip(widget, text: str):
    import tkinter as tk

    tip = {"w": None}

    def _show(_=None):
        if tip["w"] or not text:
            return
        x = widget.winfo_rootx() + 16
        y = widget.winfo_rooty() + 20
        tw = tk.Toplevel(widget)
        ensure_theme_applied(tw)
        _ensure_topmost(tw, widget)
        tw.wm_overrideredirect(True)
        tw.wm_geometry(f"+{x}+{y}")
        lbl = tk.Label(
            tw,
            text=text,
            bg=get_theme_color("card", fallback="#2A2F37"),
            fg=get_theme_color("fg", fallback="#E8E8E8"),
            bd=1,
            relief="solid",
            justify="left",
        )
        lbl.pack(ipadx=8, ipady=6)
        tip["w"] = tw

    def _hide(_=None):
        if tip["w"]:
            tip["w"].destroy()
            tip["w"] = None

    widget.bind("<Enter>", _show, add="+")
    widget.bind("<Leave>", _hide, add="+")


def save_all(options: Dict[str, tk.Variable], cfg: ConfigManager | None = None) -> None:
    """Persist all options from mapping using ConfigManager."""

    cfg = cfg or ConfigManager()
    for key, var in options.items():
        value = var.get()
        cfg.set(key, value)
    cfg.save_all()




class SettingsPanel:
    """Dynamic panel generated from :class:`ConfigManager` schema."""

    # >>> WM PATCH START: SettingsPanel schema getter
    def _get_schema(self):
        schema = None
        if hasattr(self, "cfg"):
            schema = getattr(self.cfg, "schema", None)
            if schema:
                return schema

        master_schema = getattr(self.master, "schema", None)
        if master_schema:
            return master_schema

        parent = getattr(self.master, "master", None)
        if parent is not None:
            parent_schema = getattr(parent, "schema", None)
            if parent_schema:
                return parent_schema

        schema_path = getattr(self, "schema_path", None)
        if schema_path:
            try:
                with open(schema_path, "r", encoding="utf-8") as handle:
                    loaded = json.load(handle)
                if isinstance(loaded, dict) and loaded:
                    return loaded
            except Exception:
                logger.exception("[SETTINGS] Nie udało się wczytać schematu z pliku %s", schema_path)

        return schema
    # >>> WM PATCH END

    def __init__(
        self,
        master: tk.Misc,
        config_path: str | None = None,
        schema_path: str | None = None,
    ):
        self.master = master
        self.config_path = config_path
        self.schema_path = schema_path
        if config_path is not None or schema_path is not None:
            self.cfg = ConfigManager.refresh(
                config_path=config_path, schema_path=schema_path
            )
        else:
            self.cfg = ConfigManager()
        self._dirty = False
        self._unsaved = False
        self._validation_errors: list[str] = []
        self._validate_loaded_config()
        self.settings_state = self._load_settings_state()
        wm_bind_settings_getter(lambda k: self.settings_state.get(k))
        bind_settings(self.settings_state)
        ensure_core_tree()
        settings_actions_bind(self.settings_state, on_change=self.on_setting_changed)
        self.vars: Dict[str, tk.Variable] = {}
        self._initial: Dict[str, Any] = {}
        self._defaults: Dict[str, Any] = {}
        self._options: Dict[str, dict[str, Any]] = {}
        self._fields_vars: list[tuple[tk.Variable, dict[str, Any]]] = []
        self._orders_vars: dict[str, tk.Variable] = {}
        self._orders_meta: dict[str, dict[str, Any]] = {}
        self._open_windows: dict[str, tk.Toplevel] = {}
        self._mod_vars: dict[str, tk.BooleanVar] = {}
        self._user_var: tk.StringVar | None = None
        self._tab_frames: dict[tuple[str, ...], tk.Widget] = {}

        self._container = ttk.Frame(self.master)
        self._container.pack(fill="both", expand=True)
        self._container.grid_rowconfigure(0, weight=1)
        self._container.grid_columnconfigure(0, weight=1)

        self._scroll_area = ScrollableFrame(self._container)
        self._scroll_area.grid(row=0, column=0, sticky="nsew", padx=8, pady=(8, 0))
        self._content_area = self._scroll_area.inner

        self._footer_frame = ttk.Frame(self._container)
        self._footer_frame.grid(row=1, column=0, sticky="ew", padx=8, pady=8)
        self._footer_frame.grid_columnconfigure(0, weight=1)

        self.btns = ttk.Frame(self._footer_frame)
        self.btns.grid(row=0, column=0, sticky="ew")

        try:
            self._base_title = self.master.winfo_toplevel().title() or "Ustawienia"
        except Exception:
            self._base_title = "Ustawienia"

        self._build_ui()

        self._dirty = False
        self._saving = False
        self._autosave_job: str | None = None
        self._autosave_delay_ms = self._resolve_autosave_delay()
        self._bind_defaults_shortcut()
        self._start_autosave_loop()

    def _validate_loaded_config(self) -> None:
        """Validate loaded configuration against the schema and collect issues."""

        try:
            validate = getattr(self.cfg, "_validate_all", None)
            if callable(validate):
                validate()
        except ConfigError as exc:
            message = str(exc)
            self._validation_errors.append(message)
            logger.warning("[SETTINGS] Błąd walidacji konfiguracji: %s", message)
            messagebox.showwarning(
                "Walidacja konfiguracji",
                "Niektóre wartości w config.json wymagają uwagi:\n" + message,
            )
        except Exception:
            logger.exception("[SETTINGS] Nieoczekiwany błąd walidacji konfiguracji")

    def _current_user_role(self) -> str:
        try:
            profile = profile_service.get_active_profile()
        except Exception:
            profile = None
        if isinstance(profile, dict):
            role = profile.get("rola") or profile.get("role")
            if role:
                return str(role).strip().lower()

        for owner in (self.master, getattr(self.master, "master", None)):
            if owner is None:
                continue
            if isinstance(getattr(owner, "active_profile", None), dict):
                ap = getattr(owner, "active_profile")
                role = ap.get("rola") or ap.get("role")
                if role:
                    return str(role).strip().lower()
            for attr in ("rola", "role"):
                role = getattr(owner, attr, None)
                if role:
                    return str(role).strip().lower()

        return ""

    # ------------------------------------------------------------------
    def _load_settings_state(self) -> dict[str, Any]:
        """Return mapping of config keys to their current values."""

        try:
            cfg = getattr(self, "cfg", None)
            if cfg is None:
                return {}

            schema = self._get_schema() or {}
            state: dict[str, Any] = {}

            def _iter_fields(node: dict[str, Any]):
                for field in node.get("fields", []):
                    if field.get("deprecated"):
                        continue
                    key = field.get("key")
                    if key:
                        yield key, field
                for child_key in ("tabs", "groups", "subtabs"):
                    for child in node.get(child_key, []):
                        yield from _iter_fields(child)

            for key, field in _iter_fields(schema):
                state[key] = cfg.get(key, field.get("default"))

            if "paths.data_root" not in state:
                try:
                    state["paths.data_root"] = cfg.get("paths.data_root")
                except Exception:
                    state["paths.data_root"] = None

            for option in schema.get("options", []):
                if option.get("deprecated"):
                    continue
                key = option.get("key")
                if key:
                    state[key] = cfg.get(key, option.get("default"))

            return state
        except Exception:
            return {}

    # ------------------------------------------------------------------
    def _build_ui(self) -> None:
        """Create notebook tabs and widgets based on current schema."""

        self._unsaved = False
        self._fields_vars = []
        self.settings_state.clear()
        self._tab_frames = {}
        content_parent = getattr(self, "_content_area", self.master)
        for child in content_parent.winfo_children():
            child.destroy()
        if hasattr(self, "btns"):
            for child in self.btns.winfo_children():
                child.destroy()

        schema = self._get_schema()
        print(f"[WM-DBG] using schema via _get_schema(): {schema is not None}")
        schema = schema or {}

        base_dir = Path(__file__).resolve().parent
        self._base_dir = base_dir

        self.nb = ttk.Notebook(content_parent)
        self.notebook = self.nb
        print("[WM-DBG] [SETTINGS] notebook created")
        self.nb.pack(fill="both", expand=True)
        self.nb.bind("<<NotebookTabChanged>>", self._on_tab_change)

        self.tab_ogolne = ttk.Frame(self.nb)
        self.tab_ui = ttk.Frame(self.nb)
        self.tab_paths = ttk.Frame(self.nb)
        self.tab_modules = ttk.Frame(self.nb)
        self.tab_users = ttk.Frame(self.nb)
        self.tab_tools = ttk.Frame(self.nb)
        self.tab_backup = ttk.Frame(self.nb)
        self.tab_jarvis = ttk.Frame(self.nb)
        self.tab_advanced = ttk.Frame(self.nb)

        try:
            current_role = self._current_user_role()
        except Exception:
            current_role = ""
        current_role = str(current_role or "").strip().lower()

        allow_users = (
            not current_role
            or current_role in {"administrator", "admin", "brygadzista"}
        )

        tabs_config = [
            (self.tab_ogolne, "Ogólne", ""),
            (self.tab_ui, "Wygląd", ""),
            (self.tab_paths, "Ścieżki", ""),
            (self.tab_tools, "Narzędzia", ""),
            (self.tab_modules, "Moduły", ""),
            (self.tab_backup, "Backup", ""),
            (self.tab_jarvis, "Jarvis", ""),
            (self.tab_advanced, "Zaawansowane", ""),
        ]
        if allow_users:
            tabs_config.insert(3, (self.tab_users, "Użytkownicy", ""))

        for frame, title, subtitle in tabs_config:
            if frame is not None:
                text = title if not subtitle else f"{title}  {subtitle}"
                self.nb.add(frame, text=text)

        _scroll__general_container = ScrolledFrame(self.tab_ogolne)
        _scroll__general_container.pack(fill="both", expand=True, padx=8, pady=8)
        self._general_container = ttk.LabelFrame(
            _scroll__general_container.inner,
            text="🌍 Ogólne"
        )
        self._general_container.pack(fill="both", expand=True, padx=8, pady=8)

        _scroll__users_container = ScrolledFrame(self.tab_users)
        _scroll__users_container.pack(fill="both", expand=True, padx=8, pady=8)
        self._users_container = ttk.Frame(_scroll__users_container.inner)
        self._users_container.pack(fill="both", expand=True, padx=8, pady=8)
        self._users_notebook: ttk.Notebook | None = None
        _scroll__tools_container = ScrolledFrame(self.tab_tools)
        _scroll__tools_container.pack(fill="both", expand=True, padx=8, pady=8)
        self._tools_container = ttk.LabelFrame(
            _scroll__tools_container.inner,
            text="🔧 Narzędzia"
        )
        self._tools_container.pack(fill="both", expand=True, padx=8, pady=8)

        _scroll__paths_container = ScrolledFrame(self.tab_paths)
        _scroll__paths_container.pack(fill="both", expand=True, padx=8, pady=8)
        self._paths_container = ttk.LabelFrame(
            _scroll__paths_container.inner,
            text="📁 Ścieżki i dane"
        )
        self._paths_container.pack(fill="both", expand=True, padx=8, pady=8)

        _scroll__backup_container = ScrolledFrame(self.tab_backup)
        _scroll__backup_container.pack(fill="both", expand=True, padx=8, pady=8)
        self._backup_container = ttk.LabelFrame(
            _scroll__backup_container.inner,
            text="💾 Kopie zapasowe"
        )
        self._backup_container.pack(fill="both", expand=True, padx=8, pady=8)

        _scroll__ui_container = ScrolledFrame(self.tab_ui)
        _scroll__ui_container.pack(fill="both", expand=True, padx=8, pady=8)
        self._ui_container = ttk.LabelFrame(
            _scroll__ui_container.inner,
            text="🎨 Wygląd i UI"
        )
        self._ui_container.pack(fill="both", expand=True, padx=8, pady=8)
        self._modules_nb = ttk.Notebook(self.tab_modules)
        self._modules_nb.pack(fill="both", expand=True, padx=8, pady=8)
        self._modules_nb.bind("<<NotebookTabChanged>>", self._on_modules_tab_change, add="+")
        _scroll__jarvis_container = ScrolledFrame(self.tab_jarvis)
        _scroll__jarvis_container.pack(fill="both", expand=True, padx=8, pady=8)
        self._jarvis_container = ttk.LabelFrame(
            _scroll__jarvis_container.inner,
            text="🔔 Jarvis i alerty"
        )
        self._jarvis_container.pack(fill="both", expand=True, padx=8, pady=8)

        _scroll__advanced_container = ScrolledFrame(self.tab_advanced)
        _scroll__advanced_container.pack(fill="both", expand=True, padx=8, pady=8)
        self._advanced_container = ttk.LabelFrame(
            _scroll__advanced_container.inner,
            text="⚙️ Zaawansowane"
        )
        self._advanced_container.pack(fill="both", expand=True, padx=8, pady=8)

        self._nested_tab_lookup: dict[str, tuple[str, ttk.Notebook, str]] = {}
        self._manual_fields_keys = {
            "paths.data_root",
            "paths.logs_dir",
            "paths.backup_dir",
            "backup.keep_last",
            "backup.auto_on_exit",
            "ui.theme",
            "ui.font_size",
            "ui.block_login_outside_shift",
            "profiles.ui.enable_profile_card",
            "profiles.ui.show_name_in_header",
            "profiles.avatar.enabled",
            "profiles.pin.change_allowed",
            "profiles.editable_fields",
            "jarvis.enabled",
        }
        self._build_manual_config_fields()

        # state for lazy creation of magazyn subtabs
        self._magazyn_frame: ttk.Frame | None = None
        self._magazyn_schema: dict[str, Any] | None = None
        self._magazyn_initialized = False

        handlers = self._get_tab_handlers()

        modules_ids = {
            "maszyny",
            "zlecenia",
            "magazyn",
            "zamowienia",
            "produkty",
            "dyspo",
        }

        # allow_users policzone wyżej – tutaj już tylko używamy tej wartości
        if allow_users:
            users_nb = ttk.Notebook(self._users_container)
            users_nb.pack(fill="both", expand=True, padx=4, pady=4)
            self._users_notebook = users_nb

            users_list_frame = ttk.Frame(users_nb, style="WM.TFrame")
            users_nb.add(users_list_frame, text="Lista i edycja")

            users_profile_frame = ttk.Frame(users_nb, style="WM.TFrame")
            users_nb.add(users_profile_frame, text="Profil użytkownika")

        for tab in schema.get("tabs", []):
            tab_id_raw = tab.get("id")
            tab_id = str(tab_id_raw or "").strip().lower()
            title = tab.get("title", tab.get("id", ""))
            print("[WM-DBG] [SETTINGS] add tab:", title)
            if tab_id == "uzytkownicy":
                if not allow_users:
                    print(
                        "[WM-DBG][SETTINGS] pomijam zakładkę Użytkownicy dla roli:",
                        current_role,
                    )
                # TODO: przenieść do nowego układu zakładek – sekcja użytkowników
                continue
            if tab_id == "system":
                self._render_system_tab(tab, handlers)
                continue
            if tab_id == "narzedzia":
                frame = self._tools_container
                for child in frame.winfo_children():
                    child.destroy()
                path_key = (tab_id,)
                self._remember_tab_frame(path_key, frame)
                counts = self._handle_tools_tab(frame, tab, path_key)
                if counts:
                    self._log_tab_stats(title, *counts)
                continue
            if tab_id in modules_ids:
                frame = ttk.Frame(self._modules_nb)
                self._modules_nb.add(frame, text=title)
                self._register_nested_tab(title, self.tab_modules, self._modules_nb, frame)
                path_key = tuple(filter(None, [tab_id]))
                self._remember_tab_frame(path_key, frame)
                handler = handlers.get(tab_id)
                if handler is None:
                    counts = self._handle_generic_tab(frame, tab, path_key)
                else:
                    counts = handler(frame, tab, path_key)
                if counts:
                    self._log_tab_stats(title, *counts)
                continue
            if tab_id == "aktualizacje":
                frame = ttk.LabelFrame(self._backup_container, text=title)
                frame.pack(fill="both", expand=True, padx=8, pady=6)
                path_key = (tab_id,)
                self._remember_tab_frame(path_key, frame)
                counts = self._handle_generic_tab(frame, tab, path_key)
                if counts:
                    self._log_tab_stats(title, *counts)
                continue
            if tab_id == "testy_audyt":
                frame = ttk.LabelFrame(self._advanced_container, text=title)
                frame.pack(fill="both", expand=True, padx=8, pady=6)
                path_key = (tab_id,)
                self._remember_tab_frame(path_key, frame)
                counts = self._handle_generic_tab(frame, tab, path_key)
                if counts:
                    self._log_tab_stats(title, *counts)
                continue

            frame = ttk.LabelFrame(self._advanced_container, text=title)
            frame.pack(fill="both", expand=True, padx=8, pady=6)
            path_key = tuple(filter(None, [tab_id]))
            self._remember_tab_frame(path_key, frame)
            handler = handlers.get(tab_id)
            if handler is None:
                counts = self._handle_generic_tab(frame, tab, path_key)
            else:
                counts = handler(frame, tab, path_key)
            if counts:
                self._log_tab_stats(title, *counts)

        if allow_users:
            try:
                # FIX: create_users_tab działa jak samodzielny panel,
                # więc osadzamy go w czystym kontenerze.
                for child in users_list_frame.winfo_children():
                    child.destroy()

                users_frame = create_users_tab(users_list_frame)
                if users_frame is not None:
                    users_frame.pack(fill="both", expand=True)
                users_list_frame.update_idletasks()
                self._users_container.update_idletasks()
                self._register_nested_tab(
                    "Użytkownicy",
                    self.tab_users,
                    self._users_notebook,
                    users_list_frame,
                )
            except Exception:
                logger.exception(
                    "[SETTINGS] Nie udało się zbudować zakładki Użytkownicy"
                )
            try:
                self._build_user_profile_settings(users_profile_frame)
            except Exception:
                logger.exception(
                    "[SETTINGS] Nie udało się zbudować sekcji Profil użytkownika"
                )

        try:
            self._add_patch_manager_tab(str(base_dir))
        except Exception as exc:  # pragma: no cover - diagnostyka GUI
            print("[WM-DBG] [PATCH] pominięto sekcję Patche:", exc)

        if self._magazyn_frame is None:
            self._add_magazyn_tab()

        print("[WM-DBG] [SETTINGS] notebook packed")
        self.notebook.select(self.tab_ogolne)

        self._bind_global_shortcuts()

        left_btns = ttk.Frame(self.btns)
        left_btns.pack(side="left", padx=5)
        ttk.Button(left_btns, text="Anuluj", command=self._close_window).pack(
            side="left", padx=5
        )
        self.btn_defaults: ttk.Button | None = None
        ttk.Button(left_btns, text="Odśwież", command=self.refresh_panel).pack(
            side="left", padx=5
        )

        self.btn_save: ttk.Button | None = None

        self.master.winfo_toplevel().protocol("WM_DELETE_WINDOW", self.on_close)

    def _remember_tab_frame(
        self, path: tuple[str, ...], frame: tk.Widget
    ) -> None:
        """Store reference to tab frame for later lookup."""

        if path:
            self._tab_frames[path] = frame

    def _log_tab_stats(self, title: str, groups: int, fields: int) -> None:
        """Log diagnostic information about generated tab contents."""

        print(f"[WM-DBG] tab='{title}' groups={groups} fields={fields}")

    def _register_manual_var(
        self,
        key: str,
        var: tk.Variable,
        *,
        default: Any = None,
        option_type: str | None = None,
    ) -> None:
        """Bind ``var`` to configuration bookkeeping just like schema widgets."""

        if option_type:
            self._options[key] = {"key": key, "type": option_type}
        else:
            self._options.setdefault(key, {"key": key})
        if default is None:
            default = var.get()
        self.vars[key] = var
        self._initial[key] = var.get()
        self._defaults[key] = default
        self._fields_vars.append((var, dict(self._options[key])))
        self.settings_state[key] = var.get()
        var.trace_add("write", lambda *_: self._on_var_write(key, var))

    def _register_nested_tab(
        self,
        name: str,
        top_level: tk.Widget,
        notebook: ttk.Notebook,
        tab_widget: tk.Widget,
    ) -> None:
        """Remember how to focus nested notebook tabs by display name."""

        lookup = getattr(self, "_nested_tab_lookup", None)
        if lookup is None:
            lookup = {}
            self._nested_tab_lookup = lookup
        lookup[name.strip().lower()] = (
            str(top_level),
            notebook,
            str(tab_widget),
        )

    def _choose_directory(
        self, variable: tk.StringVar, *, title: str = "Wybierz katalog"
    ) -> None:
        """Open a folder picker and assign the chosen path to ``variable``."""

        selected = filedialog.askdirectory(title=title)
        if not selected:
            return
        if re.match(r"^[A-Za-z]:\\?$", str(selected).strip()):
            messagebox.showerror(
                "Nieprawidłowa ścieżka",
                "Podaj pełny katalog, np. C\\\\WM\\\\backup.",
                parent=self.master,
            )
            return
        try:
            normalized = os.path.normpath(selected)
        except Exception:
            normalized = selected
        variable.set(normalized)

    def _get_tab_handlers(self) -> dict[str, TabHandler]:
        """Return mapping of tab identifiers to specialized handlers."""

        return {
            "magazyn": self._handle_magazyn_tab,
            "produkty": self._handle_products_tab,
            "zlecenia": self._handle_orders_tab,
            "narzedzia": self._handle_tools_tab,
            "maszyny": self._handle_machines_tab,
            "system": self._handle_system_tab,
        }

    def _render_system_tab(
        self, tab: dict[str, Any], handlers: dict[str, TabHandler]
    ) -> None:
        """Distribute legacy system subtabs between the new layout sections."""

        base_path = ("system",)
        root_holder = ttk.Frame(self._paths_container)
        root_holder.pack(fill="x", expand=False, padx=8, pady=(4, 0))
        self._remember_tab_frame(base_path + ("root",), root_holder)
        # TODO: przenieść do nowego układu zakładek – sekcja wyboru Folderu WM
        _make_system_tab(
            root_holder,
            globals().get("CONFIG_MANAGER"),
            owner=self,
            on_root_change=self._apply_root_change,
        )

        for subtab in tab.get("subtabs", []):
            sub_id_raw = subtab.get("id")
            sub_id = str(sub_id_raw or "").strip().lower()
            title = subtab.get("title", subtab.get("id", ""))
            if sub_id == "sciezki_danych":
                parent = self._paths_container
            elif sub_id == "ustawienia_startowe":
                parent = self._general_container
            elif sub_id == "interfejs":
                parent = self._ui_container
            elif sub_id in {"backup", "kopie_zapasowe", "aktualizacje"}:
                parent = self._backup_container
            else:
                parent = self._advanced_container
            frame = ttk.Frame(parent)
            frame.pack(fill="both", expand=True, padx=8, pady=6)
            path_key = base_path + (sub_id,) if sub_id else base_path
            handler = handlers.get(sub_id)
            if handler is None:
                counts = self._handle_generic_tab(frame, subtab, path_key)
            else:
                counts = handler(frame, subtab, path_key)
            if counts:
                self._log_tab_stats(title, *counts)

    def _build_manual_config_fields(self) -> None:
        """Insert hand-crafted widgets tied to configuration keys."""

        cfg = getattr(self, "cfg", None) or ConfigManager()

        security_frame = ttk.LabelFrame(
            self._general_container, text="Bezpieczeństwo"
        )
        security_frame.pack(fill="x", expand=False, padx=8, pady=(0, 8))

        self.var_block_login_outside_shift = tk.BooleanVar(
            value=cfg.get("ui.block_login_outside_shift", False)
        )
        chk_block_outside_shift = ttk.Checkbutton(
            security_frame,
            text="Blokuj logowanie poza swoją zmianą",
            variable=self.var_block_login_outside_shift,
        )
        chk_block_outside_shift.pack(anchor="w", padx=4, pady=4)
        self._register_manual_var(
            "ui.block_login_outside_shift",
            self.var_block_login_outside_shift,
            default=self.var_block_login_outside_shift.get(),
            option_type="bool",
        )

        ttk.Label(
            security_frame, text="Data startu rotacji (YYYY-MM-DD):"
        ).pack(anchor="w", padx=4, pady=(8, 2))
        self.var_attendance_rotation_start = tk.StringVar(
            value=str(cfg.get("attendance.rotation_start", "") or "")
        )
        ttk.Entry(
            security_frame, textvariable=self.var_attendance_rotation_start, width=20
        ).pack(anchor="w", padx=4, pady=(0, 4))
        self._register_manual_var(
            "attendance.rotation_start",
            self.var_attendance_rotation_start,
            default=self.var_attendance_rotation_start.get(),
            option_type="str",
        )

        paths_frame = ttk.LabelFrame(
            self._paths_container, text="Ścieżki konfiguracyjne"
        )
        paths_frame.pack(fill="x", expand=False, padx=8, pady=(0, 8))
        paths_frame.columnconfigure(1, weight=1)

        ttk.Label(paths_frame, text="Folder danych (root):").grid(
            row=0, column=0, sticky="w", padx=4, pady=4
        )
        # Powiązane z config["paths"]["data_root"]
        self.var_data_root = tk.StringVar(value=cfg.get("paths.data_root", ""))
        entry_data_root = ttk.Entry(
            paths_frame, textvariable=self.var_data_root, state="readonly"
        )
        entry_data_root.grid(row=0, column=1, sticky="ew", padx=4, pady=4)
        self._register_manual_var(
            "paths.data_root",
            self.var_data_root,
            default=self.var_data_root.get(),
            option_type="path",
        )

        ttk.Label(paths_frame, text="Katalog logów:").grid(
            row=1, column=0, sticky="w", padx=4, pady=4
        )
        # Powiązane z config["paths"]["logs_dir"]
        self.var_logs_dir = tk.StringVar(value=cfg.get("paths.logs_dir", ""))
        entry_logs_dir = ttk.Entry(paths_frame, textvariable=self.var_logs_dir)
        entry_logs_dir.grid(row=1, column=1, sticky="ew", padx=4, pady=4)
        ttk.Button(
            paths_frame,
            text="Wybierz…",
            command=lambda: self._choose_directory(
                self.var_logs_dir, title="Wybierz katalog logów"
            ),
        ).grid(row=1, column=2, sticky="ew", padx=4, pady=4)
        self._register_manual_var(
            "paths.logs_dir",
            self.var_logs_dir,
            default=self.var_logs_dir.get(),
            option_type="path",
        )

        ttk.Label(paths_frame, text="Katalog kopii zapasowych:").grid(
            row=2, column=0, sticky="w", padx=4, pady=4
        )
        # Powiązane z config["paths"]["backup_dir"]
        self.var_backups_dir = tk.StringVar(value=cfg.get("paths.backup_dir", ""))
        entry_backups_dir = ttk.Entry(paths_frame, textvariable=self.var_backups_dir)
        entry_backups_dir.grid(row=2, column=1, sticky="ew", padx=4, pady=4)
        ttk.Button(
            paths_frame,
            text="Wybierz…",
            command=lambda: self._choose_directory(
                self.var_backups_dir, title="Wybierz katalog kopii"
            ),
        ).grid(row=2, column=2, sticky="ew", padx=4, pady=4)
        self._register_manual_var(
            "paths.backup_dir",
            self.var_backups_dir,
            default=self.var_backups_dir.get(),
            option_type="path",
        )

        backup_frame = ttk.LabelFrame(
            self._backup_container, text="Parametry kopii zapasowej"
        )
        backup_frame.pack(fill="x", expand=False, padx=8, pady=(0, 8))
        backup_frame.columnconfigure(1, weight=1)

        ttk.Label(backup_frame, text="Ilość zachowanych kopii:").grid(
            row=0, column=0, sticky="w", padx=4, pady=4
        )
        # Powiązane z config["backup"]["keep_last"]
        self.var_backup_keep_last = tk.IntVar(
            value=cfg.get("backup.keep_last", 5)
        )
        spin_keep_last = ttk.Spinbox(
            backup_frame,
            from_=1,
            to=50,
            textvariable=self.var_backup_keep_last,
            width=6,
        )
        spin_keep_last.grid(row=0, column=1, sticky="w", padx=4, pady=4)
        self._register_manual_var(
            "backup.keep_last",
            self.var_backup_keep_last,
            default=self.var_backup_keep_last.get(),
            option_type="int",
        )

        # Powiązane z config["backup"]["auto_on_exit"]
        self.var_backup_on_exit = tk.BooleanVar(
            value=cfg.get("backup.auto_on_exit", True)
        )
        chk_backup_on_exit = ttk.Checkbutton(
            backup_frame,
            text="Rób kopię przy zamykaniu programu",
            variable=self.var_backup_on_exit,
        )
        chk_backup_on_exit.grid(row=1, column=0, columnspan=2, sticky="w", padx=4, pady=4)
        self._register_manual_var(
            "backup.auto_on_exit",
            self.var_backup_on_exit,
            default=self.var_backup_on_exit.get(),
            option_type="bool",
        )

        ui_frame = ttk.LabelFrame(self._ui_container, text="Wygląd interfejsu")
        ui_frame.pack(fill="x", expand=False, padx=8, pady=(0, 8))
        ui_frame.columnconfigure(1, weight=1)

        ttk.Label(ui_frame, text="Motyw interfejsu:").grid(
            row=0, column=0, sticky="w", padx=4, pady=4
        )
        # Powiązane z config["ui"]["theme"]
        self.var_theme = tk.StringVar(value=cfg.get("ui.theme", "dark"))
        combo_theme = ttk.Combobox(
            ui_frame,
            textvariable=self.var_theme,
            values=["dark", "light", "auto"],
            state="readonly",
        )
        combo_theme.grid(row=0, column=1, sticky="ew", padx=4, pady=4)
        self._register_manual_var(
            "ui.theme",
            self.var_theme,
            default=self.var_theme.get(),
            option_type="string",
        )

        ttk.Label(ui_frame, text="Rozmiar czcionki:").grid(
            row=1, column=0, sticky="w", padx=4, pady=4
        )
        # Powiązane z config["ui"]["font_size"]
        self.var_font_size = tk.IntVar(value=cfg.get("ui.font_size", 10))
        spin_font_size = ttk.Spinbox(
            ui_frame,
            from_=8,
            to=18,
            textvariable=self.var_font_size,
            width=6,
        )
        spin_font_size.grid(row=1, column=1, sticky="w", padx=4, pady=4)
        self._register_manual_var(
            "ui.font_size",
            self.var_font_size,
            default=self.var_font_size.get(),
            option_type="int",
        )

        jarvis_frame = ttk.LabelFrame(
            self._jarvis_container, text="Jarvis i powiadomienia"
        )
        jarvis_frame.pack(fill="x", expand=False, padx=8, pady=(0, 8))

        # Powiązane z config["jarvis"]["enabled"]
        self.var_jarvis_enabled = tk.BooleanVar(
            value=cfg.get("jarvis.enabled", True)
        )
        chk_jarvis = ttk.Checkbutton(
            jarvis_frame,
            text="Włącz Jarvisa",
            variable=self.var_jarvis_enabled,
        )
        chk_jarvis.pack(anchor="w", padx=4, pady=4)
        self._register_manual_var(
            "jarvis.enabled",
            self.var_jarvis_enabled,
            default=self.var_jarvis_enabled.get(),
            option_type="bool",
        )

        self._last_tab = self.nb.select()

    def _build_user_profile_settings(self, parent: tk.Widget) -> None:
        """Extend the Users tab with global profile settings."""

        cfg = getattr(self, "cfg", None) or ConfigManager()

        holder = ttk.Frame(parent)
        holder.pack(fill="both", expand=True, padx=4, pady=4)

        profile_frame = ttk.LabelFrame(
            holder, text="Ustawienia profilu użytkownika"
        )
        profile_frame.pack(fill="x", expand=False, padx=8, pady=(0, 8), anchor="nw")
        profile_frame.columnconfigure(1, weight=1)

        def _cfg_bool(key: str, fallback_key: str, default: bool) -> bool:
            value = cfg.get(key, None)
            if value is None:
                value = cfg.get(fallback_key, default)
            return bool(value)

        self.var_profile_enabled = tk.BooleanVar(
            value=_cfg_bool(
                "profiles.ui.enable_profile_card",
                "ui.profile.enabled",
                True,
            )
        )
        self.var_profile_header = tk.BooleanVar(
            value=_cfg_bool(
                "profiles.ui.show_name_in_header",
                "ui.profile.show_name_header",
                True,
            )
        )
        self.var_profile_avatar = tk.BooleanVar(
            value=_cfg_bool(
                "profiles.avatar.enabled",
                "ui.profile.avatar_enabled",
                False,
            )
        )
        self.var_profile_pin_change = tk.BooleanVar(
            value=_cfg_bool(
                "profiles.pin.change_allowed",
                "ui.profile.allow_pin_change",
                False,
            )
        )

        ttk.Checkbutton(
            profile_frame,
            text="Włącz kartę profilu",
            variable=self.var_profile_enabled,
        ).grid(row=0, column=0, columnspan=2, sticky="w", padx=4, pady=2)
        ttk.Checkbutton(
            profile_frame,
            text="Pokazuj imię w nagłówku",
            variable=self.var_profile_header,
        ).grid(row=1, column=0, columnspan=2, sticky="w", padx=4, pady=2)
        ttk.Checkbutton(
            profile_frame,
            text="Włącz avatar",
            variable=self.var_profile_avatar,
        ).grid(row=2, column=0, columnspan=2, sticky="w", padx=4, pady=2)
        ttk.Checkbutton(
            profile_frame,
            text="Zezwól na zmianę PIN",
            variable=self.var_profile_pin_change,
        ).grid(row=3, column=0, columnspan=2, sticky="w", padx=4, pady=2)

        raw_fields = []
        try:
            raw_fields = cfg.get("profiles.editable_fields", []) or []
        except Exception:
            raw_fields = []
        if not isinstance(raw_fields, list):
            raw_fields = []
        self.var_profile_editable_fields = tk.StringVar(
            value=", ".join([str(x) for x in raw_fields if str(x).strip()])
        )

        ttk.Label(
            profile_frame,
            text="Pola edytowane przez użytkownika (CSV):",
        ).grid(row=4, column=0, sticky="w", padx=4, pady=(6, 2))
        ttk.Entry(
            profile_frame,
            textvariable=self.var_profile_editable_fields,
        ).grid(row=4, column=1, sticky="ew", padx=4, pady=(6, 2))

        ttk.Label(
            profile_frame,
            text="Np.: imie, nazwisko, telefon, email",
            style="WM.Muted.TLabel",
        ).grid(row=5, column=0, columnspan=2, sticky="w", padx=4, pady=(0, 4))

        self._register_manual_var(
            "profiles.ui.enable_profile_card",
            self.var_profile_enabled,
            default=self.var_profile_enabled.get(),
            option_type="bool",
        )
        self._register_manual_var(
            "profiles.ui.show_name_in_header",
            self.var_profile_header,
            default=self.var_profile_header.get(),
            option_type="bool",
        )
        self._register_manual_var(
            "profiles.avatar.enabled",
            self.var_profile_avatar,
            default=self.var_profile_avatar.get(),
            option_type="bool",
        )
        self._register_manual_var(
            "profiles.pin.change_allowed",
            self.var_profile_pin_change,
            default=self.var_profile_pin_change.get(),
            option_type="bool",
        )
        self._register_manual_var(
            "profiles.editable_fields",
            self.var_profile_editable_fields,
            default=self.var_profile_editable_fields.get(),
            option_type="string",
        )

    # ------------------------------------------------------------------
    def open_tab(self, title: str) -> bool:
        """Activate the notebook tab matching ``title``."""

        try:
            notebook = self.nb
        except AttributeError:
            return False

        # FIX(SETTINGS): notebook mógł zostać już zniszczony, a stary callback
        # nadal próbuje wykonać notebook.tabs()/select(). To kończy się:
        # invalid command name "...!notebook"
        try:
            if notebook is None or not bool(notebook.winfo_exists()):
                return False
        except Exception:
            return False

        wanted = title.strip().lower()
        if not wanted:
            return False

        try:
            tab_ids = notebook.tabs()
        except Exception:
            return False

        for tab_id in tab_ids:
            try:
                text = str(notebook.tab(tab_id, "text") or "")
            except Exception:
                continue
            if text.strip().lower() == wanted:
                try:
                    notebook.select(tab_id)
                except Exception:
                    return False
                try:
                    notebook.focus_set()
                except Exception:
                    pass
                return True
        nested = getattr(self, "_nested_tab_lookup", {}).get(wanted)
        if nested:
            top_id, inner_nb, inner_tab = nested
            try:
                if inner_nb is None or not bool(inner_nb.winfo_exists()):
                    return False
                notebook.select(top_id)
            except Exception:
                return False
            try:
                inner_nb.select(inner_tab)
                inner_nb.focus_set()
            except Exception:
                return False
            return True
        return False

    def _bind_global_shortcuts(self) -> None:
        """Register global keybindings for frequently used tabs."""

        try:
            root = self.master.winfo_toplevel()
        except Exception:
            return

        def _make_handler(tab_title: str):
            def _handler(event):
                if self.open_tab(tab_title):
                    return "break"
                return None

            return _handler

        root.bind_all("<Key-n>", _make_handler("Narzędzia"))
        root.bind_all("<Key-m>", _make_handler("Maszyny"))
        root.bind_all("<Key-p>", _make_handler("Użytkownicy"))

    def _on_modules_tab_change(self, _event=None):
        if self._magazyn_frame is not None and not self._magazyn_initialized:
            if self._modules_nb.select() == str(self._magazyn_frame):
                self._init_magazyn_tab()

    def _handle_magazyn_tab(
        self, frame: tk.Widget, tab: dict[str, Any], tab_path: tuple[str, ...]
    ) -> tuple[int, int] | None:
        """Prepare magazyn tab for lazy initialization."""

        self._magazyn_frame = frame
        self._magazyn_schema = tab
        return None

    def _handle_products_tab(
        self, frame: tk.Widget, tab: dict[str, Any], tab_path: tuple[str, ...]
    ) -> tuple[int, int]:
        """Build products tab content using dedicated helper."""

        return self._build_products_tab(frame, tab)

    def _handle_orders_tab(
        self, frame: tk.Widget, tab: dict[str, Any], tab_path: tuple[str, ...]
    ) -> tuple[int, int]:
        """Build orders tab content using dedicated helper."""

        return self._build_orders_tab(frame, tab)

    def _handle_tools_tab(
        self, frame: tk.Widget, tab: dict[str, Any], tab_path: tuple[str, ...]
    ) -> tuple[int, int]:
        """Build tools tab content using dedicated helper."""

        return self._build_tools_tab(frame, tab)

    def _handle_machines_tab(
        self, frame: tk.Widget, tab: dict[str, Any], tab_path: tuple[str, ...]
    ) -> tuple[int, int]:
        """Build machines tab content and append supplementary sections."""

        grp_count, fld_count = self._populate_tab(frame, tab, tab_path=tab_path)
        _add_machines_bg_group(frame, self.cfg)
        _add_machines_map_group(frame, self.cfg)
        _add_readonly_info(
            frame,
            "Pliki danych są ustalane relatywnie względem Folderu WM (root). "
            "Ścieżka pliku maszyn: 'maszyny/maszyny.json'. Zmień tylko Folder WM "
            "(root) w zakładce System.",
        )
        return grp_count, fld_count

    def _handle_system_tab(
        self, frame: tk.Widget, tab: dict[str, Any], tab_path: tuple[str, ...]
    ) -> tuple[int, int]:
        """Build system tab content with system helpers and schema widgets."""

        _make_system_tab(
            frame,
            globals().get("CONFIG_MANAGER"),
            owner=self,
            on_root_change=self._apply_root_change,
        )
        return self._populate_tab(frame, tab, tab_path=tab_path)

    def _handle_generic_tab(
        self, frame: tk.Widget, tab: dict[str, Any], tab_path: tuple[str, ...]
    ) -> tuple[int, int]:
        """Fallback handler populating tab via schema definition."""

        return self._populate_tab(frame, tab, tab_path=tab_path)

    def _refresh_paths_preview(self) -> None:
        """Refresh diagnostic labels showing resolved configuration paths."""

        labels = getattr(self, "_root_paths_labels", None)
        cfg_manager = getattr(self, "cfg", None)
        if not isinstance(labels, dict) or not labels:
            return
        if cfg_manager is None:
            return

        updates: dict[str, str] = {}
        for key, getter in (
            ("root", lambda: cfg_manager.path_root()),
            ("config", lambda: cfg_manager.get_config_path()),
            ("data", lambda: cfg_manager.path_data()),
            ("backup", lambda: cfg_manager.path_backup()),
            ("logs", lambda: cfg_manager.path_logs()),
            ("assets", lambda: cfg_manager.path_assets()),
        ):
            try:
                value = getter()
                updates[key] = str(value)
            except Exception:
                updates[key] = ""

        root_value = updates.get("root", "")
        warning_keys = {"data", "backup", "logs", "assets"}

        for key, widget in labels.items():
            try:
                text, color = _format_diag_path(
                    updates.get(key, ""),
                    root_value,
                    warn=bool(root_value) and key in warning_keys,
                )
                widget.config(text=text, foreground=color)
            except Exception:
                logger.exception("[SETTINGS] Nie udało się odświeżyć etykiety %s", key)

    def _apply_root_change(self, new_root: str) -> None:
        """Apply runtime updates after changing the <root> directory."""

        raw_root = str(new_root or "").strip()
        if re.match(r"^[A-Za-z]:\\?$", raw_root):
            messagebox.showerror(
                "Nieprawidłowa ścieżka",
                "Podaj pełny katalog, np. C\\\\WM\\\\data.",
                parent=self.master,
            )
            return

        new_root = os.path.normcase(os.path.abspath(raw_root))
        self.cfg.set("paths.data_root", new_root)

        logs_dir = self.cfg.get("paths.logs_dir")
        if not logs_dir:
            logs_dir = os.path.join(new_root, "logs")
            self.cfg.set("paths.logs_dir", logs_dir)

        backup_dir = self.cfg.get("paths.backup_dir")
        if not backup_dir:
            backup_dir = os.path.join(new_root, "backup")
            self.cfg.set("paths.backup_dir", backup_dir)

        assets_dir = self.cfg.get("paths.assets_dir")
        if not assets_dir:
            assets_dir = os.path.join(new_root, "assets")
            self.cfg.set("paths.assets_dir", assets_dir)

        for path in (
            new_root,
            logs_dir,
            backup_dir,
            os.path.join(new_root, "data"),
            assets_dir,
        ):
            try:
                os.makedirs(path, exist_ok=True)
            except Exception:
                logger.exception("[SETTINGS] Tworzenie katalogu %s nieudane", path)

        self.cfg.save()
        init_logging(self.cfg)
        if hasattr(self, "var_data_root"):
            try:
                self.var_data_root.set(new_root)
            except Exception:
                pass
        if hasattr(self, "var_logs_dir") and logs_dir:
            try:
                self.var_logs_dir.set(logs_dir)
            except Exception:
                pass
        if hasattr(self, "var_backups_dir") and backup_dir:
            try:
                self.var_backups_dir.set(backup_dir)
            except Exception:
                pass
        self._refresh_paths_preview()
        print(f"[WM-DBG][SETTINGS] Ustawiono nowe <root>: {new_root}")

    def _close_window(self) -> None:
        """Invoke the standard close flow used by the Cancel button."""

        try:
            self.on_close()
        except Exception:
            try:
                self.master.winfo_toplevel().destroy()
            except Exception:
                pass

    def _add_magazyn_tab(self) -> None:
        try:
            frame = MagazynSettingsFrame(self._modules_nb, self.cfg)
        except Exception as e:
            import tkinter as tk
            from tkinter import ttk
            frame = ttk.Frame(self._modules_nb)
            lbl = ttk.Label(frame, text=f"Błąd ładowania zakładki Magazyn:\n{e}")
            lbl.pack(padx=12, pady=12)
        self._modules_nb.add(frame, text="Magazyn")
        self._register_nested_tab("Magazyn", self.tab_modules, self._modules_nb, frame)

    def _coerce_default_for_var(self, opt: dict[str, Any], default: Any) -> Any:
        """Return value adjusted for Tk variable according to option definition."""

        opt_type = opt.get("type")
        widget_type = opt.get("widget")
        if opt_type == "array":
            default_list = default or []
            return "\n".join(str(x) for x in default_list)
        if opt_type in {"dict", "object"}:
            default_dict: Dict[str, Any] = default or {}
            return "\n".join(f"{k} = {v}" for k, v in default_dict.items())
        if opt_type == "string" and widget_type == "color":
            return default or ""
        if opt_type == "path":
            return default or ""
        return default

    def _register_option_var(
        self,
        key: str,
        var: tk.Variable,
        field_def: dict[str, Any] | None,
    ) -> None:
        """Register Tk variable for config option with bookkeeping."""

        opt = (
            _normalize_field_definition(field_def) if field_def else {"key": key}
        )
        opt = dict(opt)
        self.vars[key] = var
        self._options[key] = opt
        self._initial[key] = var.get()
        self._defaults[key] = opt.get("default")
        self._fields_vars.append((var, opt))
        self.settings_state[key] = var.get()
        var.trace_add("write", lambda *_: self._on_var_write(key, var))

    def _create_button_field(
        self, parent: tk.Widget, field_def: dict[str, Any]
    ) -> ttk.Button:
        """Return ttk button configured for schema action field."""

        text = (
            field_def.get("label_pl")
            or field_def.get("label")
            or field_def.get("key")
            or "Akcja"
        )
        btn = ttk.Button(
            parent,
            text=text,
            command=lambda f=field_def, lbl=text: self._on_button_field_clicked(
                f, lbl
            ),
        )
        tip = field_def.get("help_pl") or field_def.get("help")
        if tip:
            _bind_tooltip(btn, tip)
        return btn

    def _on_var_write(self, key: str, var: tk.Variable) -> None:
        """Handle Tk variable updates by tracking unsaved state and cache."""

        setattr(self, "_unsaved", True)
        try:
            self.settings_state[key] = var.get()
        except Exception:
            pass
        if getattr(self, "_saving", False):
            return
        self._mark_dirty()
        self._status(f"Zmieniono: {key}")

    # ------------------------------------------------------------------
    # Autosave helpers
    # ------------------------------------------------------------------
    def _mark_dirty(self) -> None:
        self._dirty = True

    def _resolve_autosave_delay(self) -> int:
        try:
            raw = self.cfg.get("ui.autosave_delay_sec", 3.0)
            delay_sec = float(raw)
        except Exception:
            delay_sec = 3.0
        delay_ms = int(delay_sec * 1000)
        if delay_ms <= 0:
            return 0
        return max(delay_ms, 500)

    def _bind_defaults_shortcut(self) -> None:
        try:
            self.master.bind_all("<Control-r>", self._on_defaults_kbd)
        except Exception:
            pass

    def _start_autosave_loop(self) -> None:
        if getattr(self, "_autosave_delay_ms", 0) <= 0:
            return
        try:
            self._autosave_job = self.master.after(
                self._autosave_delay_ms, self._autosave_tick
            )
        except Exception:
            self._autosave_job = None

    def _autosave_tick(self) -> None:
        self._autosave_job = None
        if self._dirty:
            try:
                self.save()
                self._status("Zmiany zapisane automatycznie")
            except Exception as exc:
                logger.exception("[SETTINGS] autosave failed: %s", exc)
                self._status(f"Błąd zapisu: {exc}")
        self._start_autosave_loop()

    def _on_defaults_kbd(self, _event=None):
        self.restore_defaults()
        return "break"

    def _status(self, message: str) -> None:
        try:
            title_base = getattr(self, "_base_title", "") or "Ustawienia"
            top = self.master.winfo_toplevel()
            top.title(f"{title_base} – {message}")
        except Exception:
            logger.info("[SETTINGS] %s", message)

    def _on_button_field_clicked(self, field: dict[str, Any], label: str) -> None:
        """Execute configured action for schema button field."""

        action = field.get("action")
        params = field.get("params", {}) or {}
        wm_dbg("ui.button", "click", label=label, action=action, params=params)
        if not action:
            return
        try:
            result = settings_action_exec(action, params)
            ok = True
            if isinstance(result, dict) and "ok" in result:
                ok = bool(result["ok"])
            wm_info(
                "ui.button",
                "done",
                label=label,
                action=action,
                ok=ok,
                result=result,
            )
        except RuntimeError as exc:
            wm_err(
                "ui.button",
                "action failed",
                exc,
                label=label,
                action=action,
                params=params,
            )
            messagebox.showerror(
                "Błąd akcji ustawień",
                str(exc),
                parent=self.master,
            )
            return
        except Exception as exc:
            wm_err(
                "ui.button",
                "action failed",
                exc,
                label=label,
                action=action,
                params=params,
            )
            messagebox.showerror(
                "Błąd akcji ustawień",
                f"Nie udało się wykonać akcji: {exc}",
                parent=self.master,
            )
            return

        write_to_key = params.get("write_to_key")
        if write_to_key:
            self.on_setting_changed(
                write_to_key, self.settings_state.get(write_to_key)
            )

    def on_setting_changed(self, key: str, value: Any) -> None:
        """Callback invoked by action handlers when config value changes."""

        wm_info("ui.settings.change", "value updated", key=key, value=value)
        self.settings_state[key] = value
        var = self.vars.get(key)
        if var is None:
            return
        opt = self._options.get(key, {"key": key})
        try:
            coerced = self._coerce_default_for_var(opt, value)
        except Exception as exc:
            wm_err(
                "ui.settings.change",
                "coerce failed",
                exc,
                key=key,
                value=value,
            )
            coerced = value
        try:
            if isinstance(var, tk.BooleanVar):
                var.set(bool(value))
            elif isinstance(var, tk.IntVar):
                var.set(int(value))
            elif isinstance(var, tk.DoubleVar):
                var.set(float(value))
            else:
                var.set(coerced)
        except Exception as exc:
            wm_err(
                "ui.settings.change",
                "var set failed",
                exc,
                key=key,
                value=value,
            )
            try:
                var.set(value)
            except Exception as fallback_exc:
                wm_err(
                    "ui.settings.change",
                    "var fallback set failed",
                    fallback_exc,
                    key=key,
                    value=value,
                )

    def _add_group(
        self,
        parent: tk.Widget,
        title: str,
        *,
        description: str | None = None,
        namespace: str | None = None,
    ) -> ttk.LabelFrame:
        """Create labeled frame for manual settings sections."""

        group = ttk.LabelFrame(parent, text=title)
        group.pack(fill="x", padx=10, pady=(10, 6))
        group.columnconfigure(0, weight=1)
        if namespace:
            setattr(group, "_settings_namespace", namespace)
        if description:
            ttk.Label(
                group,
                text=description,
                wraplength=560,
                font=("", 9, "italic"),
            ).pack(anchor="w", padx=8, pady=(6, 2))
        return group

    def _add_field(
        self,
        group: ttk.LabelFrame,
        key: str,
        label: str,
        *,
        field_type: str = "string",
        default: Any = None,
        description: str = "",
    ) -> tk.Variable:
        """Create a labeled field inside ``group`` and register variable."""

        namespace = getattr(group, "_settings_namespace", None)
        if namespace and "." not in key:
            full_key = f"{namespace}.{key}"
        else:
            full_key = key

        frame = ttk.Frame(group)
        frame.pack(fill="x", padx=8, pady=4)
        ttk.Label(frame, text=label).pack(anchor="w")

        def format_list(value: Any) -> str:
            if isinstance(value, (list, tuple, set)):
                return "\n".join(str(item).strip() for item in value if str(item).strip())
            if isinstance(value, str):
                return value
            return ""

        def format_dict(value: Any) -> str:
            if isinstance(value, dict):
                return "\n".join(f"{k} = {v}" for k, v in value.items())
            if isinstance(value, str):
                return value
            return ""

        def format_nested(value: Any) -> str:
            if isinstance(value, dict):
                lines: list[str] = []
                for k, vals in value.items():
                    key_str = str(k).strip()
                    if not key_str:
                        continue
                    vals_list = [str(item).strip() for item in vals if str(item).strip()]
                    joined = ", ".join(vals_list)
                    lines.append(f"{key_str}: {joined}" if joined else key_str)
                return "\n".join(lines)
            if isinstance(value, str):
                return value
            return ""

        widget: tk.Widget
        var: tk.Variable
        field_def: dict[str, Any]

        if field_type == "list":
            text_value = format_list(default)
            var = StrListVar(master=frame)
            var.set(text_value)
            text = tk.Text(frame, height=4, wrap="word")
            text.insert("1.0", text_value)

            def update_list(*_args: Any) -> None:
                var.set(text.get("1.0", "end").strip())

            text.bind("<KeyRelease>", update_list)
            text.pack(fill="x", expand=True, pady=(2, 0))
            widget = text
            field_def = {
                "key": full_key,
                "type": "array",
                "value_type": "string",
                "default": default if isinstance(default, list) else [],
            }
        elif field_type == "dict_float":
            text_value = format_dict(default)
            var = FloatDictVar(master=frame)
            var.set(text_value)
            text = tk.Text(frame, height=4, wrap="word")
            text.insert("1.0", text_value)

            def update_fdict(*_args: Any) -> None:
                var.set(text.get("1.0", "end").strip())

            text.bind("<KeyRelease>", update_fdict)
            text.pack(fill="x", expand=True, pady=(2, 0))
            widget = text
            field_def = {
                "key": full_key,
                "type": "dict",
                "value_type": "float",
                "default": default if isinstance(default, dict) else {},
            }
        elif field_type == "dict":
            text_value = format_dict(default)
            var = StrDictVar(master=frame)
            var.set(text_value)
            text = tk.Text(frame, height=4, wrap="word")
            text.insert("1.0", text_value)

            def update_dict(*_args: Any) -> None:
                var.set(text.get("1.0", "end").strip())

            text.bind("<KeyRelease>", update_dict)
            text.pack(fill="x", expand=True, pady=(2, 0))
            widget = text
            field_def = {
                "key": full_key,
                "type": "dict",
                "value_type": "string",
                "default": default if isinstance(default, dict) else {},
            }
        elif field_type == "nested_list":
            text_value = format_nested(default)
            var = NestedListVar(master=frame)
            var.set(text_value)
            text = tk.Text(frame, height=6, wrap="word")
            text.insert("1.0", text_value)

            def update_nested(*_args: Any) -> None:
                var.set(text.get("1.0", "end").strip())

            text.bind("<KeyRelease>", update_nested)
            text.pack(fill="x", expand=True, pady=(2, 0))
            widget = text
            field_def = {
                "key": full_key,
                "type": "dict",
                "default": default if isinstance(default, dict) else {},
            }
        elif field_type == "int":
            value = default if isinstance(default, int) else 0
            var = tk.IntVar(master=frame, value=value)
            spin = ttk.Spinbox(frame, from_=1, to=10, textvariable=var, width=6)
            spin.pack(anchor="w", pady=(2, 0))
            widget = spin
            field_def = {"key": full_key, "type": "int", "default": value}
        else:
            value = "" if default is None else str(default)
            var = tk.StringVar(master=frame, value=value)
            entry = ttk.Entry(frame, textvariable=var)
            entry.pack(fill="x", expand=True, pady=(2, 0))
            widget = entry
            field_def = {"key": full_key, "type": "string", "default": value}

        if description:
            ttk.Label(
                frame,
                text=description,
                wraplength=560,
                font=("", 9),
                foreground="#565656",
            ).pack(anchor="w", pady=(2, 0))

        self._register_option_var(full_key, var, field_def)

        if full_key.startswith("_orders."):
            if not hasattr(self, "_orders_vars"):
                self._orders_vars = {}
            if not hasattr(self, "_orders_meta"):
                self._orders_meta = {}
            name = full_key.split(".", 1)[1]
            self._orders_vars[name] = var
            self._orders_meta[name] = {
                "type": field_type,
                "widget": widget,
            }

        return var

    def _build_products_tab(
        self, parent: tk.Widget, tab: dict[str, Any]
    ) -> tuple[int, int]:
        """Create combined products tab with custom manager and schema config."""

        nb = ttk.Notebook(parent)
        nb.pack(fill="both", expand=True)

        self.products_tab = ProductsMaterialsTab(nb, base_dir=self._base_dir)
        nb.add(self.products_tab, text="Zarządzanie danymi")
        print("[WM-DBG] [SETTINGS] zakładka Produkty i materiały: OK")

        config_frame = ttk.Frame(nb)
        nb.add(config_frame, text="Konfiguracja")

        config_path = ("produkty", "konfiguracja")
        self._tab_frames[config_path] = config_frame
        grp_count, fld_count = self._populate_tab(
            config_frame,
            tab,
            skip_products_override=True,
            tab_path=config_path,
        )
        return grp_count, fld_count

    def _populate_tab(
        self,
        parent: tk.Widget,
        tab: dict[str, Any],
        *,
        skip_products_override: bool = False,
        tab_path: tuple[str, ...] | None = None,
    ) -> tuple[int, int]:
        """Populate a single tab or subtab frame and return counts."""

        tab_path = tuple(tab_path or ())
        self._remember_tab_frame(tab_path, parent)
        grp_count = 0
        fld_count = 0

        tab_id_raw = tab.get("id")
        tab_id = str(tab_id_raw or "").strip().lower()

        if not skip_products_override and tab_id == "produkty":
            return self._build_products_tab(parent, tab)

        for group in tab.get("groups", []):
            if _is_deprecated(group):
                ident = group.get("label") or group.get("id") or "group"
                print(
                    f"[WM-DBG][SETTINGS] pomijam deprecated {ident}"
                )
                continue
            grp_count += 1
            grp_label = group.get("label") or group.get("title") or ""
            grp_frame = ttk.LabelFrame(parent, text=grp_label)
            grp_frame.pack(fill="x", expand=False, padx=5, pady=5)
            tip = group.get("tooltip") or group.get("description")
            if tip:
                _bind_tooltip(grp_frame, tip)
            inner = ttk.Frame(grp_frame)
            inner.pack(fill="x", expand=False, padx=8, pady=6)

            for field_def in group.get("fields", []):
                if _is_deprecated(field_def):
                    ident = field_def.get("key", "field")
                    print(
                        f"[WM-DBG][SETTINGS] pomijam deprecated {ident}"
                    )
                    continue
                if _is_legacy_system_field(field_def):
                    ident = field_def.get("key", "field")
                    print(
                        f"[WM-DBG][SETTINGS] pomijam legacy field {ident}"
                    )
                    continue
                if field_def.get("type") == "info":
                    label = field_def.get("label_pl") or field_def.get("label")
                    text = field_def.get("text_pl") or field_def.get("text") or ""
                    _add_readonly_info(inner, text, label=label)
                    fld_count += 1
                    continue
                key = field_def.get("key")
                if not key:
                    continue
                if key in getattr(self, "_manual_fields_keys", set()):
                    # TODO: przenieść do nowego układu zakładek – pole obsługiwane ręcznie
                    continue
                fld_count += 1
                normalized_def = _normalize_field_definition(field_def)
                if key == "auth.auto_login_profile":
                    normalized_def = self._prepare_auto_login_field(normalized_def)
                original_default = normalized_def.get("default")
                current_value = self.cfg.get(key, original_default)
                current_value = _coerce_value_for_type(current_value, normalized_def)
                option_for_widget = dict(normalized_def)
                option_for_widget["default"] = current_value
                self._options[key] = dict(normalized_def)
                if normalized_def.get("type") == "button":
                    btn = self._create_button_field(inner, field_def)
                    btn.pack(fill="x", padx=5, pady=2)
                    continue
                field, var = _create_widget(option_for_widget, inner)
                field.pack(fill="x", padx=5, pady=2)
                self.vars[key] = var
                self._initial[key] = current_value
                self._defaults[key] = original_default
                self._fields_vars.append((var, dict(normalized_def)))
                self.settings_state[key] = current_value
                var.trace_add("write", lambda *_: self._on_var_write(key, var))

            if tab_id in {"zarzadzanie_lista", "users_list", "profiles_list"}:
                users_frame = ttk.Frame(inner)
                users_frame.pack(fill="both", expand=True)
                owner = (
                    self.master.winfo_toplevel()
                    if hasattr(self.master, "winfo_toplevel")
                    else self.master
                )
                panel_uzytkownicy(owner, users_frame)

            group_key = str(group.get("key") or "").strip().lower()
            # (usunięto duplikat przycisku „Edytor definicji zadań…”, jest w _build_tools_tab)

        if subtabs := tab.get("subtabs"):
            nb = ttk.Notebook(parent)
            nb.pack(fill="both", expand=True)
            for sub in subtabs:
                title = sub.get("title", sub.get("id", ""))
                sub_frame = ttk.Frame(nb)
                nb.add(sub_frame, text=title)
                sub_id = str(sub.get("id") or "").strip().lower()
                sub_path = tuple(filter(None, (*tab_path, sub_id)))
                self._remember_tab_frame(sub_path, sub_frame)
                g, f = self._populate_tab(
                    sub_frame,
                    sub,
                    tab_path=sub_path,
                    skip_products_override=skip_products_override,
                )
                grp_count += g
                fld_count += f

        if tab_id in {"profile", "profiles", "users", "uzytkownicy"}:
            users = [u.get("login", "") for u in profile_service.get_all_users()]
            sel = ttk.Frame(parent)
            sel.pack(fill="x", padx=5, pady=5)
            ttk.Label(sel, text="Użytkownik:").pack(side="left")
            self._user_var = tk.StringVar()
            cmb = ttk.Combobox(
                sel, values=users, textvariable=self._user_var, state="readonly"
            )
            cmb.pack(side="left", fill="x", expand=True, padx=5)
            cmb.bind(
                "<<ComboboxSelected>>",
                lambda _e: self._load_user_modules(self._user_var.get()),
            )
            box = ttk.LabelFrame(parent, text="Widoczność modułów")
            box.pack(fill="x", padx=5, pady=5)
            self._mod_vars = {}
            for key, label in SIDEBAR_MODULES:
                var = tk.BooleanVar(value=True)
                self._mod_vars[key] = var
                ttk.Checkbutton(box, text=label, variable=var).pack(anchor="w")
            ttk.Button(
                parent,
                text="Zastosuj teraz",
                command=self._apply_user_modules,
            ).pack(fill="x", padx=5, pady=(0, 5))

        if tab_id in {"update", "updates", "aktualizacje", "aktualizacja"}:
            self._add_patch_section(parent)

        return grp_count, fld_count

    def focus_tab(self, *path: str) -> bool:
        """Select notebook tab specified by ``path`` (e.g. ("uzytkownicy", "profile_config"))."""

        normalized = tuple(
            filter(None, (str(part).strip().lower() for part in path))
        )
        if not normalized:
            return False
        success = False
        for depth in range(1, len(normalized) + 1):
            segment = normalized[:depth]
            frame = self._tab_frames.get(segment)
            if frame is None:
                continue
            notebook = getattr(frame, "master", None)
            if hasattr(notebook, "select"):
                try:
                    notebook.select(frame)
                    success = True
                except Exception:
                    continue
        return success

    def _prepare_auto_login_field(self, option: dict[str, Any]) -> dict[str, Any]:
        """Populate enum choices for the auto-login profile selector."""

        option = dict(option)
        entries: list[dict[str, str]] = [{"label": "— brak —", "value": ""}]
        allowed: list[str] = [""]
        labels: dict[str, str] = {"": "— brak —"}

        try:
            users = profile_service.get_all_users()
        except Exception:
            logger.exception("[SETTINGS] Nie udało się pobrać listy profili do auto-logowania")
            users = []

        seen: set[str] = set()
        for user in users:
            try:
                if not user.get("active", True):
                    continue
                login_raw = str(user.get("login", "")).strip()
                if not login_raw:
                    continue
                canonical = login_raw.lower()
                if canonical in seen:
                    continue
                seen.add(canonical)
                role = str(user.get("rola", "")).strip()
                label = f"{login_raw} ({role})" if role else login_raw
                allowed.append(login_raw)
                labels[login_raw] = label
                entries.append({"label": label, "value": login_raw})
            except Exception:
                logger.exception("[SETTINGS] Błąd podczas budowania listy profili auto-logowania")

        option["enum"] = allowed
        option["enum_labels"] = labels
        option["options"] = entries

        current_default = option.get("default") or ""
        current_default = str(current_default).strip()
        if current_default and current_default.lower() not in seen:
            current_default = ""
        option["default"] = current_default
        return option

    def _build_orders_tab(
        self, parent: tk.Widget, tab: dict[str, Any]
    ) -> tuple[int, int]:
        """Create advanced editor for the Orders settings tab."""

        self._orders_vars = {}
        self._orders_meta = {}

        orders_cfg_raw = self.cfg.get("orders", {})
        orders_cfg = orders_cfg_raw if isinstance(orders_cfg_raw, dict) else {}
        orders_cfg = copy.deepcopy(orders_cfg)

        types_raw = orders_cfg.get("types", {})
        base_types: dict[str, dict[str, Any]] = {
            code: copy.deepcopy(data) for code, data in DEFAULT_ORDER_TYPES.items()
        }
        if isinstance(types_raw, dict):
            for code, data in types_raw.items():
                if isinstance(data, dict):
                    target = base_types.setdefault(
                        code, copy.deepcopy(DEFAULT_ORDER_TYPES.get(code, {}))
                    )
                    target.update(copy.deepcopy(data))

        for code, data in base_types.items():
            data.setdefault("label", DEFAULT_ORDER_TYPES.get(code, {}).get("label", code))
            data.setdefault(
                "prefix", DEFAULT_ORDER_TYPES.get(code, {}).get("prefix", f"{code}-")
            )
            statuses_default = DEFAULT_ORDER_TYPES.get(code, {}).get("statuses", ["nowe"])
            data.setdefault("statuses", list(statuses_default))
            if "enabled" not in data:
                data["enabled"] = bool(
                    DEFAULT_ORDER_TYPES.get(code, {}).get("enabled", True)
                )

        enabled_types = [
            code for code, data in base_types.items() if data.get("enabled", True)
        ]
        prefixes = {
            code: data.get("prefix", f"{code}-") for code, data in base_types.items()
        }
        statuses_map = {
            code: list(data.get("statuses", [])) for code, data in base_types.items()
        }

        id_width_raw = orders_cfg.get("id_width", 4)
        try:
            id_width = int(id_width_raw)
        except (TypeError, ValueError):
            id_width = 4

        colors = orders_cfg.get("status_colors")
        colors = colors if isinstance(colors, dict) else {}

        tasks = orders_cfg.get("tasks")
        if not isinstance(tasks, list):
            fallback_tasks = self.cfg.get("czynnosci_technologiczne", [])
            tasks = fallback_tasks if isinstance(fallback_tasks, list) else []

        alerts_raw = orders_cfg.get("alert_thresholds_pct")
        alerts: dict[str, float] = {}
        if isinstance(alerts_raw, dict):
            for key, value in alerts_raw.items():
                try:
                    alerts[str(key)] = float(value)
                except (TypeError, ValueError):
                    continue
        elif isinstance(alerts_raw, list):
            codes = list(base_types.keys())
            for idx, code in enumerate(codes):
                try:
                    alerts[code] = float(alerts_raw[idx])
                except (IndexError, TypeError, ValueError):
                    alerts[code] = 50.0
        if not alerts:
            alerts = {code: 50.0 for code in base_types.keys()}

        links = orders_cfg.get("module_links")
        if not isinstance(links, dict) or not links:
            links = {
                "ZN": "Powiąż z kartoteką narzędzi (zakres SN 500–1000)",
                "ZM": "Powiąż z modułem Maszyny",
                "ZZ": "Powiąż z modułem Magazyn → Zamówienia",
            }

        defaults_cfg = orders_cfg.get("defaults")
        if isinstance(defaults_cfg, dict):
            defaults = {
                str(k): str(v) if not isinstance(v, (int, float)) else str(v)
                for k, v in defaults_cfg.items()
            }
        else:
            defaults = {}
        defaults.setdefault("author", "zalogowany_uzytkownik")
        defaults.setdefault("status", "nowe")
        defaults.setdefault("id_width", str(id_width))

        grp_count = 0
        fld_count = 0

        group = self._add_group(
            parent,
            "Definicje typów zleceń (ZW/ZN/ZM/ZZ)",
            namespace="_orders",
        )
        grp_count += 1
        self._add_field(
            group,
            "enabled_types",
            "Typy aktywne",
            field_type="list",
            default=enabled_types,
            description=(
                "Lista aktywnych typów zleceń. Dostępne: ZW (wewnętrzne), "
                "ZN (na narzędzie), ZM (maszyny), ZZ (zakup). Podaj jeden kod w "
                "wierszu."
            ),
        )
        fld_count += 1
        self._add_field(
            group,
            "prefixes",
            "Prefiksy ID",
            field_type="dict",
            default=prefixes,
            description=(
                "Prefiks używany w numeracji zleceń (np. ZW- lub ZN-). Wpisuj "
                "w formacie `ZW = ZW-`."
            ),
        )
        fld_count += 1
        self._add_field(
            group,
            "id_width",
            "Szerokość ID",
            field_type="int",
            default=id_width,
            description="Ile cyfr ma mieć numer zlecenia. Np. 4 = ZW-0001.",
        )
        fld_count += 1

        group_status = self._add_group(
            parent,
            "Statusy i kolory",
            namespace="_orders",
        )
        grp_count += 1
        self._add_field(
            group_status,
            "statuses",
            "Lista statusów",
            field_type="nested_list",
            default=statuses_map,
            description=(
                "Definicje statusów dla poszczególnych typów zleceń. Każdy wiersz "
                "w formacie `ZW: nowe, w przygotowaniu, w realizacji`."
            ),
        )
        fld_count += 1
        self._add_field(
            group_status,
            "colors",
            "Kolory statusów",
            field_type="dict",
            default=colors,
            description=(
                "Kolor przypisany do statusu. Przykład: `w realizacji = #F1C40F`, "
                "`awaria = blink_red`."
            ),
        )
        fld_count += 1

        group_tasks = self._add_group(
            parent,
            "Czynności technologiczne",
            namespace="_orders",
        )
        grp_count += 1
        self._add_field(
            group_tasks,
            "tasks",
            "Lista czynności",
            field_type="list",
            default=tasks,
            description=(
                "Standardowe czynności technologiczne przypisywane do statusów "
                "(np. Sprawdź magazyn, Zarezerwuj półprodukty). Jeden wpis na "
                "wiersz."
            ),
        )
        fld_count += 1

        group_alerts = self._add_group(
            parent,
            "Progi alertów (%)",
            namespace="_orders",
        )
        grp_count += 1
        self._add_field(
            group_alerts,
            "alerts",
            "Alerty dla typów",
            field_type="dict_float",
            default=alerts,
            description=(
                "Próg procentowy poniżej którego system zgłasza alert (osobno "
                "dla ZW, ZN, ZM, ZZ). Format: `ZW = 75`."
            ),
        )
        fld_count += 1

        group_links = self._add_group(
            parent,
            "Powiązania modułowe",
            namespace="_orders",
        )
        grp_count += 1
        self._add_field(
            group_links,
            "links",
            "Powiązania",
            field_type="dict",
            default=links,
            description=(
                "Powiązania z innymi modułami: ZN → narzędzia SN (500–1000), "
                "ZM → maszyny, ZZ → magazyn (zamówienia)."
            ),
        )
        fld_count += 1

        group_defaults = self._add_group(
            parent,
            "Domyślne wartości",
            namespace="_orders",
        )
        grp_count += 1
        self._add_field(
            group_defaults,
            "defaults",
            "Ustawienia domyślne",
            field_type="dict",
            default=defaults,
            description=(
                "Domyślny autor = zalogowany użytkownik, domyślny status = nowe, "
                "szerokość ID = 4 cyfry. Możesz rozszerzyć o inne pola np. "
                "`priority = normal`."
            ),
        )
        fld_count += 1

        return grp_count, fld_count

    def _apply_orders_config(self, values: dict[str, Any]) -> None:
        """Persist composite Orders settings based on manual fields."""

        orders_cfg_raw = self.cfg.get("orders", {})
        orders_cfg = orders_cfg_raw if isinstance(orders_cfg_raw, dict) else {}
        orders_cfg = copy.deepcopy(orders_cfg)

        types_raw = orders_cfg.get("types", {})
        types_cfg: dict[str, dict[str, Any]] = {}
        if isinstance(types_raw, dict):
            for code, data in types_raw.items():
                if isinstance(data, dict):
                    types_cfg[code] = copy.deepcopy(data)
                else:
                    types_cfg[code] = {}

        for code, defaults in DEFAULT_ORDER_TYPES.items():
            base = types_cfg.setdefault(code, copy.deepcopy(defaults))
            if not isinstance(base, dict):
                base = copy.deepcopy(defaults)
                types_cfg[code] = base
            base.setdefault("label", defaults.get("label", code))
            base.setdefault("prefix", defaults.get("prefix", f"{code}-"))
            base.setdefault("statuses", list(defaults.get("statuses", ["nowe"])))
            base.setdefault("enabled", bool(defaults.get("enabled", True)))

        def ensure_type_entry(code: str) -> dict[str, Any]:
            norm = code.strip().upper()
            if not norm:
                return {}
            entry = types_cfg.get(norm)
            if not isinstance(entry, dict):
                defaults = DEFAULT_ORDER_TYPES.get(norm, {})
                entry = copy.deepcopy(defaults) if isinstance(defaults, dict) else {}
                entry.setdefault("label", defaults.get("label", norm) if isinstance(defaults, dict) else norm)
                entry.setdefault("prefix", defaults.get("prefix", f"{norm}-") if isinstance(defaults, dict) else f"{norm}-")
                entry.setdefault(
                    "statuses",
                    list(defaults.get("statuses", ["nowe"]))
                    if isinstance(defaults, dict)
                    else ["nowe"],
                )
                entry.setdefault("enabled", True)
                types_cfg[norm] = entry
            return entry

        if "prefixes" in values:
            prefixes: dict[str, Any] = values.get("prefixes", {}) or {}
            for code, prefix in prefixes.items():
                entry = ensure_type_entry(str(code))
                prefix_str = str(prefix).strip()
                if prefix_str:
                    entry["prefix"] = prefix_str

        if "statuses" in values:
            statuses_val: dict[str, list[str]] = values.get("statuses", {}) or {}
            for code, statuses in statuses_val.items():
                entry = ensure_type_entry(str(code))
                cleaned = [
                    str(item).strip()
                    for item in statuses
                    if isinstance(item, str) and str(item).strip()
                ]
                if cleaned:
                    entry["statuses"] = cleaned

        if "enabled_types" in values:
            enabled = {
                str(code).strip().upper()
                for code in values.get("enabled_types", [])
                if str(code).strip()
            }
            for code in list(types_cfg.keys()):
                types_cfg[code]["enabled"] = code in enabled if code else False
            for code in enabled:
                ensure_type_entry(code)["enabled"] = True

        if "id_width" in values:
            try:
                orders_cfg["id_width"] = max(1, int(values["id_width"]))
            except (TypeError, ValueError):
                pass

        if "colors" in values:
            colors_src: dict[str, Any] = values.get("colors", {}) or {}
            colors: dict[str, str] = {}
            for key, val in colors_src.items():
                key_str = str(key).strip()
                val_str = str(val).strip()
                if key_str and val_str:
                    colors[key_str] = val_str
            orders_cfg["status_colors"] = colors

        if "tasks" in values:
            tasks_raw = values.get("tasks", []) or []
            tasks_list = [
                str(item).strip()
                for item in tasks_raw
                if isinstance(item, str) and str(item).strip()
            ]
            orders_cfg["tasks"] = tasks_list

        if "alerts" in values:
            alerts_src: dict[str, Any] = values.get("alerts", {}) or {}
            alerts_dict: dict[str, float] = {}
            for key, val in alerts_src.items():
                key_str = str(key).strip().upper()
                if not key_str:
                    continue
                try:
                    alerts_dict[key_str] = float(val)
                except (TypeError, ValueError):
                    continue
            orders_cfg["alert_thresholds_pct"] = alerts_dict

        if "links" in values:
            links_src: dict[str, Any] = values.get("links", {}) or {}
            links_dict: dict[str, str] = {}
            for key, val in links_src.items():
                key_str = str(key).strip()
                val_str = str(val).strip()
                if key_str and val_str:
                    links_dict[key_str] = val_str
            orders_cfg["module_links"] = links_dict

        if "defaults" in values:
            defaults_src: dict[str, Any] = values.get("defaults", {}) or {}
            defaults_dict: dict[str, str] = {}
            for key, val in defaults_src.items():
                key_str = str(key).strip()
                if not key_str:
                    continue
                defaults_dict[key_str] = str(val).strip()
            if "id_width" in values and "id_width" not in defaults_dict:
                defaults_dict["id_width"] = str(values["id_width"])
            orders_cfg["defaults"] = defaults_dict
        elif "id_width" in values:
            defaults_existing = orders_cfg.get("defaults")
            if isinstance(defaults_existing, dict):
                defaults_dict = dict(defaults_existing)
            else:
                defaults_dict = {}
            defaults_dict.setdefault("id_width", str(values["id_width"]))
            orders_cfg["defaults"] = defaults_dict

        orders_cfg["types"] = types_cfg
        self.cfg.set("orders", orders_cfg)

    def _build_tools_tab(
        self, parent: tk.Widget, tab: dict[str, Any]
    ) -> tuple[int, int]:
        """Custom layout for the Tools tab with preview tree."""

        print(
            "[INFO][WM-DBG] Buduję zakładkę: Ustawienia → Narzędzia (UI-only patch)"
        )
        field_defs: dict[str, dict[str, Any]] = {}
        for group in tab.get("groups", []):
            if _is_deprecated(group):
                continue
            for field_def in group.get("fields", []):
                if _is_deprecated(field_def):
                    continue
                key = field_def.get("key")
                if key:
                    field_defs[key] = field_def

        grp_count = 0
        fld_count = 0

        # === Import narzędzi z Excela ===
        import_group = ttk.LabelFrame(parent, text="Import narzędzi z Excela")
        import_group.pack(fill="x", padx=10, pady=(10, 6))
        import_group.grid_columnconfigure(0, weight=1)
        grp_count += 1

        instructions = (
            "Wymagane kolumny w arkuszu:\n"
            "• id – numer narzędzia, np. 001, 002, 010\n"
            "• nazwa – nazwa narzędzia\n"
            "• typ – typ narzędzia (zgodny z typami w WM)\n"
            "• status – status początkowy (np. OK, USZKODZONE, DO_PRZEGLĄDU)\n\n"
            "Każdy wiersz utworzy osobny plik JSON w katalogu narzędzi."
        )

        def _on_import_click() -> None:
            excel_path = filedialog.askopenfilename(
                title="Wybierz plik Excela z narzędziami",
                filetypes=[
                    ("Pliki Excela", "*.xlsx *.xlsm *.xls"),
                    ("Wszystkie pliki", "*.*"),
                ],
            )
            if not excel_path:
                return

            tools_dir_path = _resolve_tools_dir_from_cfg(self.cfg)
            if not tools_dir_path:
                messagebox.showerror(
                    "Import narzędzi",
                    "Nie udało się ustalić katalogu narzędzi."
                    "\nUstaw Folder WM (<root>) w zakładce Ścieżki i spróbuj ponownie.",
                )
                return

            try:
                total_rows, saved_rows, skipped_rows = _import_tools_from_excel_file(
                    excel_path, tools_dir_path
                )
            except FileNotFoundError:
                messagebox.showerror(
                    "Import narzędzi",
                    "Wybrany plik Excela nie istnieje lub jest niedostępny.",
                )
                return
            except RuntimeError as exc:
                messagebox.showerror("Import narzędzi", str(exc))
                return
            except Exception as exc:
                logger.exception("[SETTINGS][TOOLS] Błąd importu z Excela")
                messagebox.showerror("Import narzędzi", f"Błąd importu:\n{exc}")
                return

            msg_lines = [
                f"Wczytane wiersze: {total_rows}",
                f"Zapisane narzędzia: {saved_rows}",
                f"Pominięte wiersze: {skipped_rows}",
                "",
                "Katalog docelowy:",
                tools_dir_path,
            ]
            messagebox.showinfo("Import narzędzi", "\n".join(msg_lines))
            logger.info(
                "[SETTINGS][TOOLS] Import z Excela – wczytane=%s zapisane=%s pominięte=%s",
                total_rows,
                saved_rows,
                skipped_rows,
            )

        def _on_export_click() -> None:
            if Workbook is None:
                messagebox.showerror(
                    "Eksport narzędzi",
                    "Brak biblioteki openpyxl – nie można eksportować do Excela.",
                )
                return

            tools_dir_path = _resolve_tools_dir_from_cfg(self.cfg)
            if not tools_dir_path:
                messagebox.showerror(
                    "Eksport narzędzi",
                    "Nie ustawiono katalogu narzędzi w konfiguracji WM.",
                )
                return

            out_path = filedialog.asksaveasfilename(
                title="Zapisz plik Excela z narzędziami",
                defaultextension=".xlsx",
                filetypes=[
                    ("Pliki Excela", "*.xlsx"),
                    ("Wszystkie pliki", "*.*"),
                ],
            )
            if not out_path:
                return

            try:
                exported = export_tools_to_excel(out_path, tools_dir_path)
            except Exception as exc:
                logger.exception("[SETTINGS][TOOLS] Błąd eksportu do Excela")
                messagebox.showerror(
                    "Eksport narzędzi", f"Nie udało się wyeksportować:\n{exc}"
                )
                return

            messagebox.showinfo(
                "Eksport narzędzi",
                f"Wyeksportowano narzędzia: {exported}\n\nPlik:\n{out_path}",
            )
            logger.info(
                "[SETTINGS][TOOLS] Eksport do Excela – wyeksportowano=%s",
                exported,
            )

        ttk.Button(
            import_group,
            text="Import narzędzi z Excela…",
            command=_on_import_click,
        ).grid(row=0, column=0, sticky="w", padx=8, pady=(8, 4))

        ttk.Button(
            import_group,
            text="Eksport narzędzi do Excela…",
            command=_on_export_click,
        ).grid(row=0, column=1, sticky="w", padx=8, pady=(8, 4))

        ttk.Label(
            import_group,
            text=instructions,
            justify="left",
            wraplength=520,
        ).grid(row=1, column=0, sticky="w", padx=8, pady=(0, 8))

        fld_count += 1

        # === Podgląd zdjęć narzędzi ===
        preview_group = ttk.LabelFrame(parent, text="Podgląd zdjęć")
        preview_group.pack(fill="x", padx=10, pady=(10, 6))
        preview_group.grid_columnconfigure(1, weight=1)
        grp_count += 1

        ttk.Label(
            preview_group,
            text="Czas wyświetlania zdjęcia (s):",
        ).grid(row=0, column=0, sticky="w", padx=8, pady=6)

        preview_field = field_defs.get("tools.preview_delay_sec", {})
        delay_raw = self.cfg.get(
            "tools.preview_delay_sec", preview_field.get("default", 3)
        )
        try:
            delay_value = int(float(delay_raw))
        except (TypeError, ValueError):
            delay_value = int(preview_field.get("default", 3) or 3)
        delay_value = max(1, min(delay_value, 3))

        preview_delay_var = tk.StringVar(value=str(delay_value))
        preview_delay_combo = ttk.Combobox(
            preview_group,
            width=6,
            textvariable=preview_delay_var,
            values=("1", "2", "3"),
            state="readonly",
        )
        preview_delay_combo.grid(row=0, column=1, sticky="w", padx=8, pady=6)
        fld_count += 1

        self._register_option_var(
            "tools.preview_delay_sec", preview_delay_var, preview_field
        )

        # === Kolekcje narzędzi ===
        collections_group = ttk.LabelFrame(parent, text="Kolekcje narzędzi")
        collections_group.pack(fill="x", padx=10, pady=6)
        grp_count += 1
        collections_group.grid_columnconfigure(1, weight=1)

        ttk.Label(
            collections_group,
            text="Włączone kolekcje (NN, SN):",
        ).grid(row=0, column=0, sticky="w", padx=8, pady=6)
        collections_var = CSVListVar(master=collections_group)
        collections_value = self.cfg.get(
            "tools.collections_enabled",
            field_defs.get("tools.collections_enabled", {}).get("default", []),
        )
        collections_var.set(collections_value or [])
        self.entry_tools_collections_enabled = ttk.Entry(
            collections_group, width=40, textvariable=collections_var
        )
        self.entry_tools_collections_enabled.grid(
            row=0, column=1, sticky="we", padx=8, pady=6
        )
        fld_count += 1
        self._register_option_var(
            "tools.collections_enabled",
            collections_var,
            field_defs.get("tools.collections_enabled"),
        )

        default_values = ["NN", "SN"]
        for item in collections_var.get():
            if item not in default_values:
                default_values.append(item)

        ttk.Label(collections_group, text="Domyślna kolekcja:").grid(
            row=1, column=0, sticky="w", padx=8, pady=6
        )
        default_field = field_defs.get("tools.default_collection")
        default_value = self.cfg.get(
            "tools.default_collection",
            default_field.get("default") if default_field else None,
        )
        if not isinstance(default_value, str):
            default_value = str(default_value or "")
        if default_value and default_value not in default_values:
            default_values.append(default_value)
        if not default_value and default_values:
            default_value = default_values[0]
        state = "readonly" if default_values else "normal"
        default_var = tk.StringVar(value=default_value)
        self.combo_tools_default_collection = ttk.Combobox(
            collections_group,
            width=20,
            textvariable=default_var,
            values=default_values,
            state=state,
        )
        self.combo_tools_default_collection.grid(
            row=1, column=1, sticky="w", padx=8, pady=6
        )
        fld_count += 1
        self._register_option_var(
            "tools.default_collection", default_var, default_field
        )

        # === Statusy globalne ===
        global_group = ttk.LabelFrame(parent, text="Statusy globalne (zakończenia)")
        global_group.pack(fill="x", padx=10, pady=6)
        grp_count += 1

        ttk.Label(
            global_group,
            text=(
                "Lista statusów traktowanych jako globalne zakończenia "
                "(np. sprawne, zakończone):"
            ),
        ).grid(row=0, column=0, sticky="w", padx=8, pady=(8, 4))
        statuses_var = CSVListVar(master=global_group)
        statuses_value = self.cfg.get(
            "tools.auto_check_on_status_global",
            field_defs.get("tools.auto_check_on_status_global", {}).get(
                "default", []
            ),
        )
        statuses_var.set(statuses_value or [])
        self.entry_tools_global_statuses = ttk.Entry(
            global_group, width=60, textvariable=statuses_var
        )
        self.entry_tools_global_statuses.grid(
            row=1, column=0, sticky="we", padx=8, pady=(0, 8)
        )
        global_group.grid_columnconfigure(0, weight=1)
        fld_count += 1
        self._register_option_var(
            "tools.auto_check_on_status_global",
            statuses_var,
            field_defs.get("tools.auto_check_on_status_global"),
        )

        # === Podgląd definicji ===
        preview_group = ttk.LabelFrame(
            parent, text="Podgląd definicji NN/SN (tylko do odczytu)"
        )
        preview_group.pack(fill="both", expand=True, padx=10, pady=(6, 10))
        grp_count += 1

        self.tools_preview_tree = ttk.Treeview(
            preview_group,
            columns=("info",),
            show="tree headings",
            selectmode="browse",
            height=14,
        )
        self.tools_preview_tree.heading("#0", text="Struktura")
        self.tools_preview_tree.heading("info", text="Info")
        self.tools_preview_tree.column("#0", width=420, stretch=True)
        self.tools_preview_tree.column("info", width=180, stretch=False)
        self.tools_preview_tree.pack(fill="both", expand=True, padx=8, pady=8)

        counters_frame = ttk.Frame(preview_group)
        counters_frame.pack(fill="x", padx=8, pady=(0, 8))
        self.lbl_counter_types = ttk.Label(counters_frame, text="Typy: 0")
        self.lbl_counter_statuses = ttk.Label(counters_frame, text="Statusy: 0")
        self.lbl_counter_tasks = ttk.Label(counters_frame, text="Zadania: 0")
        self.lbl_counter_types.pack(side="left", padx=(0, 16))
        self.lbl_counter_statuses.pack(side="left", padx=(0, 16))
        self.lbl_counter_tasks.pack(side="left")

        btns = ttk.Frame(preview_group)
        btns.pack(fill="x", padx=8, pady=(0, 8))
        self.btn_open_tools_def_editor = ttk.Button(
            btns,
            text="Otwórz edytor definicji zadań…",
            command=self._open_tools_definitions_editor,
        )
        self.btn_open_tools_def_editor.pack(side="left")

        self._tools_paths_hidden = True

        self._refresh_tools_def_preview()

        return grp_count, fld_count

    def _open_tools_definitions_editor(self) -> None:
        try:
            self._open_tools_config()
        except Exception as exc:
            print("[INFO] Edytor definicji zadań niedostępny w tej wersji UI:", exc)

    def _refresh_tools_def_preview(self) -> None:
        """Build read-only tree with NN/SN definitions summary."""

        print("[WM-DBG] Odświeżam podgląd definicji NN/SN (read-only)")
        tree = getattr(self, "tools_preview_tree", None)
        if tree is None:
            return

        for item in tree.get_children():
            tree.delete(item)

        # --- helpers: zgodność PL/EN i różne kształty ---
        def _pick(dct: Any, *names: str) -> Any:
            if isinstance(dct, dict):
                for name in names:
                    if name in dct:
                        return dct[name]
            return None

        def _as_dict(
            obj: Any,
            key_name_for_items: str = "name",
            alt_key_name: str = "nazwa",
        ) -> dict[str, Any]:
            """Return mapping created from list of dicts or existing dict."""

            if isinstance(obj, dict):
                return obj
            if isinstance(obj, list):
                result: dict[str, Any] = {}
                for item in obj:
                    if isinstance(item, dict):
                        key = (
                            item.get(key_name_for_items)
                            or item.get(alt_key_name)
                            or item.get("key")
                            or item.get("id")
                        )
                        if key is None:
                            key = f"poz_{len(result) + 1}"
                        result[str(key)] = item
                    elif isinstance(item, str):
                        result[str(item)] = {}
                return result
            return {}

        def _extract_tasks(val: Any) -> list[str]:
            """Return list of tasks supporting PL/EN keys and plain lists."""

            if isinstance(val, list):
                return [str(x) for x in val]
            if isinstance(val, dict):
                tasks = _pick(val, "tasks", "zadania")
                if isinstance(tasks, list):
                    return [str(x) for x in tasks]
            return []

        try:
            definitions_path = self.cfg.get("tools.definitions_path")
        except Exception:
            definitions_path = None

        resolved_path = definitions_path or _default_tools_definitions_path()
        if resolved_path:
            resolved_path = os.path.abspath(resolved_path)
        definitions = load_tools_config(resolved_path) or {}
        if definitions:
            print(f"[WM-DBG][TOOLS] definicje z pliku: {resolved_path}")

        if not definitions and hasattr(self, "tools_definitions"):
            definitions = getattr(self, "tools_definitions") or {}
            print("[WM-DBG][TOOLS] używam definicji z pamięci (self.tools_definitions)")

        lbl_types = getattr(self, "lbl_counter_types", None)
        lbl_statuses = getattr(self, "lbl_counter_statuses", None)
        lbl_tasks = getattr(self, "lbl_counter_tasks", None)

        if not isinstance(definitions, dict):
            print(
                "[WM-DBG] Brak definicji lub niepoprawny format. Podgląd będzie pusty."
            )
            if lbl_types:
                lbl_types.config(text="Typy: 0")
            if lbl_statuses:
                lbl_statuses.config(text="Statusy: 0")
            if lbl_tasks:
                lbl_tasks.config(text="Zadania: 0")
            return

        # Zgodność PL/EN: collections/kolekcje → types/typy → statuses/statusy → tasks/zadania
        collections = _pick(definitions, "collections", "kolekcje") or {}
        collections_dict = _as_dict(collections)

        total_types = 0
        total_statuses = 0
        total_tasks = 0

        for coll_key, coll_val in collections_dict.items():
            coll_id = tree.insert(
                "",
                "end",
                text=str(coll_key),
                values=("kolekcja",),
            )
            types = _pick(coll_val, "types", "typy") or {}
            types_dict = _as_dict(types)

            for type_name, type_val in types_dict.items():
                statuses = _pick(type_val, "statuses", "statusy") or {}
                statuses_dict = _as_dict(statuses)

                status_count = 0
                task_count = 0
                type_id = tree.insert(
                    coll_id,
                    "end",
                    text=f"• {type_name}",
                    values=("typ",),
                )

                for status_name, status_val in statuses_dict.items():
                    status_count += 1
                    tasks = _extract_tasks(status_val)
                    task_count += len(tasks)
                    status_id = tree.insert(
                        type_id,
                        "end",
                        text=f"- {status_name}",
                        values=(f"{len(tasks)}",),
                    )
                    for task in tasks:
                        tree.insert(
                            status_id,
                            "end",
                            text=f"· {task}",
                            values=("zadanie",),
                        )

                tree.item(
                    type_id,
                    values=(f"{status_count} status., {task_count} zadań",),
                )
                total_types += 1
                total_statuses += status_count
                total_tasks += task_count

        if lbl_types:
            lbl_types.config(text=f"Typy: {total_types}")
        if lbl_statuses:
            lbl_statuses.config(text=f"Statusy: {total_statuses}")
        if lbl_tasks:
            lbl_tasks.config(text=f"Zadania: {total_tasks}")
        print(
            "[WM-DBG] Podgląd OK: "
            f"typy={total_types}, statusy={total_statuses}, zadania={total_tasks} "
            f"(plik: {definitions_path})"
        )

    def _open_tools_config(self) -> None:
        """Otwiera alias ``gui_tools_config`` (advanced lub fallback JSON)."""

        if ToolsConfigDialog is None:
            messagebox.showerror(
                "Błąd",
                "Moduł edytora narzędzi jest niedostępny.",
            )
            return

        if not hasattr(self, "_open_windows"):
            self._open_windows = {}

        try:
            existing = self._open_windows.get("tools_config")
            if existing is not None and existing.winfo_exists():
                try:
                    existing.attributes("-topmost", True)
                except Exception:
                    pass
                try:
                    existing.lift()
                    existing.focus_force()
                except Exception:
                    pass
                return
        except Exception:
            pass

        path = self._get_tools_config_path()
        try:
            # master=None, żeby uniknąć błędu
            # "SettingsWindow object has no attribute 'tk'"
            dlg = ToolsConfigDialog(
                master=None, path=path, on_save=self._on_tools_config_saved
            )
        except Exception as exc:
            # Pokaż dokładny błąd – łatwiej debugować.
            messagebox.showerror(
                "Błąd",
                (
                    "Nie udało się otworzyć edytora narzędzi:\n"
                    f"{type(exc).__name__}: {exc}"
                ),
            )
            return

        self._open_windows["tools_config"] = dlg

        def _cleanup(*_args: object) -> None:
            if self._open_windows.get("tools_config") is dlg:
                self._open_windows.pop("tools_config", None)

        try:
            dlg.bind("<Destroy>", _cleanup, add="+")
        except Exception:
            pass

        try:
            dlg.attributes("-topmost", True)
            dlg.lift()
            dlg.focus_force()
        except Exception:
            pass

        try:
            print("[WM-DBG][SETTINGS] Tools editor opened (advanced).")
            messagebox.showinfo(
                "Edytor narzędzi",
                "Otworzono edytor definicji zadań (wersja advanced).",
            )
        except Exception:
            pass

        def _fallback_topmost(win: tk.Misc) -> None:
            try:
                win.transient(self)
                win.grab_set()
                win.lift()
            except Exception:
                pass

        # A-2g: preferujemy nowe API helpera, ale obsłuż oba warianty.
        try:
            _ensure_topmost(dlg, self)
        except TypeError:
            try:
                _ensure_topmost(dlg)
            except Exception:
                _fallback_topmost(dlg)
        except Exception:
            _fallback_topmost(dlg)

        try:
            self.wait_window(dlg)
        except Exception as exc:  # pragma: no cover - wait_window error handling
            log_akcja(f"[SETTINGS] ToolsConfigDialog wait failed: {exc}")
        finally:
            _cleanup()

    def _get_tools_config_path(self) -> str:
        """Ścieżka do definicji typów/statusów narzędzi (NN/SN)."""

        try:
            cfg = getattr(self, "cfg", None)
            if cfg is not None:
                path = cfg.get("tools.definitions_path", None)  # type: ignore[attr-defined]
                if isinstance(path, str) and path.strip():
                    return path
        except Exception:
            pass
        return _default_tools_definitions_path()

    def _on_tools_config_saved(self) -> None:
        """Callback po zapisie konfiguracji narzędzi."""

        invalidate = getattr(LZ, "invalidate_cache", None)
        if callable(invalidate):
            try:
                invalidate()
            except Exception:
                pass
        try:
            self._reload_tools_section()
        except Exception:
            pass

    def _reload_tools_section(self) -> None:
        """Odświeża sekcję Narzędzia po zmianie definicji."""

        try:
            self.refresh_panel()
        except Exception:
            pass

    def _add_patch_section(self, parent: tk.Widget) -> None:
        """Append patching and version controls to the Updates tab."""

        from tools import patcher

        frame = ttk.LabelFrame(parent, text="Paczowanie i wersje")
        frame.pack(fill="x", expand=True, padx=5, pady=5)

        def audit(action: str, detail: str) -> None:
            rec = {
                "time": datetime.datetime.now().isoformat(timespec="seconds"),
                "user": "system",
                "key": action,
                "before": "",
                "after": detail,
            }
            os.makedirs(cm.AUDIT_DIR, exist_ok=True)
            path = Path(cm.AUDIT_DIR) / "config_changes.jsonl"
            with open(path, "a", encoding="utf-8") as f:
                f.write(json.dumps(rec, ensure_ascii=False) + "\n")

        def run_patch(dry: bool) -> None:
            owner = frame.winfo_toplevel() if hasattr(frame, "winfo_toplevel") else None
            reason = f"patch:{'dry' if dry else 'apply'}"
            patch_path = _safe_pick_json(
                owner,
                reason,
                title="Wybierz plik patcha",
                filetypes=[("Pakiet WM", "*.wmpatch *.zip"), ("Wszystkie pliki", "*.*")],
            )
            if not patch_path:
                return
            print(f"[WM-DBG] apply_patch dry_run={dry} path={patch_path}")
            patcher.apply_patch(patch_path, dry_run=dry)
            audit("patch.dry_run" if dry else "patch.apply", patch_path)

        ttk.Button(
            frame,
            text="Sprawdź patch (dry-run)",
            command=lambda: run_patch(True),
        ).pack(side="left", padx=5, pady=5)
        ttk.Button(
            frame,
            text="Zastosuj patch",
            command=lambda: run_patch(False),
        ).pack(side="left", padx=5, pady=5)

        commits = patcher.get_commits()
        print(f"[WM-DBG] available commits: {len(commits)}")
        roll_frame = ttk.Frame(frame)
        roll_frame.pack(fill="x", padx=5, pady=5)
        commit_var = tk.StringVar()
        ttk.Combobox(
            roll_frame,
            textvariable=commit_var,
            values=commits,
            state="readonly",
        ).pack(side="left", fill="x", expand=True)

        def rollback() -> None:
            commit = commit_var.get()
            if not commit:
                return
            print(f"[WM-DBG] rollback to {commit}")
            patcher.rollback_to(commit)
            audit("patch.rollback", commit)

        ttk.Button(
            roll_frame,
            text="Cofnij do wersji",
            command=rollback,
        ).pack(side="left", padx=5)

    def _load_magazyn_dicts(self) -> dict[str, list[str]]:
        try:
            with MAG_DICT_PATH.open("r", encoding="utf-8") as f:
                data = json.load(f)
                if isinstance(data, dict):
                    return {
                        k: [str(v) for v in data.get(k, [])]
                        for k in ("kategorie", "typy_materialu", "jednostki")
                    }
        except Exception:
            pass
        return {"kategorie": [], "typy_materialu": [], "jednostki": []}

    def _save_magazyn_dicts(self, data: dict[str, list[str]]) -> None:
        MAG_DICT_PATH.parent.mkdir(parents=True, exist_ok=True)
        with MAG_DICT_PATH.open("w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)

    def _build_slowniki_tab(self, parent: tk.Widget) -> None:
        data = self._load_magazyn_dicts()
        editors: list[tuple[str, tk.Listbox]] = []

        def make_editor(key: str, label: str) -> None:
            frame = ttk.LabelFrame(parent, text=label)
            frame.pack(side="left", fill="both", expand=True, padx=5, pady=5)

            lb = tk.Listbox(frame, height=8, exportselection=False)
            for item in data.get(key, []):
                lb.insert("end", item)
            lb.pack(fill="both", expand=True, padx=5, pady=5)
            _bind_tooltip(lb, f"Lista: {label.lower()}")

            entry = ttk.Entry(frame)
            entry.pack(fill="x", padx=5, pady=(0, 5))

            btns = ttk.Frame(frame)
            btns.pack(pady=2)

            def add_item() -> None:
                val = entry.get().strip()
                if val:
                    lb.insert("end", val)
                    entry.delete(0, "end")

            def del_item() -> None:
                for idx in reversed(lb.curselection()):
                    lb.delete(idx)

            def move_up() -> None:
                sel = lb.curselection()
                if not sel:
                    return
                idx = sel[0]
                if idx <= 0:
                    return
                val = lb.get(idx)
                lb.delete(idx)
                lb.insert(idx - 1, val)
                lb.selection_set(idx - 1)

            def move_down() -> None:
                sel = lb.curselection()
                if not sel:
                    return
                idx = sel[0]
                if idx >= lb.size() - 1:
                    return
                val = lb.get(idx)
                lb.delete(idx)
                lb.insert(idx + 1, val)
                lb.selection_set(idx + 1)

            b_add = ttk.Button(btns, text="Dodaj", command=add_item)
            b_del = ttk.Button(btns, text="Usuń", command=del_item)
            b_up = ttk.Button(btns, text="Góra", command=move_up)
            b_down = ttk.Button(btns, text="Dół", command=move_down)
            b_add.grid(row=0, column=0, padx=2)
            b_del.grid(row=0, column=1, padx=2)
            b_up.grid(row=0, column=2, padx=2)
            b_down.grid(row=0, column=3, padx=2)
            _bind_tooltip(b_add, "Dodaj wpis do listy")
            _bind_tooltip(b_del, "Usuń zaznaczony wpis")
            _bind_tooltip(b_up, "Przesuń w górę")
            _bind_tooltip(b_down, "Przesuń w dół")

            editors.append((key, lb))

        make_editor("kategorie", "Kategorie")
        make_editor("typy_materialu", "Typy materiału")
        make_editor("jednostki", "Jednostki")

        def save_all_dicts() -> None:
            payload = {key: list(lb.get(0, "end")) for key, lb in editors}
            self._save_magazyn_dicts(payload)

        btn_save = ttk.Button(parent, text="Zapisz", command=save_all_dicts)
        btn_save.pack(anchor="e", padx=5, pady=5)
        _bind_tooltip(btn_save, "Zapisz słowniki")

    def _init_magazyn_tab(self) -> None:
        """Create subtabs for the 'magazyn' section on first use."""
        if self._magazyn_frame is None or self._magazyn_schema is None:
            return
        nb = ttk.Notebook(self._magazyn_frame)
        nb.pack(fill="both", expand=True)
        # (usunięto zbędny bind na wewnętrznym notebooku Magazynu)

        ustawienia_frame = ttk.Frame(nb)
        nb.add(ustawienia_frame, text="Ustawienia magazynu")
        print("[WM-DBG] init magazyn tab")
        self._populate_tab(ustawienia_frame, self._magazyn_schema)

        slowniki_frame = ttk.Frame(nb)
        nb.add(slowniki_frame, text="Słowniki")
        self._build_slowniki_tab(slowniki_frame)

        self._magazyn_initialized = True

    def _add_patch_manager_tab(self, base_dir: str) -> None:
        """Add the "Patche" tab with helpers for scanning and restoring patches."""

        base_dir = os.fspath(base_dir)
        tab = ttk.LabelFrame(self._advanced_container, text="Patche")
        tab.pack(fill="both", expand=True, padx=8, pady=8)

        ttk.Label(tab, text="Folder patchy (*.wmpatch):").grid(
            row=0, column=0, sticky="w", padx=8, pady=(8, 2)
        )
        patches_var = tk.StringVar(value=os.path.join(base_dir, "patches"))
        entry = ttk.Entry(tab, textvariable=patches_var, width=60)
        entry.grid(row=0, column=1, sticky="ew", padx=8, pady=(8, 2))

        def _open_patch_dir() -> None:
            path = patches_var.get().strip()
            if not path:
                return
            if os.name == "nt":
                os.startfile(path)  # type: ignore[attr-defined]
                return
            for cmd in ("xdg-open", "open"):
                try:
                    subprocess.Popen([cmd, path])
                    break
                except FileNotFoundError:
                    continue

        ttk.Button(tab, text="Otwórz folder", command=_open_patch_dir).grid(
            row=0, column=2, padx=8, pady=(8, 2)
        )

        out = scrolledtext.ScrolledText(tab, height=18, width=100, state="disabled")
        out.grid(row=1, column=0, columnspan=3, sticky="nsew", padx=8, pady=8)

        btns = ttk.Frame(tab)
        btns.grid(row=2, column=0, columnspan=3, sticky="w", padx=8, pady=(0, 8))
        run_btn = ttk.Button(btns, text="Skanuj (Dry-run)")
        apply_btn = ttk.Button(btns, text="Zastosuj (Apply)")
        run_btn.pack(side="left", padx=(0, 8))
        apply_btn.pack(side="left")

        tab.columnconfigure(1, weight=1)
        tab.rowconfigure(1, weight=1)
        tab.rowconfigure(5, weight=1)

        def _append(text: str) -> None:
            out.configure(state="normal")
            out.insert("end", text + "\n")
            out.see("end")
            out.configure(state="disabled")

        def _run_patcher(do_apply: bool) -> None:
            patch_dir = patches_var.get().strip()
            if not os.path.isdir(patch_dir):
                _append(f"[WM-DBG] [PATCH] brak katalogu: {patch_dir}")
                messagebox.showwarning(
                    "Brak katalogu",
                    f"Nie znaleziono: {patch_dir}",
                    parent=tab,
                )
                return

            _append(f"[WM-DBG] [PATCH] base={base_dir}")
            _append(f"[WM-DBG] [PATCH] patches={patch_dir}")
            _append(
                "[WM-DBG] [PATCH] mode="
                + ("APPLY" if do_apply else "DRY-RUN")
            )

            try:
                try:
                    from wm_patcher import apply_patches as _apply
                except Exception:
                    _apply = None

                if _apply is not None:
                    results = _apply(
                        base_dir=base_dir,
                        patches_dir=patch_dir,
                        dry_run=not do_apply,
                        verbose=True,
                    )
                    changed = [r for r in results if r.get("changed")]
                    for result in results:
                        status = "CHANGED" if result.get("changed") else "OK"
                        _append(f"- {result.get('file')}: {status}")
                        for detail in result.get("details", []):
                            _append(f"    • {detail}")
                    messagebox.showinfo(
                        "Patche",
                        f"Zastosowano {len(results)} plików, zmian: {len(changed)}",
                        parent=tab,
                    )
                else:
                    exe = sys.executable
                    cmd = [
                        exe,
                        os.path.join(base_dir, "wm_patcher.py"),
                        "--base",
                        base_dir,
                        "--patches",
                        patch_dir,
                    ]
                    if do_apply:
                        cmd.append("--apply")
                    _append("[WM-DBG] uruchamiam: " + " ".join(cmd))
                    proc = subprocess.run(
                        cmd,
                        capture_output=True,
                        text=True,
                        cwd=base_dir,
                    )
                    if proc.stdout:
                        for line in proc.stdout.splitlines():
                            _append(line)
                    if proc.stderr:
                        _append("[STDERR] " + proc.stderr.strip())
                    messagebox.showinfo(
                        "Patche",
                        f"Zakończono (rc={proc.returncode}). Szczegóły powyżej.",
                        parent=tab,
                    )
            except Exception as exc:  # pragma: no cover - zależne od środowiska
                _append(f"[WM-DBG] [ERROR] {exc}")
                messagebox.showerror("Błąd patchera", str(exc), parent=tab)

        run_btn.configure(command=lambda: _run_patcher(False))
        apply_btn.configure(command=lambda: _run_patcher(True))

        ttk.Separator(tab).grid(row=3, column=0, columnspan=3, sticky="ew", padx=8, pady=4)
        ttk.Label(tab, text="Kopie patchy (backup/patches):").grid(
            row=4, column=0, sticky="w", padx=8
        )

        backups = ttk.Treeview(tab, columns=("files",), show="tree")
        backups.grid(row=5, column=0, columnspan=3, sticky="nsew", padx=8, pady=(2, 8))

        def _refresh_backups() -> None:
            root_dir = os.path.join(base_dir, "backup", "patches")
            for item in backups.get_children():
                backups.delete(item)
            if not os.path.isdir(root_dir):
                _append(f"[WM-DBG] brak folderu kopii: {root_dir}")
                return
            for name in sorted(os.listdir(root_dir)):
                directory = os.path.join(root_dir, name)
                if not os.path.isdir(directory):
                    continue
                files = [
                    f
                    for f in os.listdir(directory)
                    if os.path.isfile(os.path.join(directory, f))
                ]
                node = backups.insert("", "end", iid=name, text=name, values=(len(files),))
                for filename in sorted(files):
                    backups.insert(node, "end", text=filename)

        def _restore_selected() -> None:
            selected = backups.selection()
            if not selected:
                messagebox.showinfo(
                    "Przywracanie",
                    "Wybierz katalog kopii (timestamp).",
                    parent=tab,
                )
                return
            timestamp = selected[0]
            root_dir = os.path.join(base_dir, "backup", "patches", timestamp)
            if not os.path.isdir(root_dir):
                messagebox.showerror(
                    "Błąd",
                    f"Brak katalogu: {root_dir}",
                    parent=tab,
                )
                return
            if not messagebox.askyesno(
                "Potwierdź",
                f"Przywrócić kopię: {timestamp}?",
                parent=tab,
            ):
                return

            restored = 0
            skipped = 0
            for filename in os.listdir(root_dir):
                src = os.path.join(root_dir, filename)
                if not os.path.isfile(src):
                    continue
                dst = os.path.join(base_dir, filename)
                if not os.path.exists(dst):
                    candidate: str | None = None
                    for dirpath, _dirnames, filenames in os.walk(base_dir):
                        if filename in filenames:
                            found = os.path.join(dirpath, filename)
                            if candidate is None:
                                candidate = found
                            else:
                                candidate = None
                                break
                    if candidate:
                        dst = candidate
                try:
                    shutil.copy2(src, dst)
                    restored += 1
                    _append(f"[WM-DBG] [RESTORE] {filename} -> {dst}")
                except Exception as exc:  # pragma: no cover - zależne od środowiska
                    skipped += 1
                    _append(f"[WM-DBG] [RESTORE] SKIP {filename}: {exc}")

            messagebox.showinfo(
                "Przywracanie",
                f"Przywrócono: {restored}, pominięto: {skipped}",
                parent=tab,
            )

        ctrl = ttk.Frame(tab)
        ctrl.grid(row=6, column=0, columnspan=3, sticky="w", padx=8, pady=(0, 8))
        ttk.Button(ctrl, text="Odśwież kopie", command=_refresh_backups).pack(
            side="left", padx=(0, 8)
        )
        ttk.Button(
            ctrl,
            text="Przywróć zaznaczoną kopię",
            command=_restore_selected,
        ).pack(side="left")

        _refresh_backups()
        _append("[WM-DBG] [SETTINGS] zakładka Patche: OK")

    def _confirm_save_changes(self, *, parent=None, allow_cancel: bool = False) -> bool:
        dirty = getattr(self, "_dirty", False)
        unsaved = getattr(self, "_unsaved", False)
        if not (dirty or unsaved):
            return True

        parent = parent or self.master
        prompt = messagebox.askyesnocancel(
            "Niezapisane zmiany",
            "Masz niezapisane zmiany. Zapisać teraz?",
            parent=parent,
        )
        if prompt is None and allow_cancel:
            return False
        if prompt:
            try:
                self.save()
            except Exception:
                logger.exception("[SETTINGS] save after prompt failed")
        return True

    def _on_tab_change(self, _=None):
        previous_tab = getattr(self, "_last_tab", None)
        if not self._confirm_save_changes(parent=self.master, allow_cancel=True):
            if previous_tab:
                try:
                    self.nb.select(previous_tab)
                except Exception:
                    pass
            return

        if self._magazyn_frame is not None and not self._magazyn_initialized:
            current_top = self.nb.select()
            if current_top == str(self.tab_modules):
                try:
                    current_inner = self._modules_nb.select()
                except Exception:
                    current_inner = ""
                if current_inner == str(self._magazyn_frame):
                    self._init_magazyn_tab()

        self._last_tab = self.nb.select()

    def restore_defaults(self) -> None:
        for var, opt in self._fields_vars:
            default = opt.get("default")
            try:
                value = self._coerce_default_for_var(opt, default)
                var.set(value)
            except Exception:
                if opt.get("type") == "bool":
                    var.set(str(bool(default)))
        self._status("Przywrócono wartości domyślne")

    def on_close(self) -> None:
        if self._autosave_job is not None:
            try:
                self.master.after_cancel(self._autosave_job)
            except Exception:
                pass
            finally:
                self._autosave_job = None
        if not self._confirm_save_changes(parent=self.master, allow_cancel=True):
            return
        self.master.winfo_toplevel().destroy()

    def _load_user_modules(self, user_id: str) -> None:
        user = profile_service.get_user(user_id) or {}
        disabled = {
            str(m).strip().lower()
            for m in user.get("disabled_modules", [])
            if m
        }
        for key, var in self._mod_vars.items():
            var.set(key not in disabled)

    def _save_user_modules(self, user_id: str) -> list[str]:
        if not user_id:
            return []
        user = profile_service.get_user(user_id) or {"login": user_id}
        disabled = [k for k, v in self._mod_vars.items() if not v.get()]
        user["disabled_modules"] = disabled
        profile_service.save_user(user)
        return disabled

    def _apply_user_modules(self) -> None:
        if self._user_var is None:
            return
        uid = self._user_var.get().strip()
        if not uid:
            return
        disabled = self._save_user_modules(uid)
        try:
            root = self.master.winfo_toplevel().master
            root.event_generate("<<SidebarReload>>", when="tail")
        except Exception:
            pass
        log_akcja(f"[SETTINGS] zastosowano moduły {uid}: {', '.join(disabled)}")

    def save(self) -> None:
        special_orders: dict[str, Any] = {}
        drive_only_re = re.compile(r"^[A-Za-z]:$")
        for key, var in self.vars.items():
            if key.startswith("_orders."):
                name = key.split(".", 1)[1]
                special_orders[name] = var.get()
                continue
            opt = self._options.get(key, {})
            value = var.get()
            if key == "profiles.editable_fields" and isinstance(value, str):
                value = [item.strip() for item in value.split(",") if item.strip()]
            if opt.get("type") == "path" and isinstance(value, str):
                stripped = value.strip()
                if drive_only_re.match(stripped):
                    messagebox.showerror(
                        "Nieprawidłowa ścieżka",
                        (
                            f"Nieprawidłowa ścieżka: {stripped}. "
                            "Podaj pełny katalog, np. C\\WM\\backup."
                        ),
                        parent=self.master,
                    )
                    return
            if opt.get("type") == "bool" and isinstance(value, str):
                if value in {"0", "1"}:
                    value = value == "1"
            if opt.get("type") == "enum":
                allowed = (
                    opt.get("allowed")
                    or opt.get("enum")
                    or opt.get("values")
                    or []
                )
                if allowed and value not in allowed:
                    value = allowed[0]
            self.cfg.set(key, value)
            self._initial[key] = value
        if special_orders:
            self._apply_orders_config(special_orders)
            for name, value in special_orders.items():
                self._initial[f"_orders.{name}"] = value
        self._saving = True
        try:
            self.cfg.save_all()
        finally:
            self._saving = False
        self._unsaved = False
        self._dirty = False
        if self._user_var is not None:
            uid = self._user_var.get().strip()
            if uid:
                disabled = self._save_user_modules(uid)
                log_akcja(
                    f"[SETTINGS] zapisano moduły {uid}: {', '.join(disabled)}"
                )

    def refresh_panel(self) -> None:
        """Reload configuration and rebuild widgets."""

        self.cfg = ConfigManager.refresh(
            config_path=self.config_path, schema_path=self.schema_path
        )
        self.vars.clear()
        self._initial.clear()
        self._defaults.clear()
        self._options.clear()
        self._fields_vars.clear()
        self._build_ui()
        if self._autosave_job is not None:
            try:
                self.master.after_cancel(self._autosave_job)
            except Exception:
                pass
            finally:
                self._autosave_job = None
        self._autosave_delay_ms = self._resolve_autosave_delay()
        self._start_autosave_loop()


class SettingsWindow(SettingsPanel):
    """Okno ustawień oparte na :class:`SettingsPanel`."""

    def __init__(
        self,
        master: tk.Misc,
        config_path: str | None = None,
        schema_path: str | None = None,
    ) -> None:
        cfg_mgr = None
        try:
            cfg_mgr = ConfigManager()
        except Exception:
            cfg_mgr = None

        if cfg_mgr is not None:
            if config_path is None:
                try:
                    config_path = str(p_config(cfg_mgr))
                except Exception:
                    config_path = None
            if schema_path is None:
                try:
                    schema_path = str(p_settings_schema(cfg_mgr))
                except Exception:
                    schema_path = None

        base_dir = Path(__file__).resolve().parent
        if config_path is None:
            config_path = str((base_dir / "config.json").resolve())
        if schema_path is None:
            schema_path = str((base_dir / "settings_schema.json").resolve())

        config_path = str(config_path)
        schema_path = str(schema_path)

        config_parent = Path(config_path).resolve().parent
        config_parent.mkdir(parents=True, exist_ok=True)

        self.config_path = config_path
        self.schema_path = schema_path
        print(f"[WM-DBG] config_path={self.config_path}")
        print(f"[WM-DBG] schema_path={self.schema_path}")

        super().__init__(master, config_path=config_path, schema_path=schema_path)
        self.schema = self.cfg.schema
        print(f"[WM-DBG] tabs loaded: {len(self.schema.get('tabs', []))}")
        self._init_audit_tab()
        self._reorder_tabs()
        try:
            setattr(master, "_wm_settings_panel", self)
        except Exception:
            pass

    def _reorder_tabs(self) -> None:
        """Ensure deterministic order for dynamically added tabs."""

        tabs = {self.nb.tab(t, "text"): t for t in self.nb.tabs()}
        audit_id = tabs.get("Audyt")
        target_order = [
            "Magazyn",
            "Zamówienia",
            "Aktualizacje & Kopie",
            "Testy/Audyt",
        ]

        if not audit_id:
            return

        for label in target_order:
            target = tabs.get(label)
            if target is not None:
                index = self.nb.index(target)
                self.nb.insert(index, audit_id)
                return

        # fallback – keep audit as last tab
        self.nb.insert("end", audit_id)

    def _init_audit_tab(self) -> None:
        """Create the Audit tab with controls."""

        frame = ttk.LabelFrame(self._advanced_container, text="Audyt")
        frame.pack(fill="both", expand=True, padx=8, pady=8)

        btn = ttk.Button(frame, text="Uruchom audyt", command=self._run_audit_now)
        btn.pack(anchor="w", padx=5, pady=5)

        txt = tk.Text(frame, height=12)
        txt.pack(fill="x", expand=False, padx=5, pady=(0, 5))

        ctrl_bar = tk.Frame(frame)
        ctrl_bar.pack(fill="x", padx=5, pady=(0, 5))
        btn_copy = tk.Button(
            ctrl_bar,
            text="Kopiuj raport",
            command=lambda: _wm_copy_audit_report(frame, getattr(self, "cfg", None), txt),
        )
        btn_copy.pack(side="right")

        def _bind_copy_shortcut(_event=None):
            _wm_copy_audit_report(frame, getattr(self, "cfg", None), txt)
            return "break"

        txt.bind("<Control-c>", _bind_copy_shortcut)
        self.btn_audit_run = btn
        self.txt_audit = txt

        tree_frame = ttk.Frame(frame)
        tree_frame.pack(fill="both", expand=True, padx=5, pady=(0, 5))
        columns = ("time", "user", "action", "details", "file")
        tree = ttk.Treeview(tree_frame, columns=columns, show="headings", height=12)
        vsb = ttk.Scrollbar(tree_frame, orient="vertical", command=tree.yview)
        tree.configure(yscrollcommand=vsb.set)
        tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        tree_frame.grid_rowconfigure(0, weight=1)
        tree_frame.grid_columnconfigure(0, weight=1)

        headers = {
            "time": "Czas",
            "user": "Użytkownik",
            "action": "Akcja/klucz",
            "details": "Szczegóły",
            "file": "Plik",
        }
        for key, label in headers.items():
            tree.heading(key, text=label)
        tree.column("time", width=150, anchor="w")
        tree.column("user", width=120, anchor="w")
        tree.column("action", width=220, anchor="w")
        tree.column("details", width=320, anchor="w")
        tree.column("file", width=140, anchor="w")
        self.audit_tree = tree

        def _populate_audit_tree() -> None:
            tree.delete(*tree.get_children())
            records: list[dict[str, Any]] = []
            try:
                audit_dir = Path(cm.AUDIT_DIR)
                if audit_dir.exists():
                    for audit_file in sorted(audit_dir.glob("*.jsonl")):
                        try:
                            with audit_file.open("r", encoding="utf-8") as handle:
                                for raw in handle:
                                    line = raw.strip()
                                    if not line:
                                        continue
                                    try:
                                        record = json.loads(line)
                                    except Exception:
                                        continue
                                    record["_audit_file"] = audit_file.name
                                    records.append(record)
                        except Exception:
                            continue
            except Exception:
                records = []

            def _ts_value(rec: dict[str, Any]) -> str:
                ts = rec.get("time") or rec.get("ts") or ""
                return str(ts)

            records.sort(key=_ts_value, reverse=True)

            for rec in records:
                time_val = rec.get("time") or rec.get("ts") or ""
                user_val = rec.get("user") or rec.get("who") or ""
                action_val = (
                    rec.get("key")
                    or rec.get("action")
                    or rec.get("event")
                    or ""
                )
                detail_val: Any = rec.get("after")
                if detail_val in ({}, [], None, ""):
                    detail_val = (
                        rec.get("detail")
                        or rec.get("path")
                        or rec.get("branch")
                        or rec.get("commit")
                    )
                if isinstance(detail_val, (dict, list)):
                    try:
                        detail_val = json.dumps(detail_val, ensure_ascii=False)
                    except Exception:
                        detail_val = str(detail_val)
                if detail_val is None:
                    detail_val = ""
                file_val = rec.get("_audit_file", "")
                tree.insert(
                    "",
                    "end",
                    values=(
                        str(time_val),
                        str(user_val),
                        str(action_val),
                        str(detail_val),
                        str(file_val),
                    ),
                )
            try:
                logger.info("[AUDYT] Wyświetlono %s pozycji audytu", len(records))
            except Exception:
                pass

        _populate_audit_tree()
        self._refresh_audit_history = _populate_audit_tree

        crash_frame = ttk.LabelFrame(frame, text="Raport błędów WM")
        crash_frame.pack(fill="x", expand=False, padx=5, pady=(0, 5))

        self._crash_log_path_var = tk.StringVar(value="Plik raportu: …")
        self._crash_log_stats_var = tk.StringVar(value="Ładuję dane o błędach…")

        lbl_path = ttk.Label(
            crash_frame,
            textvariable=self._crash_log_path_var,
            wraplength=640,
            justify="left",
        )
        lbl_path.pack(anchor="w", padx=5, pady=(5, 0))

        lbl_stats = ttk.Label(
            crash_frame,
            textvariable=self._crash_log_stats_var,
            justify="left",
        )
        lbl_stats.pack(anchor="w", padx=5, pady=(2, 5))

        btns = ttk.Frame(crash_frame)
        btns.pack(fill="x", padx=5, pady=(0, 5))
        btn_show = ttk.Button(
            btns,
            text="Pokaż ostatnie błędy",
            command=self._open_crash_log_file,
        )
        btn_show.pack(side="left", padx=(0, 5))
        btn_mark = ttk.Button(
            btns,
            text="Oznacz jako przeczytane",
            command=self._mark_crash_log_as_read,
        )
        btn_mark.pack(side="left", padx=(0, 5))
        btn_clear = ttk.Button(
            btns,
            text="Wyczyść raport błędów",
            command=self._clear_crash_log_file,
        )
        btn_clear.pack(side="left")

        self._refresh_crash_report_panel()

    def _append_audit_out(self, s: str) -> None:
        try:
            self.txt_audit.insert("end", s)
            self.txt_audit.see("end")
        except Exception:
            pass

    def _run_audit_now(self) -> None:
        try:
            self.btn_audit_run.config(state="disabled")
        except Exception:
            pass
        self._append_audit_out("\n[INFO] Uruchamiam audyt...\n")

        def _worker() -> None:
            try:
                import wm_audit_runtime

                result = wm_audit_runtime.run_audit()
                path = "audit_wm_report.txt"
                with open(path, "w", encoding="utf-8") as f:
                    f.write(result)
                msg = result + f"\n[INFO] Raport zapisano do {path}\n"
                self.txt_audit.after(0, self._append_audit_out, msg)
                refresh = getattr(self, "_refresh_audit_history", None)
                if callable(refresh):
                    try:
                        self.txt_audit.after(0, refresh)
                    except Exception:
                        pass
            except Exception as exc:
                self.txt_audit.after(
                    0, self._append_audit_out, f"[ERROR] {exc!r}\n"
                )
            finally:
                try:
                    self.btn_audit_run.after(
                        0, lambda: self.btn_audit_run.config(state="normal")
                    )
                except Exception:
                    pass

        threading.Thread(target=_worker, daemon=True).start()

    def _refresh_crash_report_panel(self) -> None:
        try:
            path = get_crash_log_path()
            self._crash_log_path_var.set(f"Plik raportu: {path}")
            total, unread = get_crash_log_stats()
            if total <= 0:
                text = "Brak zarejestrowanych błędów."
            else:
                text = f"Ostatnie błędy: {total} (nieprzeczytane: {unread})"
            self._crash_log_stats_var.set(text)
        except Exception as exc:
            self._crash_log_stats_var.set(f"Błąd odczytu raportu: {exc}")

    def _open_crash_log_file(self) -> None:
        path = get_crash_log_path()
        if not path.exists():
            messagebox.showinfo(
                "Raport błędów WM", "Brak pliku z błędami do wyświetlenia."
            )
            return
        try:
            if sys.platform.startswith("win"):
                os.startfile(path)  # type: ignore[attr-defined]
            elif sys.platform == "darwin":
                subprocess.Popen(["open", str(path)])
            else:
                subprocess.Popen(["xdg-open", str(path)])
            mark_crash_log_read()
        except Exception as exc:
            messagebox.showerror(
                "Raport błędów WM",
                f"Nie udało się otworzyć pliku raportu:\n{exc}",
            )
        finally:
            self._refresh_crash_report_panel()

    def _mark_crash_log_as_read(self) -> None:
        try:
            mark_crash_log_read()
            messagebox.showinfo(
                "Raport błędów WM", "Wszystkie wpisy oznaczono jako przeczytane."
            )
        except Exception as exc:
            messagebox.showerror(
                "Raport błędów WM",
                f"Nie udało się zaktualizować stanu raportu:\n{exc}",
            )
        finally:
            self._refresh_crash_report_panel()

    def _clear_crash_log_file(self) -> None:
        if not messagebox.askyesno(
            "Raport błędów WM",
            "Na pewno wyczyścić raport błędów? Tego działania nie można cofnąć.",
        ):
            return
        try:
            clear_crash_log()
            messagebox.showinfo(
                "Raport błędów WM", "Raport błędów został wyczyszczony."
            )
        except Exception as exc:
            messagebox.showerror(
                "Raport błędów WM",
                f"Nie udało się wyczyścić raportu błędów:\n{exc}",
            )
        finally:
            self._refresh_crash_report_panel()

    def _append_tests_out(self, s: str):
        try:
            self.txt_tests.insert("end", s)
            self.txt_tests.see("end")
        except Exception:
            pass

    def _run_all_tests(self):
        try:
            self.btn_tests_run.config(state="disabled")
        except Exception:
            pass
        self._append_tests_out("\n[INFO] Uruchamiam: pytest -q\n")
        print("[WM-DBG][SETTINGS][TESTS] start")

        def _worker():
            try:
                cmd = [sys.executable, "-m", "pytest", "-q"]
                proc = subprocess.Popen(
                    cmd,
                    cwd=os.getcwd(),
                    stdout=subprocess.PIPE,
                    stderr=subprocess.STDOUT,
                    universal_newlines=True,
                    bufsize=1,
                )
                for line in proc.stdout:
                    self.txt_tests.after(0, self._append_tests_out, line)
                ret = proc.wait()
                self.txt_tests.after(
                    0, self._append_tests_out, f"\n[INFO] Zakończono: kod wyjścia = {ret}\n"
                )
                print(f"[WM-DBG][SETTINGS][TESTS] finished ret={ret}")
            except FileNotFoundError:
                self.txt_tests.after(
                    0,
                    self._append_tests_out,
                    "\n[ERROR] Nie znaleziono pytest. Zainstaluj: pip install pytest\n",
                )
                print("[WM-DBG][SETTINGS][TESTS] pytest not found")
            except Exception as e:
                self.txt_tests.after(
                    0, self._append_tests_out, f"\n[ERROR] Błąd uruchamiania testów: {e!r}\n"
                )
                print(f"[WM-DBG][SETTINGS][TESTS] error: {e!r}")
            finally:
                try:
                    self.btn_tests_run.after(0, lambda: self.btn_tests_run.config(state="normal"))
                except Exception:
                    pass

        threading.Thread(target=_worker, daemon=True).start()

    def _install_pytest(self):
        try:
            self.btn_install_pytest.config(state="disabled")
        except Exception:
            pass
        self._append_tests_out(
            "\n[INFO] Uruchamiam: python -m pip install -U pytest\n"
        )
        print("[WM-DBG][SETTINGS][TESTS] install start")

        def _worker():
            try:
                cmd = [
                    sys.executable,
                    "-m",
                    "pip",
                    "install",
                    "-U",
                    "pytest",
                ]
                proc = subprocess.Popen(
                    cmd,
                    cwd=os.getcwd(),
                    stdout=subprocess.PIPE,
                    stderr=subprocess.STDOUT,
                    universal_newlines=True,
                    bufsize=1,
                )
                for line in proc.stdout:
                    self.txt_tests.after(0, self._append_tests_out, line)
                ret = proc.wait()
                self.txt_tests.after(
                    0, self._append_tests_out, f"\n[INFO] Zakończono: kod wyjścia = {ret}\n"
                )
                print(f"[WM-DBG][SETTINGS][TESTS] install finished ret={ret}")
            except Exception as e:
                self.txt_tests.after(
                    0,
                    self._append_tests_out,
                    f"\n[ERROR] Błąd instalacji pytest: {e!r}\n",
                )
                print(f"[WM-DBG][SETTINGS][TESTS] install error: {e!r}")
            finally:
                try:
                    self.btn_install_pytest.after(
                        0, lambda: self.btn_install_pytest.config(state="normal")
                    )
                except Exception:
                    pass

        threading.Thread(target=_worker, daemon=True).start()


if __name__ == "__main__":
    root = tk.Tk()
    ensure_theme_applied(root)
    root.title("Ustawienia")
    SettingsPanel(root)
    root.mainloop()

# ⏹ KONIEC KODU
