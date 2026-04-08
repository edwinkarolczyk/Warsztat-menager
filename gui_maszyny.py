# version: 1.0
from __future__ import annotations

import datetime as dt
import os
import subprocess
import tkinter as tk
from logging import getLogger
from tkinter import filedialog, messagebox, ttk
from typing import Any, Callable, Dict, Iterable, List, Optional, Tuple

from core.settings_manager import Settings
from gui_maszyny_view import MachinesView

try:
    from PIL import Image, ImageTk  # type: ignore
except Exception:  # pragma: no cover - PIL opcjonalne
    Image = None
    ImageTk = None

try:  # pragma: no cover - opcjonalny moduł nowego kreatora
    from wm.dyspo_wizard import open_dyspo_wizard
except Exception:  # pragma: no cover - brak nowego modułu w starszych instalacjach
    open_dyspo_wizard = None  # type: ignore

try:  # pragma: no cover - skróty dostępne tylko w nowej wersji
    from wm.gui.shortcuts import bind_ctrl_d
except Exception:  # pragma: no cover - zachowaj kompatybilność
    def bind_ctrl_d(*_args, **_kwargs):  # type: ignore
        return None


def _maybe_open_dyspo(root, context):
    if open_dyspo_wizard is None:
        return
    target = root
    if hasattr(root, "winfo_toplevel"):
        try:
            target = root.winfo_toplevel()
        except Exception:
            target = root
    if getattr(target, "tk", None) is None:
        local_tk = globals().get("tk")
        local_ttk = globals().get("ttk")
        dialog = None
        if hasattr(local_tk, "Toplevel"):
            try:
                dialog = local_tk.Toplevel(target)
            except Exception:
                dialog = None
        proceed = None
        if hasattr(local_ttk, "Button"):
            try:
                proceed = lambda: None
                local_ttk.Button(
                    dialog or target, text="Dalej", command=proceed
                )
            except Exception:
                proceed = None
        if dialog is not None and hasattr(dialog, "bind") and proceed is not None:
            try:
                dialog.bind("<Return>", proceed)
            except Exception:
                pass
        return
    open_dyspo_wizard(target, context=context)


CANVAS_W = 1000
CANVAS_H = 1000

GRID_BASE_BG_PX_X = 25
GRID_BASE_BG_PX_Y = 25

DEFAULT_BG_COLOR = "#1e1e1e"

SCHEDULE_YEAR = 2025
SCHEDULE_SOON_THRESHOLD_DAYS = 7
SCHEDULE_STATUS_COLORS = {
    "overdue": "#dc2626",
    "soon": "#ca8a04",
    "ok": "#16a34a",
    "done": "#0f766e",
    "none": "#64748b",
}
SCHEDULE_STATUS_ROW_COLORS = {
    "overdue": {"background": "#fee2e2", "foreground": "#7f1d1d"},
    "soon": {"background": "#fef3c7", "foreground": "#854d0e"},
    "ok": {"background": "#dcfce7", "foreground": "#166534"},
    "done": {"background": "#e0f2fe", "foreground": "#0c4a6e"},
    "none": {"background": "#e2e8f0", "foreground": "#475569"},
}
_TREE_STATUS_TAG_CACHE: Dict[str, bool] = {}

from core.path_utils import resolve_root_path
from utils_json import (
    normalize_doc_list_or_dict,
    safe_read_json as _safe_read_json,
    safe_write_json as _safe_write_json,
)


def _safe_clamp(v: int, lo: int, hi: int) -> int:
    return max(lo, min(hi, v))


def _canvas_bounds(canvas) -> tuple[int, int]:
    try:
        w = int(canvas.winfo_width())
        h = int(canvas.winfo_height())
        return max(1, w), max(1, h)
    except Exception:
        return (800, 600)

logger = getLogger(__name__)


def _normalize_machine_key(value: object) -> str:
    text = str(value or "").strip().lower()
    if not text:
        return ""
    filtered = [ch for ch in text if ch.isalnum()]
    normalized = "".join(filtered)
    if normalized.isdigit():
        normalized = normalized.lstrip("0") or "0"
    return normalized


def _machine_identifiers(machine: Dict[str, Any]) -> List[str]:
    identifiers: set[str] = set()
    for key in ("id", "nr_ewid", "nr", "nazwa"):
        normalized = _normalize_machine_key(machine.get(key))
        if normalized:
            identifiers.add(normalized)
    return list(identifiers)


def _entry_identifiers(entry: Dict[str, Any]) -> List[str]:
    identifiers: set[str] = set()
    for key in ("machine_id", "id", "machine", "machine_name", "nazwa"):
        normalized = _normalize_machine_key(entry.get(key))
        if normalized:
            identifiers.add(normalized)
    return list(identifiers)


def _match_schedule_entry(machine: Dict[str, Any], entry: Dict[str, Any]) -> bool:
    if not isinstance(machine, dict) or not isinstance(entry, dict):
        return False
    machine_keys = set(_machine_identifiers(machine))
    entry_keys = set(_entry_identifiers(entry))
    return bool(machine_keys & entry_keys)


def _parse_schedule_date(value: object) -> Optional[dt.date]:
    if isinstance(value, dt.datetime):
        return value.date()
    if isinstance(value, dt.date):
        return value
    if isinstance(value, (int, float)):
        ordinal = int(value)
        if ordinal > 59:  # Excel 1900 date system offset
            try:
                return dt.date.fromordinal(ordinal + 693594)
            except ValueError:
                pass
    if isinstance(value, str):
        raw = value.strip()
        if not raw:
            return None
        for fmt in (
            "%Y-%m-%d",
            "%d.%m.%Y",
            "%Y.%m.%d",
            "%d/%m/%Y",
            "%Y/%m/%d",
            "%d-%m-%Y",
            "%Y-%m-%d %H:%M:%S",
        ):
            try:
                return dt.datetime.strptime(raw, fmt).date()
            except ValueError:
                continue
        try:
            return dt.date.fromisoformat(raw)
        except ValueError:
            return None
    return None


def _format_next_label(entry: Dict[str, Any], date_obj: dt.date) -> str:
    label = date_obj.isoformat()
    typ = str(entry.get("type") or entry.get("typ") or "").strip()
    if typ:
        label = f"{label} ({typ})"
    return label


def _normalize_schedule_entry(raw: Dict[str, Any]) -> Optional[Dict[str, Any]]:
    if not isinstance(raw, dict):
        return None

    machine_id = str(raw.get("machine_id") or raw.get("id") or "").strip()
    machine_name = str(
        raw.get("machine_name")
        or raw.get("machine")
        or raw.get("nazwa")
        or raw.get("maszyna")
        or ""
    ).strip()
    date_value = raw.get("date") or raw.get("termin") or raw.get("data")
    date_obj = _parse_schedule_date(date_value)
    if date_obj is None:
        return None

    inspection_type = str(raw.get("type") or raw.get("typ") or raw.get("rodzaj") or "").strip()
    responsible = str(
        raw.get("responsible")
        or raw.get("osoba")
        or raw.get("odpowiedzialny")
        or ""
    ).strip()
    notes = str(raw.get("notes") or raw.get("uwagi") or raw.get("komentarz") or "").strip()

    status_raw = str(raw.get("status") or "").strip().lower()
    status = "wykonany" if status_raw in {"wykonany", "done", "completed"} else "planowany"

    entry: Dict[str, Any] = {
        "machine_id": machine_id,
        "machine_name": machine_name or machine_id,
        "date": date_obj.isoformat(),
        "type": inspection_type,
        "responsible": responsible,
        "notes": notes,
        "status": status,
    }

    completed_value = raw.get("completed_at") or raw.get("wykonano")
    completed_date = _parse_schedule_date(completed_value) if completed_value else None
    if status == "wykonany":
        if completed_date is not None:
            entry["completed_at"] = completed_date.isoformat()
        elif isinstance(completed_value, str) and completed_value.strip():
            entry["completed_at"] = completed_value.strip()

    card_value = str(raw.get("card") or raw.get("karta") or "").strip()
    if card_value:
        entry["card"] = card_value

    return entry


def _schedule_entry_keys(entry: Dict[str, Any]) -> List[Tuple[str, str, str]]:
    keys: List[Tuple[str, str, str]] = []
    machine_keys = _entry_identifiers(entry)
    if not machine_keys:
        machine_keys = [""]
    date_key = str(entry.get("date") or "")
    type_key = str(entry.get("type") or "").strip().lower()
    for mk in machine_keys:
        keys.append((mk, date_key, type_key))
    return keys


def _merge_schedule_status(
    new_entries: List[Dict[str, Any]], existing_entries: List[Dict[str, Any]]
) -> List[Dict[str, Any]]:
    lookup: Dict[Tuple[str, str, str], Dict[str, Any]] = {}
    for entry in existing_entries or []:
        for key in _schedule_entry_keys(entry):
            lookup[key] = entry

    seen: set[Tuple[str, str, str]] = set()
    for entry in new_entries:
        for key in _schedule_entry_keys(entry):
            seen.add(key)
            previous = lookup.get(key)
            if not previous:
                continue
            prev_status = str(previous.get("status") or "").strip().lower()
            if prev_status == "wykonany":
                entry["status"] = "wykonany"
                if previous.get("completed_at"):
                    entry["completed_at"] = previous["completed_at"]
            if previous.get("card") and not entry.get("card"):
                entry["card"] = previous["card"]

    extras: List[Dict[str, Any]] = []
    for entry in existing_entries or []:
        keys = _schedule_entry_keys(entry)
        if not keys:
            continue
        if all(key not in seen for key in keys):
            if str(entry.get("status") or "").strip().lower() == "wykonany":
                extras.append(dict(entry))

    merged = list(new_entries) + extras
    merged.sort(key=lambda item: (str(item.get("machine_id") or item.get("machine_name") or ""), str(item.get("date") or "")))
    return merged


def _schedule_summary(
    entries: Iterable[Dict[str, Any]],
    *,
    today: Optional[dt.date] = None,
    soon_threshold: int = SCHEDULE_SOON_THRESHOLD_DAYS,
) -> Dict[str, Any]:
    today = today or dt.date.today()
    parsed: List[Tuple[dt.date, Dict[str, Any]]] = []
    for entry in entries or []:
        date_obj = _parse_schedule_date(entry.get("date"))
        if date_obj is None:
            continue
        parsed.append((date_obj, entry))
    parsed.sort(key=lambda item: item[0])

    upcoming = [(date_obj, entry) for date_obj, entry in parsed if str(entry.get("status") or "").strip().lower() != "wykonany"]
    history = [
        (date_obj, entry)
        for date_obj, entry in parsed
        if str(entry.get("status") or "").strip().lower() == "wykonany"
    ]
    history.sort(key=lambda item: item[0], reverse=True)

    summary: Dict[str, Any] = {
        "upcoming": [entry for _, entry in upcoming],
        "history": [entry for _, entry in history],
        "next_entry": None,
        "next_date": None,
        "next_label": "—",
        "days": None,
        "status_key": "none",
        "status_label": "Brak danych",
        "status_text": "Brak zaplanowanych przeglądów",
        "color": SCHEDULE_STATUS_COLORS["none"],
    }

    if upcoming:
        next_date, next_entry = upcoming[0]
        days = (next_date - today).days
        if days < 0:
            status_key = "overdue"
            status_label = "Po terminie"
        elif days <= soon_threshold:
            status_key = "soon"
            status_label = "Wkrótce"
        else:
            status_key = "ok"
            status_label = "Planowany"
        summary.update(
            {
                "next_entry": next_entry,
                "next_date": next_date,
                "next_label": _format_next_label(next_entry, next_date),
                "days": days,
                "status_key": status_key,
                "status_label": status_label,
                "status_text": f"{status_label} – {summary['next_label']}",
                "color": SCHEDULE_STATUS_COLORS[status_key],
            }
        )
    elif history:
        last_date, last_entry = history[0]
        summary.update(
            {
                "next_entry": None,
                "next_date": None,
                "next_label": "—",
                "days": None,
                "status_key": "done",
                "status_label": "Wykonane",
                "status_text": f"Wykonano {last_date.isoformat()}",
                "color": SCHEDULE_STATUS_COLORS["done"],
            }
        )
    return summary


def _attach_schedule(
    rows: Iterable[Dict[str, Any]],
    schedule_entries: List[Dict[str, Any]],
    *,
    today: Optional[dt.date] = None,
    soon_threshold: int = SCHEDULE_SOON_THRESHOLD_DAYS,
) -> None:
    today = today or dt.date.today()
    for machine in rows or []:
        if not isinstance(machine, dict):
            continue
        matching = [entry for entry in schedule_entries if _match_schedule_entry(machine, entry)]
        machine["__schedule_entries"] = matching
        machine["__schedule_summary"] = _schedule_summary(
            matching,
            today=today,
            soon_threshold=soon_threshold,
        )


def _strip_schedule_fields(machine: Dict[str, Any]) -> Dict[str, Any]:
    return {k: v for k, v in machine.items() if not k.startswith("__schedule")}


def _schedule_status_key(machine: Dict[str, Any]) -> str:
    summary = machine.get("__schedule_summary") if isinstance(machine, dict) else None
    key = str((summary or {}).get("status_key") or "none")
    return key


def _ensure_tree_schedule_tag(tree: ttk.Treeview, status_key: str) -> str:
    tag = f"SCHEDULE::{status_key}"
    if tag in _TREE_STATUS_TAG_CACHE:
        return tag
    colors = SCHEDULE_STATUS_ROW_COLORS.get(status_key)
    if colors:
        tree.tag_configure(tag, **colors)
    _TREE_STATUS_TAG_CACHE[tag] = True
    return tag


def _describe_entry_status(
    entry: Dict[str, Any], *, today: Optional[dt.date] = None
) -> Tuple[str, str]:
    today = today or dt.date.today()
    status = str(entry.get("status") or "").strip().lower()
    date_obj = _parse_schedule_date(entry.get("date"))
    if status == "wykonany":
        label = "Wykonany"
        if date_obj:
            label = f"Wykonany {date_obj.isoformat()}"
        return label, "done"
    if date_obj is None:
        return "Brak daty", "none"
    days = (date_obj - today).days
    if days < 0:
        return f"Po terminie ({abs(days)} dni)", "overdue"
    if days <= SCHEDULE_SOON_THRESHOLD_DAYS:
        return f"Wkrótce ({days} dni)", "soon"
    return f"Planowany ({days} dni)", "ok"


def _serialize_schedule_entries(entries: Iterable[Dict[str, Any]]) -> List[Dict[str, Any]]:
    payload: List[Dict[str, Any]] = []
    for entry in entries or []:
        if not isinstance(entry, dict):
            continue
        clean: Dict[str, Any] = {
            "machine_id": entry.get("machine_id", ""),
            "machine_name": entry.get("machine_name", ""),
            "date": entry.get("date", ""),
            "type": entry.get("type", ""),
            "responsible": entry.get("responsible", ""),
            "notes": entry.get("notes", ""),
            "status": entry.get("status", "planowany"),
        }
        if entry.get("completed_at"):
            clean["completed_at"] = entry.get("completed_at")
        if entry.get("card"):
            clean["card"] = entry.get("card")
        payload.append(clean)
    return payload


def _resolve_card_storage(path: str, cfg: Any | None = None) -> str:
    normalized = os.path.normpath(path)
    try:
        if cfg is None:
            from config_manager import ConfigManager

            cfg = ConfigManager()
        data_root = cfg.path_data()
        project_root = cfg.path_root()
    except Exception:
        data_root = resolve_root_path("<root>", "data")
        project_root = resolve_root_path("<root>", "")

    for base, prefix in (
        (data_root, os.path.join("<root>", "data")),
        (project_root, "<root>"),
    ):
        try:
            if os.path.commonpath([normalized, base]) == os.path.normpath(base):
                rel = os.path.relpath(normalized, base)
                return os.path.join(prefix, rel).replace(os.sep, "/")
        except Exception:
            continue
    return normalized.replace(os.sep, "/")


def _resolve_card_absolute(path: str, cfg: Any | None = None) -> str:
    if not path:
        return ""
    try:
        if cfg is None:
            from config_manager import ConfigManager

            cfg = ConfigManager()
        root = cfg.path_root()
    except Exception:
        root = os.getcwd()
    return resolve_root_path(root, path)


def _open_external(path: str) -> bool:
    try:
        if os.name == "nt":
            os.startfile(path)  # type: ignore[attr-defined]
            return True
    except Exception:
        return False
    try:
        subprocess.Popen(["xdg-open", path])
        return True
    except Exception:
        try:
            subprocess.Popen(["open", path])
            return True
        except Exception:
            return False


_EXCEL_FIELD_ALIASES: Dict[str, Tuple[str, ...]] = {
    "machine_id": ("id", "nr", "nr ewid", "nr maszyny", "identyfikator"),
    "machine_name": ("maszyna", "nazwa", "nazwa maszyny"),
    "date": ("data", "termin", "termin przegl", "plan"),
    "type": ("typ", "rodzaj", "przegl", "rodzaj przegl"),
    "responsible": ("osoba", "odpowiedzial", "odpowiedzialny"),
    "notes": ("uwagi", "komentarz", "notat"),
    "status": ("status", "stan"),
    "completed_at": ("wykonano", "data wykonania", "zrealizowano"),
    "card": ("karta", "plik", "załącznik", "zalacznik"),
}


def _normalize_excel_header(value: object) -> str:
    text = str(value or "").strip().lower()
    text = text.replace("\n", " ").replace("\xa0", " ")
    return " ".join(text.split())


def _map_excel_headers(headers: List[str]) -> List[Optional[str]]:
    mapped: List[Optional[str]] = []
    for header in headers:
        normalized = _normalize_excel_header(header)
        field: Optional[str] = None
        for key, tokens in _EXCEL_FIELD_ALIASES.items():
            if any(token in normalized for token in tokens):
                field = key
                break
        mapped.append(field)
    return mapped


def _read_excel_schedule(path: str) -> List[Dict[str, Any]]:
    try:
        from openpyxl import load_workbook  # type: ignore
    except Exception as exc:  # pragma: no cover - optional dependency
        raise RuntimeError(
            "Brak biblioteki openpyxl – zainstaluj ją aby importować harmonogram."
        ) from exc

    wb = load_workbook(filename=path, data_only=True, read_only=True)
    sheet = wb.active
    rows_iter = sheet.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        return []

    headers = [_normalize_excel_header(cell) for cell in header_row]
    mapping = _map_excel_headers(headers)

    entries: List[Dict[str, Any]] = []
    for row in rows_iter:
        if not row:
            continue
        if all(cell is None or str(cell).strip() == "" for cell in row):
            continue
        raw: Dict[str, Any] = {}
        for idx, cell in enumerate(row):
            field = mapping[idx] if idx < len(mapping) else None
            if not field:
                continue
            if field in {"machine_id", "machine_name", "type", "responsible", "notes", "status", "card"}:
                raw[field] = str(cell or "").strip()
            else:
                raw[field] = cell
        normalized = _normalize_schedule_entry(raw)
        if normalized:
            entries.append(normalized)
    return entries


def _import_schedule_from_excel(path: str) -> Tuple[List[Dict[str, Any]], Dict[str, Any]]:
    entries = _read_excel_schedule(path)
    if not entries:
        raise ValueError("Nie znaleziono danych harmonogramu w pliku Excel.")

    years = {
        _parse_schedule_date(entry.get("date")).year
        for entry in entries
        if _parse_schedule_date(entry.get("date")) is not None
    }
    year = next(iter(years)) if len(years) == 1 else SCHEDULE_YEAR
    meta = {
        "year": int(year),
        "source": os.path.basename(path),
        "imported_at": dt.datetime.now().isoformat(),
    }
    return entries, meta
def _coalesce_data_root(cfg: dict | None = None) -> str:
    """Return an absolute data root derived from *cfg* or environment."""

    cfg = cfg or {}
    paths_cfg = cfg.get("paths") or {}
    candidates = (
        paths_cfg.get("data_root"),
        cfg.get("data_root"),
        os.environ.get("WM_DATA_ROOT"),
    )
    for candidate in candidates:
        if isinstance(candidate, str) and candidate.strip():
            value = os.path.expanduser(candidate.strip())
            if not os.path.isabs(value):
                value = os.path.normpath(os.path.join(os.getcwd(), value))
            return os.path.normpath(value)
    return os.path.normpath(os.path.join(os.getcwd(), "data"))


# ---- bezpieczny import config_manager / fallback (jak w Twojej bazie) ----
try:
    from config_manager import get_config, resolve_rel
except Exception:
    def resolve_rel(cfg: dict, what: str) -> str:
        root = _coalesce_data_root(cfg)
        mapping = {"machines": os.path.join("maszyny", "maszyny.json")}
        rel = mapping.get(what)
        if not rel:
            return root
        return os.path.normpath(os.path.join(root, rel))

    def get_config() -> dict:
        try:
            from config_manager import ConfigManager  # type: ignore
            return ConfigManager().load()
        except Exception:
            return {}

from ui_theme import ensure_theme_applied

# ---- funkcje z utils_maszyny (masz po R-06Jb) ----
from utils_maszyny import (
    load_machines_rows_with_fallback,
    ensure_machines_sample_if_empty,
    load_machines,
    upsert_machine,
    delete_machine,
    merge_rows_union_by_id,
    resolve_schedule_path,
)
def _iter_inspection_dates(machine: dict) -> list[dt.date]:
    """
    Zwraca listę poprawnych dat (datetime.date) z machine['zadania'].
    Ignoruje rekordy bez 'data' lub o błędnym formacie.
    """

    dates: list[dt.date] = []
    schedule_entries = machine.get("__schedule_entries") if isinstance(machine, dict) else None
    if schedule_entries:
        for entry in schedule_entries:
            status = str(entry.get("status") or "").strip().lower()
            if status == "wykonany":
                continue
            parsed = _parse_schedule_date(entry.get("date"))
            if parsed is not None:
                dates.append(parsed)
        return dates

    try:
        tasks = machine.get("zadania", []) or []
        for t in tasks:
            dstr = (t or {}).get("data")
            if not dstr or not isinstance(dstr, str):
                continue
            try:
                if len(dstr) == 7 and dstr.count("-") == 1:
                    d = dt.datetime.strptime(f"{dstr}-01", "%Y-%m-%d").date()
                else:
                    d = dt.datetime.strptime(dstr, "%Y-%m-%d").date()
                dates.append(d)
            except Exception:
                continue
    except Exception:
        logger.debug("[Maszyny] _iter_inspection_dates: ignoruję błędy formatu")
    return dates


def _next_inspection_date_safe(
    machine: dict, today: dt.date | None = None
) -> dt.date | None:
    """
    Zwraca NAJBLIŻSZĄ PRZYSZŁĄ datę przeglądu lub None, gdy nie istnieje.
    Nigdy nie rzuca ValueError przy pustej liście.
    """

    today = today or dt.date.today()
    dates = _iter_inspection_dates(machine)
    if not dates:
        return None
    future = [d for d in dates if d >= today]
    if not future:
        return None
    try:
        return min(future)
    except Exception:
        return None


def _days_to_next_inspection_safe(
    machine: dict, today: dt.date | None = None
) -> int | None:
    """
    Różnica dni do najbliższego przyszłego przeglądu; None gdy brak przyszłych terminów.
    """

    today = today or dt.date.today()
    nxt = _next_inspection_date_safe(machine, today=today)
    if nxt is None:
        return None
    try:
        return (nxt - today).days
    except Exception:
        return None


def _next_inspection_date(
    machine: dict, today: dt.date | None = None
) -> Optional[dt.date]:
    return _next_inspection_date_safe(machine, today=today)


def _days_to_next_inspection(machine: dict, today: dt.date | None = None) -> Optional[int]:
    return _days_to_next_inspection_safe(machine, today=today)


def _status_color(machine: dict) -> str:
    """Zwraca kolor statusu z uwzględnieniem liczby dni do przeglądu."""

    status = (machine.get("status") or "").strip().lower()
    if status in ("awaria", "uszkodzona", "error"):
        return "#dc2626"  # red-600

    summary = machine.get("__schedule_summary") if isinstance(machine, dict) else None
    if summary and summary.get("color"):
        return str(summary["color"])

    if status in ("ok", "sprawna", "sprawne"):
        return "#16a34a"  # green-600
    if status in ("uwaga", "serwis", "warning"):
        return "#ca8a04"  # yellow-600

    days = _days_to_next_inspection(machine)
    if days is None:
        return "#64748b"  # slate-500 – brak danych
    if days < 0:
        return "#dc2626"
    if days <= 30:
        return "#ca8a04"
    return "#16a34a"


def _map_label_text(machine: dict, label_mode: str) -> str:
    label_mode = (label_mode or "id").lower()
    if label_mode == "typ":
        return str(machine.get("typ", "") or "")
    if label_mode == "nazwa":
        return str(machine.get("nazwa", "") or "")
    return str(machine.get("id") or machine.get("nr_ewid") or "")


def _render_days_label_on_canvas(
    canvas, x: int, y: int, machine: dict
) -> Optional[int]:
    """
    Rysuje etykietę „dni do przeglądu” pod kropką maszyny.
    - '—' (neutralny szary) gdy brak przyszłych terminów (tylko przeszłe albo brak zadań).
    - czerwony, gdy days < 0 (spóźnione).
    """

    days_value = _days_to_next_inspection_safe(machine)
    if days_value is None:
        days_label = "—"
        fill_color = "#d1d5db"
    else:
        days_label = f"{days_value} dni"
        fill_color = "#d1d5db" if days_value >= 0 else "#fca5a5"
    try:
        return canvas.create_text(
            x, y, text=days_label, fill=fill_color, font=("Segoe UI", 8)
        )
    except Exception:
        logger.debug("[Maszyny] canvas text draw skipped (timing)")
        return None


class ImageTooltip:
    def __init__(self, parent: tk.Misc):
        self.parent = parent
        self.top: Optional[tk.Toplevel] = None
        self._img_ref = None

    def show(self, x: int, y: int, machine: dict) -> None:
        self.hide()
        self.top = tk.Toplevel(self.parent)
        self.top.wm_overrideredirect(True)
        self.top.attributes("-topmost", True)
        self.top.configure(bg="#111214")
        self.top.geometry(f"+{x + 12}+{y + 12}")

        path = machine.get("image") or machine.get("obraz") or ""
        name = machine.get("nazwa") or ""
        ttk.Label(self.top, text=name or "(bez nazwy)").pack(fill="x", padx=6, pady=(6, 2))

        if Image and ImageTk and path and os.path.exists(path):
            try:
                image = Image.open(path)
                image.thumbnail((256, 256))
                photo = ImageTk.PhotoImage(image)
                self._img_ref = photo
                tk.Label(self.top, image=photo, bd=0).pack(padx=6, pady=(0, 6))
            except Exception:
                ttk.Label(self.top, text="Nie można wczytać obrazu").pack(padx=6, pady=(0, 6))
        else:
            ttk.Label(self.top, text="Brak obrazu").pack(padx=6, pady=(0, 6))

        days = _days_to_next_inspection(machine)
        info = (
            f"Typ: {machine.get('typ', '') or '-'}   •   "
            f"Dni do przeglądu: {days if days is not None else '—'}"
        )
        ttk.Label(self.top, text=info).pack(padx=6, pady=(0, 6))

    def hide(self) -> None:
        if self.top is not None:
            self.top.destroy()
            self.top = None
            self._img_ref = None


class MonthYearDialog(tk.Toplevel):
    def __init__(self, master, title="Wybierz miesiąc i rok", init_year=None, init_month=None):
        super().__init__(master)
        self.title(title)
        self.resizable(False, False)
        self.result: Optional[str] = None
        yr = init_year or tk.IntVar(value=int(self._now("%Y")))
        mo = init_month or tk.StringVar(value=self._now("%m"))
        self._yr, self._mo = yr, mo

        body = ttk.Frame(self, padding=12)
        body.pack(fill="both", expand=True)

        ttk.Label(body, text="Rok:").grid(row=0, column=0, sticky="w", padx=(0, 6))
        self.spn_year = ttk.Spinbox(body, from_=2000, to=2100, width=6, textvariable=yr)
        self.spn_year.grid(row=0, column=1, sticky="w")

        ttk.Label(body, text="Miesiąc:").grid(row=1, column=0, sticky="w", padx=(0, 6), pady=(8, 0))
        self.cbo_month = ttk.Combobox(
            body,
            width=9,
            state="readonly",
            values=[f"{i:02d}" for i in range(1, 13)],
            textvariable=mo,
        )
        self.cbo_month.grid(row=1, column=1, sticky="w", pady=(8, 0))

        btns = ttk.Frame(body)
        btns.grid(row=2, column=0, columnspan=2, pady=(12, 0), sticky="e")
        ttk.Button(btns, text="Anuluj", command=self._on_cancel).pack(side="right")
        ttk.Button(btns, text="OK", command=self._on_ok).pack(side="right", padx=(0, 8))

        self.grab_set()
        self.transient(master)
        self.cbo_month.focus_set()

    def _now(self, fmt: str) -> str:
        import datetime as _dt

        return _dt.datetime.now().strftime(fmt)

    def _on_ok(self):
        y = int(self._yr.get())
        m = int(self._mo.get())
        self.result = f"{y:04d}-{m:02d}-01"
        self.destroy()

    def _on_cancel(self):
        self.result = None
        self.destroy()


def pick_machine_image(parent) -> Optional[str]:
    path = filedialog.askopenfilename(
        parent=parent,
        title="Wybierz zdjęcie maszyny",
        filetypes=[
            ("Obrazy", "*.jpg;*.jpeg;*.png;*.bmp;*.gif"),
            ("Wszystkie pliki", "*.*"),
        ],
    )
    return path or None


def _add_inspection_date(parent, machine: dict):
    dlg = MonthYearDialog(parent)
    parent.wait_window(dlg)
    if dlg.result:
        zadania = machine.setdefault("zadania", [])
        zadania.append({"data": dlg.result, "typ_zadania": "przegląd", "uwagi": ""})


def _set_machine_image(parent, machine: dict):
    path = pick_machine_image(parent)
    if not path:
        return
    machine["image"] = os.path.normpath(path)


def _build_edit_footer(panel, machine: dict, on_changed):
    footer = ttk.Frame(panel)
    footer.grid(row=7, column=0, columnspan=2, sticky="we", padx=6, pady=(8, 0))
    footer.columnconfigure(2, weight=1)

    info_var = tk.StringVar()

    def _summary() -> str:
        tasks = machine.get("zadania") or []
        image_path = machine.get("image") or machine.get("obraz") or ""
        parts = [
            f"Przeglądy: {len(tasks)}" if tasks else "Przeglądy: brak",
            f"Zdjęcie: {os.path.basename(image_path)}" if image_path else "Zdjęcie: brak",
        ]
        return " | ".join(parts)

    def _trigger_changed() -> None:
        info_var.set(_summary())
        if callable(on_changed):
            on_changed()

    def _handle_add():
        before = list(machine.get("zadania") or [])
        _add_inspection_date(panel, machine)
        after = machine.get("zadania") or []
        if after != before:
            _trigger_changed()

    def _handle_image():
        before = machine.get("image") or machine.get("obraz")
        _set_machine_image(panel, machine)
        after = machine.get("image") or machine.get("obraz")
        if after != before:
            _trigger_changed()

    ttk.Button(
        footer,
        text="Dodaj przegląd (miesiąc/rok)",
        command=_handle_add,
    ).grid(row=0, column=0, sticky="w")
    ttk.Button(
        footer,
        text="Ustaw zdjęcie…",
        command=_handle_image,
    ).grid(row=0, column=1, sticky="w", padx=(8, 0))
    ttk.Label(footer, textvariable=info_var).grid(row=0, column=2, sticky="w", padx=(12, 0))

    _trigger_changed()
    return footer


def _save_machines(primary_path: str, rows: list[dict]) -> bool:
    try:
        payload = {"maszyny": rows}
        _safe_write_json(primary_path, payload)
        logger.info("[Maszyny] Zapisano %d rekordów -> %s", len(rows), primary_path)
        return True
    except Exception:
        logger.exception("[Maszyny] Błąd zapisu danych maszyn (%s)", primary_path)
        return False


_TREE_COLUMN_LAYOUT: Tuple[Tuple[str, str, int, str], ...] = (
    ("id", "ID", 110, "w"),
    ("nazwa", "Nazwa", 220, "w"),
    ("typ", "Typ", 120, "w"),
    ("status", "Status", 90, "center"),
    ("przeglad", "Najbliższy przegląd", 190, "center"),
    ("przeglad_status", "Status przeglądu", 150, "center"),
    ("dni", "Dni", 70, "center"),
)


def _ensure_tree_columns(tree: ttk.Treeview) -> None:
    columns = [cid for cid, _, _, _ in _TREE_COLUMN_LAYOUT]
    tree.configure(columns=columns)
    for cid, label, width, anchor in _TREE_COLUMN_LAYOUT:
        tree.heading(cid, text=label)
        tree.column(cid, width=width, anchor=anchor)


def _tree_insert_row(tree: ttk.Treeview, machine: dict) -> str:
    columns = list(tree["columns"]) if tree["columns"] else []
    summary = machine.get("__schedule_summary") if isinstance(machine, dict) else {}
    next_label = (summary or {}).get("next_label") or "—"
    status_label = (summary or {}).get("status_label") or "Brak danych"
    days = (summary or {}).get("days")

    def _value_for(col: str) -> str | int:
        if col == "dni":
            return days if days is not None else "—"
        if col == "przeglad":
            return next_label
        if col == "przeglad_status":
            return status_label
        if col == "id":
            return machine.get("id", "") or machine.get("nr_ewid", "") or ""
        return machine.get(col, "")

    values = tuple(_value_for(col) for col in columns)
    identifier = str(machine.get("id") or machine.get("nr_ewid") or "")
    item_id = tree.insert("", "end", iid=identifier or None, values=values)
    tag_identifier = identifier or item_id
    schedule_key = _schedule_status_key(machine)
    schedule_tag = _ensure_tree_schedule_tag(tree, schedule_key)
    tags = [f"ROW::{tag_identifier}"]
    if schedule_tag:
        tags.append(schedule_tag)
    tree.item(item_id, tags=tuple(tags))
    return item_id


def _bind_tree_tooltips(tree: ttk.Treeview, rows: list[dict], root_widget: tk.Misc) -> None:
    tip = ImageTooltip(root_widget)
    id_by_item: dict[str, dict] = {}
    for item in tree.get_children(""):
        tags = tree.item(item, "tags") or ()
        tag = next((t for t in tags if t.startswith("ROW::")), None)
        if not tag:
            continue
        row_id = tag.split("::", 1)[1]
        machine = next(
            (m for m in rows if str(m.get("id") or m.get("nr_ewid") or "") == row_id),
            None,
        )
        if machine is None:
            values = tree.item(item, "values")
            if values:
                fallback_id = str(values[0])
                machine = next(
                    (
                        m
                        for m in rows
                        if str(m.get("id") or m.get("nr_ewid") or "") == fallback_id
                    ),
                    None,
                )
        if machine:
            id_by_item[item] = machine

    def _on_motion(event: tk.Event) -> None:
        item = tree.identify_row(event.y)
        if item and item in id_by_item:
            tip.show(event.x_root, event.y_root, id_by_item[item])
        else:
            tip.hide()

    def _on_leave(_event: tk.Event) -> None:
        tip.hide()

    tree.bind("<Motion>", _on_motion)
    tree.bind("<Leave>", _on_leave)


# ============================================================
# Renderer hali: kropki z etykietą w środku, drag&drop, tooltip z miniaturą
# ============================================================
class MachineHallRenderer:
    COLORS = {
        "SELECTED": "#60a5fa",
        "_": "#1f2937",
        "OK": "#16a34a",
        "WARN": "#ca8a04",
        "ALERT": "#dc2626",
    }
    RADIUS = 18

    def __init__(
        self,
        parent: tk.Misc,
        rows: List[Dict],
        cfg: dict | None = None,
        on_drag_commit=None,
        bg_path: str | None = None,
    ):
        self.parent = parent
        self.rows = rows or []
        self.cfg = cfg or {}
        self.on_drag_commit = on_drag_commit
        self.canvas = tk.Canvas(
            parent,
            bg=DEFAULT_BG_COLOR,
            highlightthickness=0,
            width=CANVAS_W,
            height=CANVAS_H,
        )
        self.nodes_by_id: dict[str, int] = {}
        self.text_by_id: dict[str, int] = {}
        self.labels_by_id: dict[str, int] = {}
        self.rows_by_id: dict[str, dict] = {}
        self.selected_id: Optional[str] = None
        self._drag_active = False
        self._drag_id: Optional[str] = None
        self._offset: tuple[int, int] = (0, 0)
        self._bg_path = bg_path
        self._bg_image_path: Optional[str] = None
        self._bg_img_pil: Optional["Image.Image"] = None
        self._bg_img_tk: Optional["ImageTk.PhotoImage"] = None
        self._bg_fallback: Optional[tk.PhotoImage] = None
        self._bg_w = 0
        self._bg_h = 0
        self._bg_anchor_xy: Tuple[int, int] = (0, 0)
        self._scale_x = 1.0
        self._scale_y = 1.0
        self.tip = ImageTooltip(parent)
        self.items_meta: dict[int, dict] = {}
        self._current_radius = self.RADIUS

    # ---------- public ----------
    def render(self) -> None:
        self.canvas.config(width=CANVAS_W, height=CANVAS_H)
        self.canvas.pack(fill="none", expand=False, padx=8, pady=8)
        self._load_background()
        self._draw_all()
        self._bind_drag()
        self.canvas.bind("<Motion>", self._on_canvas_motion, add="+")
        self.canvas.bind("<Leave>", lambda _e: self.tip.hide(), add="+")

    def select(self, machine_id: str | None):
        self.selected_id = machine_id
        self._redraw_selection()

    def update_rows(self, rows: List[Dict]):
        self.rows = rows or []
        self._draw_all()

    # ---------- internals ----------
    def _load_background(self):
        self._reset_background_state()

        path = None
        machines_cfg = self.cfg.get("machines") if isinstance(self.cfg, dict) else None
        if isinstance(machines_cfg, dict):
            candidate = machines_cfg.get("background_image")
            if isinstance(candidate, str) and candidate.strip():
                path = candidate
        if not path:
            cfg_paths = (self.cfg.get("paths", {}) or {})
            cfg_bg = (cfg_paths.get("hall", {}) or {}).get("background_image") or cfg_paths.get(
                "hall.background_image"
            )
            if isinstance(cfg_bg, str) and cfg_bg.strip():
                path = cfg_bg
        if not path and isinstance(self._bg_path, str):
            path = self._bg_path
        if not path:
            return
        if not os.path.isabs(path):
            cfg_context = self.cfg if isinstance(self.cfg, dict) else {}
            root = _coalesce_data_root(cfg_context)
            path = os.path.join(root, path)
        path = os.path.normpath(path)
        if not os.path.exists(path):
            logger.info("[Maszyny][HALL] Tło nie istnieje: %s", path)
            return

        self._bg_image_path = path
        self._load_bg_image_assets(path)

    def _reset_background_state(self) -> None:
        self._bg_image_path = None
        self._bg_img_pil = None
        self._bg_img_tk = None
        self._bg_fallback = None
        self._bg_w = 0
        self._bg_h = 0
        self._bg_anchor_xy = (0, 0)
        self._scale_x = 1.0
        self._scale_y = 1.0

    def _set_bg_geometry(self, width: int, height: int) -> None:
        self._bg_w = max(0, int(width))
        self._bg_h = max(0, int(height))
        off_x = max(0, (CANVAS_W - self._bg_w) // 2)
        off_y = max(0, (CANVAS_H - self._bg_h) // 2)
        self._bg_anchor_xy = (off_x, off_y)
        # Tło nie jest skalowane – współczynniki 1.0 pozwalają mapować px->canvas.
        self._scale_x = 1.0
        self._scale_y = 1.0

    def _load_bg_image_assets(self, path: str) -> None:
        if Image and ImageTk:
            try:
                img = Image.open(path)
            except Exception:
                img = None
            else:
                self._bg_img_pil = img
                self._set_bg_geometry(img.width, img.height)
                try:
                    self._bg_img_tk = ImageTk.PhotoImage(img)
                except Exception:
                    self._bg_img_tk = None
        if self._bg_img_tk is not None:
            self._bg_fallback = None
            return
        try:
            tk_img = tk.PhotoImage(file=path)
        except Exception:
            self._bg_fallback = None
            self._set_bg_geometry(0, 0)
            return
        self._bg_fallback = tk_img
        try:
            width = int(tk_img.width())
            height = int(tk_img.height())
        except Exception:
            width = height = 0
        self._set_bg_geometry(width, height)

    def _draw_background_and_grid(self) -> None:
        self.canvas.config(width=CANVAS_W, height=CANVAS_H)
        self.canvas.create_rectangle(
            0,
            0,
            CANVAS_W,
            CANVAS_H,
            fill=DEFAULT_BG_COLOR,
            outline="",
        )

        ax, ay = self._bg_anchor_xy
        if self._bg_img_tk is not None:
            self.canvas.create_image(ax, ay, image=self._bg_img_tk, anchor="nw")
        elif self._bg_fallback is not None:
            self.canvas.create_image(ax, ay, image=self._bg_fallback, anchor="nw")

        if self._bg_w > 0 and self._bg_h > 0:
            step_x = max(1, int(GRID_BASE_BG_PX_X * self._scale_x))
            step_y = max(1, int(GRID_BASE_BG_PX_Y * self._scale_y))

            x = ax
            while x <= ax + self._bg_w:
                self.canvas.create_line(x, ay, x, ay + self._bg_h, fill="#2a2a2a")
                x += step_x

            y = ay
            while y <= ay + self._bg_h:
                self.canvas.create_line(ax, y, ax + self._bg_w, y, fill="#2a2a2a")
                y += step_y

            self.canvas.create_rectangle(
                ax,
                ay,
                ax + self._bg_w,
                ay + self._bg_h,
                outline="#3a3a3a",
            )

    def _map_bg_to_canvas(self, x_bg: int, y_bg: int) -> Tuple[int, int]:
        ax, ay = self._bg_anchor_xy
        cx = int(ax + x_bg * self._scale_x)
        cy = int(ay + y_bg * self._scale_y)
        return self._clamp_to_canvas(cx, cy)

    def _map_canvas_to_bg(self, x_canvas: int, y_canvas: int) -> Tuple[int, int]:
        if self._bg_w > 0 and self._bg_h > 0:
            ax, ay = self._bg_anchor_xy
            try:
                bx = int(round((x_canvas - ax) / self._scale_x))
                by = int(round((y_canvas - ay) / self._scale_y))
            except ZeroDivisionError:
                bx, by = x_canvas, y_canvas
            bx = _safe_clamp(bx, 0, max(0, self._bg_w - 1))
            by = _safe_clamp(by, 0, max(0, self._bg_h - 1))
            return bx, by
        return x_canvas, y_canvas

    def _clamp_to_canvas(self, x: int, y: int) -> Tuple[int, int]:
        clamped_x = _safe_clamp(int(x), 0, CANVAS_W - 1)
        clamped_y = _safe_clamp(int(y), 0, CANVAS_H - 1)
        return clamped_x, clamped_y

    def _status_color(self, status):
        key = str(status or "").strip().upper()
        return self.COLORS.get(key, self.COLORS["_"])

    def _node_center(self, r: Dict, idx: int, radius: int) -> tuple[int, int]:
        cols, cell_w, cell_h, pad_x, pad_y = 6, 120, 110, 70, 70
        x, y = r.get("x"), r.get("y")
        if isinstance(x, int) and isinstance(y, int):
            margin = max(radius, 4)
            if self._bg_w > 0 and self._bg_h > 0:
                cx, cy = self._map_bg_to_canvas(x, y)
                ax, ay = self._bg_anchor_xy
                min_x = ax + margin
                max_x = ax + self._bg_w - margin
                min_y = ay + margin
                max_y = ay + self._bg_h - margin
                if max_x < min_x:
                    min_x, max_x = margin, max(margin, CANVAS_W - margin)
                if max_y < min_y:
                    min_y, max_y = margin, max(margin, CANVAS_H - margin)
            else:
                width, height = _canvas_bounds(self.canvas)
                min_x = margin
                min_y = margin
                max_x = max(margin, width - margin)
                max_y = max(margin, height - margin)
                cx, cy = x, y
            return (
                _safe_clamp(int(cx), int(min_x), int(max_x)),
                _safe_clamp(int(cy), int(min_y), int(max_y)),
            )
        gx, gy = idx % cols, idx // cols
        cx, cy = pad_x + gx * cell_w, pad_y + gy * cell_h
        width, height = _canvas_bounds(self.canvas)
        margin = max(radius, 4)
        max_x = max(margin, width - margin)
        max_y = max(margin, height - margin)
        return (
            _safe_clamp(int(cx), margin, max_x),
            _safe_clamp(int(cy), margin, max_y),
        )

    def _short_id(self, mid: str) -> str:
        mid = (mid or "").strip()
        if len(mid) <= 5:
            return mid
        # skróć typu ABC-1234 -> A-1234
        parts = mid.split("-")
        if len(parts) >= 2 and parts[0]:
            return f"{parts[0][0]}-{parts[-1][:4]}"
        return mid[:5]

    def _draw_all(self):
        self.canvas.delete("all")
        self.nodes_by_id.clear()
        self.text_by_id.clear()
        self.labels_by_id.clear()
        self.rows_by_id.clear()
        self.items_meta.clear()

        radius = self._resolve_radius()
        self._current_radius = radius

        self._draw_background_and_grid()

        label_mode = self._label_mode()
        for idx, r in enumerate(self.rows):
            if not isinstance(r, dict):
                continue
            mid = str(r.get("id") or r.get("nr_ewid") or f"row{idx}")
            self.rows_by_id[mid] = r

            cx, cy = self._node_center(r, idx, radius)
            node = self.canvas.create_oval(
                cx - radius,
                cy - radius,
                cx + radius,
                cy + radius,
                fill=_status_color(r),
                outline="#0b0c0f",
                width=1,
            )
            self.nodes_by_id[mid] = node
            self.items_meta[node] = r

            label_text = _map_label_text(r, label_mode).strip()
            label_text = label_text[:6]
            font_size = max(8, min(14, int(radius * 0.55)))
            text_id = self.canvas.create_text(
                cx,
                cy,
                text=label_text,
                fill="#0b0c0f",
                font=("TkDefaultFont", font_size, "bold"),
            )
            self.text_by_id[mid] = text_id
            self.items_meta[text_id] = r

            label_id = _render_days_label_on_canvas(
                self.canvas, cx, cy + radius + 14, r
            )
            if label_id is not None:
                self.labels_by_id[mid] = label_id
                self.items_meta[label_id] = r

        self._redraw_selection()

    def _redraw_selection(self):
        for mid, node in self.nodes_by_id.items():
            sel = mid == self.selected_id
            try:
                self.canvas.itemconfigure(
                    node,
                    outline=self.COLORS["SELECTED"] if sel else "#0b0c0f",
                    width=3 if sel else 1,
                )
            except Exception:
                pass

    def _find_node_at(self, x: int, y: int) -> Optional[str]:
        radius = self._current_radius or self.RADIUS
        items = self.canvas.find_overlapping(
            x - radius,
            y - radius,
            x + radius,
            y + radius,
        )
        inv = {v: k for k, v in self.nodes_by_id.items()}
        for it in items:
            if it in inv:
                return inv[it]
        return None

    def _bind_drag(self):
        self.canvas.bind("<Button-1>", self._on_press, add="+")
        self.canvas.bind("<B1-Motion>", self._on_motion, add="+")
        self.canvas.bind("<ButtonRelease-1>", self._on_release, add="+")

    def _on_press(self, ev):
        mid = self._find_node_at(ev.x, ev.y)
        self._drag_active = bool(mid)
        self._drag_id = mid
        if mid:
            self.select(mid)
            node = self.nodes_by_id.get(mid)
            if node:
                x1, y1, x2, y2 = self.canvas.coords(node)
                cx = (x1 + x2) / 2
                cy = (y1 + y2) / 2
                self._offset = (int(cx - ev.x), int(cy - ev.y))
        else:
            self._offset = (0, 0)

    def _move_group(self, mid: str, cx: int, cy: int):
        node = self.nodes_by_id.get(mid)
        radius = self._current_radius or self.RADIUS
        margin = max(radius, 4)
        if self._bg_w > 0 and self._bg_h > 0:
            ax, ay = self._bg_anchor_xy
            min_x = ax + margin
            min_y = ay + margin
            max_x = ax + self._bg_w - margin
            max_y = ay + self._bg_h - margin
            if max_x < min_x:
                width, height = _canvas_bounds(self.canvas)
                min_x = margin
                max_x = max(margin, width - margin)
            if max_y < min_y:
                width, height = _canvas_bounds(self.canvas)
                min_y = margin
                max_y = max(margin, height - margin)
        else:
            width, height = _canvas_bounds(self.canvas)
            min_x = margin
            min_y = margin
            max_x = max(margin, width - margin)
            max_y = max(margin, height - margin)
        cx = _safe_clamp(cx, int(min_x), int(max_x))
        cy = _safe_clamp(cy, int(min_y), int(max_y))
        if node:
            self.canvas.coords(
                node,
                cx - radius,
                cy - radius,
                cx + radius,
                cy + radius,
            )
        t = self.text_by_id.get(mid)
        if t:
            self.canvas.coords(t, cx, cy)
        lab = self.labels_by_id.get(mid)
        if lab:
            self.canvas.coords(lab, cx, cy + (radius + 14))

    def _on_motion(self, ev):
        if not self._drag_active or not self._drag_id:
            return
        self.tip.hide()
        offx, offy = self._offset
        cx, cy = ev.x + offx, ev.y + offy
        grid = 10
        cx = int(round(cx / grid) * grid)
        cy = int(round(cy / grid) * grid)
        self._move_group(self._drag_id, cx, cy)

    def _on_release(self, ev):
        if not self._drag_active or not self._drag_id:
            return
        node = self.nodes_by_id.get(self._drag_id)
        if not node:
            self._drag_active = False
            self._drag_id = None
            return
        x1, y1, x2, y2 = self.canvas.coords(node)
        cx, cy = int((x1 + x2) / 2), int((y1 + y2) / 2)
        bx, by = self._map_canvas_to_bg(cx, cy)

        # aktualizuj model i zapisz
        r = self.rows_by_id.get(self._drag_id)
        if r is not None:
            r["x"], r["y"] = bx, by
        if callable(self.on_drag_commit):
            try:
                self.on_drag_commit(self._drag_id, bx, by)
            except Exception:
                logger.exception("[Maszyny][HALL] Błąd zapisu po drag&drop")

        self._drag_active = False
        self._drag_id = None

    def _on_canvas_motion(self, event: tk.Event) -> None:
        if self._drag_active:
            return
        current = self.canvas.find_withtag("current")
        if not current:
            self.tip.hide()
            return
        item_id = current[0]
        machine = self.items_meta.get(item_id)
        if machine is None:
            pair = self._pair(item_id)
            if pair is not None:
                machine = self.items_meta.get(pair)
        if machine:
            self.tip.show(event.x_root, event.y_root, machine)
        else:
            self.tip.hide()

    def _pair(self, item_id: int) -> Optional[int]:
        bbox = self.canvas.bbox(item_id)
        if not bbox:
            return None
        x1, y1, x2, y2 = bbox
        nearby = self.canvas.find_overlapping(x1 - 2, y1 - 2, x2 + 2, y2 + 2)
        for candidate in nearby:
            if candidate != item_id and candidate in self.items_meta:
                return candidate
        return None

    def _label_mode(self) -> str:
        machines_cfg = self.cfg.get("machines") if isinstance(self.cfg, dict) else {}
        mode = "id"
        if isinstance(machines_cfg, dict):
            raw = (machines_cfg.get("map_label") or "id").strip().lower()
            if raw in {"id", "typ", "nazwa"}:
                mode = raw
        return mode

    def _resolve_radius(self) -> int:
        machines_cfg = self.cfg.get("machines") if isinstance(self.cfg, dict) else {}
        if isinstance(machines_cfg, dict):
            try:
                radius = int(machines_cfg.get("map_dot_radius") or self.RADIUS)
            except Exception:
                radius = self.RADIUS
        else:
            radius = self.RADIUS
        radius = radius or self.RADIUS
        return max(10, min(60, radius))

# ============================================================
# Reszta panelu — zostaje jak w Twojej wersji R-06Jc,
# poniżej fragmenty, które muszą zapewnić domyślne nr_hali=1
# przy edycji/zapisie oraz integrację z rendererem.
# ============================================================
def _build_tree(parent: tk.Misc, rows: List[Dict]) -> ttk.Treeview:
    tree = ttk.Treeview(
        parent,
        columns=tuple(cid for cid, _, _, _ in _TREE_COLUMN_LAYOUT),
        show="headings",
        height=18,
    )
    _ensure_tree_columns(tree)
    for r in rows:
        _tree_insert_row(tree, r)
    tree.pack(fill="both", expand=True, padx=8, pady=(0, 8))
    return tree


def _detect_real_source(rows_from_fallback: List[Dict], primary_path: str, cfg: dict) -> str:
    primary_rows, _ = load_machines(primary_path)
    if rows_from_fallback and not primary_rows:
        legacy_path = resolve_rel(cfg, r"maszyny.json")
        legacy_rows, _ = load_machines(legacy_path)
        if legacy_rows:
            return legacy_path
    return primary_path


def _open_machines_panel(root: tk.Misc, container: tk.Misc, Renderer=None):
    for child in container.winfo_children():
        child.destroy()

    paned = ttk.Panedwindow(container, orient="horizontal")
    paned.pack(fill="both", expand=True)
    left, right = ttk.Frame(paned), ttk.Frame(paned)
    paned.add(left, weight=1)
    paned.add(right, weight=1)

    toolbar = ttk.Frame(left)
    toolbar.pack(fill="x", padx=8, pady=(8, 0))
    info = tk.StringVar(value="Maszyny")
    ttk.Label(toolbar, textvariable=info).pack(side="left")

    filter_var = tk.StringVar(value="Wszystkie")
    ttk.Label(toolbar, text="Filtr:").pack(side="left", padx=(12, 4))
    filter_box = ttk.Combobox(
        toolbar,
        state="readonly",
        width=14,
        values=("Wszystkie", "Po terminie", "Wkrótce", "Planowane", "Wykonane"),
        textvariable=filter_var,
    )
    filter_box.pack(side="left")

    btn_import = ttk.Button(toolbar, text="Importuj harmonogram…")
    btn_import.pack(side="left", padx=(12, 0))

    btn_add, btn_edit, btn_del, btn_save = (ttk.Button(toolbar, text=t) for t in ("Dodaj", "Edytuj", "Usuń", "Zapisz"))
    for button in (btn_save, btn_del, btn_edit, btn_add):
        button.pack(side="right", padx=4)

    schedule_info = tk.StringVar(value="")
    ttk.Label(left, textvariable=schedule_info).pack(fill="x", padx=8, pady=(4, 4))

    cfg: Dict[str, Any] = {}
    try:
        cfg = get_config()
    except Exception:
        logger.exception("[Maszyny] Nie udało się wczytać konfiguracji.")

    cfg_manager = None
    try:
        from config_manager import ConfigManager

        cfg_manager = ConfigManager()
    except Exception:
        cfg_manager = None

    rows, primary_path = load_machines_rows_with_fallback(cfg, resolve_rel)
    had_rows = bool(rows)
    rows = ensure_machines_sample_if_empty(rows, primary_path)
    source_path = _detect_real_source(rows, primary_path, cfg)
    rows_cache: List[Dict] = list(rows)

    schedule_year = SCHEDULE_YEAR
    schedule_path = resolve_schedule_path(schedule_year, cfg_manager)
    schedule_payload = _safe_read_json(schedule_path, default={})
    schedule_meta: Dict[str, Any] = {}
    raw_schedule_entries: List[Dict[str, Any]] = []
    if isinstance(schedule_payload, dict):
        schedule_meta = {k: v for k, v in schedule_payload.items() if k != "entries"}
        raw_schedule_entries = [
            entry for entry in schedule_payload.get("entries", []) if isinstance(entry, dict)
        ]
        if schedule_meta.get("year"):
            try:
                schedule_year = int(schedule_meta["year"])
            except Exception:
                schedule_year = SCHEDULE_YEAR
    elif isinstance(schedule_payload, list):
        raw_schedule_entries = [entry for entry in schedule_payload if isinstance(entry, dict)]
        schedule_meta = {"year": schedule_year}
    else:
        schedule_meta = {"year": schedule_year}
    if schedule_year != SCHEDULE_YEAR:
        schedule_path = resolve_schedule_path(schedule_year, cfg_manager)

    schedule_entries: List[Dict[str, Any]] = []
    for raw_entry in raw_schedule_entries:
        entry = _normalize_schedule_entry(raw_entry)
        if not entry:
            continue
        if raw_entry.get("card"):
            entry["card"] = str(raw_entry.get("card")).strip()
        if raw_entry.get("completed_at") and "completed_at" not in entry:
            entry["completed_at"] = str(raw_entry.get("completed_at"))
        schedule_entries.append(entry)

    schedule_meta.setdefault("year", schedule_year)
    if schedule_meta.get("source") is None:
        schedule_meta["source"] = (
            os.path.basename(schedule_path) if os.path.exists(schedule_path) else ""
        )
    schedule_meta.setdefault("imported_at", schedule_meta.get("imported_at"))
    schedule_meta.setdefault("updated_at", schedule_meta.get("updated_at"))

    _attach_schedule(rows_cache, schedule_entries)
    visible_rows: List[Dict] = list(rows_cache)

    info.set(
        f"Wczytano {len(rows_cache)} maszyn." if had_rows else "Brak danych – dodano przykładowe pozycje."
    )

    tree = _build_tree(left, visible_rows)
    _bind_tree_tooltips(tree, visible_rows, root)

    selected_machine_id: Optional[str] = None
    hall: MachineHallRenderer | None = None
    upcoming_items: Dict[str, Dict[str, Any]] = {}
    history_items: Dict[str, Dict[str, Any]] = {}

    def _refresh_schedule_info() -> None:
        year = schedule_meta.get("year", schedule_year)
        if schedule_entries:
            parts = [f"Harmonogram {year}: {len(schedule_entries)} wpisów"]
        else:
            parts = [f"Harmonogram {year}: brak danych"]
        source = schedule_meta.get("source")
        if source:
            parts.append(f"Źródło: {source}")
        imported = schedule_meta.get("imported_at")
        if imported:
            parts.append(f"Import: {imported}")
        schedule_info.set(" • ".join(parts))

    def _update_info() -> None:
        info.set(f"Wczytano {len(rows_cache)} maszyn • widocznych: {len(visible_rows)}")

    def _recompute_visible_rows() -> None:
        nonlocal visible_rows
        mode = filter_var.get()

        def predicate(machine: Dict[str, Any]) -> bool:
            key = _schedule_status_key(machine)
            if mode == "Po terminie":
                return key == "overdue"
            if mode == "Wkrótce":
                return key == "soon"
            if mode == "Planowane":
                return key == "ok"
            if mode == "Wykonane":
                return key == "done"
            return True

        visible_rows = [row for row in rows_cache if predicate(row)]

    def _find_machine(machine_id: Optional[str]) -> Optional[Dict]:
        if not machine_id:
            return None
        return next(
            (
                row
                for row in rows_cache
                if str(row.get("id") or row.get("nr_ewid") or "") == machine_id
            ),
            None,
        )

    def _save_schedule_state() -> None:
        nonlocal schedule_path, schedule_year
        try:
            schedule_year = int(schedule_meta.get("year", schedule_year) or schedule_year)
        except Exception:
            schedule_year = SCHEDULE_YEAR
        schedule_path = resolve_schedule_path(schedule_year, cfg_manager)
        payload = dict(schedule_meta)
        payload["year"] = schedule_year
        payload["entries"] = _serialize_schedule_entries(schedule_entries)
        _safe_write_json(schedule_path, payload)

    def _refresh_tree() -> None:
        tree.delete(*tree.get_children())
        _ensure_tree_columns(tree)
        for row in visible_rows:
            _tree_insert_row(tree, row)
        _bind_tree_tooltips(tree, visible_rows, root)
        visible_ids = {
            str(row.get("id") or row.get("nr_ewid") or "")
            for row in visible_rows
        }
        if selected_machine_id and selected_machine_id in visible_ids:
            try:
                tree.selection_set(selected_machine_id)
            except Exception:
                tree.selection_remove(tree.selection())
        else:
            tree.selection_remove(tree.selection())
        _update_info()

    def _reload_from(path: str) -> List[Dict]:
        payload = _safe_read_json(path, default=[])
        return normalize_doc_list_or_dict(payload, "maszyny", fallback_keys=("machines",))

    def _on_rows_changed() -> None:
        _attach_schedule(rows_cache, schedule_entries)
        _recompute_visible_rows()
        _refresh_tree()
        if hall is not None:
            hall.update_rows(rows_cache)
        if selected_machine_id:
            machine = _find_machine(selected_machine_id)
            _populate_details(machine)
        else:
            _populate_details(None)

    def _on_schedule_changed(save: bool = True) -> None:
        schedule_meta.setdefault("year", schedule_year)
        schedule_meta["updated_at"] = dt.datetime.now().isoformat()
        _on_rows_changed()
        _refresh_schedule_info()
        if save:
            _save_schedule_state()

    def _save_rows(rows_to_save: List[Dict]) -> List[Dict]:
        nonlocal source_path
        cleaned = [_strip_schedule_fields(row) for row in rows_to_save]
        if os.path.normpath(source_path) != os.path.normpath(primary_path):
            legacy_rows = _reload_from(source_path)
            prim_rows = _reload_from(primary_path)
            merged = merge_rows_union_by_id(prim_rows, legacy_rows)
            merged = merge_rows_union_by_id(merged, cleaned)
            if _save_machines(primary_path, merged):
                source_path = primary_path
                _attach_schedule(merged, schedule_entries)
                return merged
            return rows_to_save
        if _save_machines(primary_path, cleaned):
            _attach_schedule(cleaned, schedule_entries)
            return cleaned
        return rows_to_save

    def _apply_filter(*_args) -> None:
        _recompute_visible_rows()
        _refresh_tree()
        if selected_machine_id and not any(
            str(row.get("id") or row.get("nr_ewid") or "") == selected_machine_id
            for row in visible_rows
        ):
            _set_selected_machine(None)
        else:
            if selected_machine_id:
                _populate_details(_find_machine(selected_machine_id))
            else:
                _populate_details(None)

    def _drag_commit(mid: str, x: int, y: int) -> None:
        nonlocal rows_cache
        update = None
        for row in rows_cache:
            if str(row.get("id") or row.get("nr_ewid") or "") == mid:
                update = dict(row)
                update["x"], update["y"] = x, y
                if "nr_hali" not in update or update.get("nr_hali") in (None, ""):
                    update["nr_hali"] = "1"
                break
        if update is None:
            return
        new_rows = upsert_machine(rows_cache, update)
        persisted = _save_rows(new_rows)
        rows_cache = list(persisted)
        _on_rows_changed()

    hall = MachineHallRenderer(right, rows_cache, cfg=cfg, on_drag_commit=_drag_commit)
    hall.render()

    details = ttk.LabelFrame(right, text="Przeglądy")
    details.pack(fill="both", expand=True, padx=8, pady=(0, 8))
    summary_var = tk.StringVar(value="Wybierz maszynę, aby zobaczyć harmonogram.")
    ttk.Label(details, textvariable=summary_var).pack(fill="x", padx=8, pady=(6, 4))

    columns_details = ("data", "typ", "status", "uwagi")
    column_setup = {
        "data": ("Data", 110, "center"),
        "typ": ("Typ", 120, "center"),
        "status": ("Status", 160, "w"),
        "uwagi": ("Uwagi", 220, "w"),
    }

    upcoming_section = ttk.LabelFrame(details, text="Zaplanowane")
    upcoming_section.pack(fill="both", expand=True, padx=8, pady=(0, 8))
    upcoming_tree = ttk.Treeview(upcoming_section, columns=columns_details, show="headings", height=6)
    for cid, (label, width, anchor) in column_setup.items():
        upcoming_tree.heading(cid, text=label)
        upcoming_tree.column(cid, width=width, anchor=anchor)
    upcoming_tree.pack(fill="both", expand=True, padx=6, pady=(0, 4))

    upcoming_buttons = ttk.Frame(upcoming_section)
    upcoming_buttons.pack(fill="x", padx=6, pady=(0, 4))
    btn_mark_done = ttk.Button(upcoming_buttons, text="Oznacz jako wykonany", state="disabled")
    btn_mark_done.pack(side="left")
    btn_assign_card = ttk.Button(upcoming_buttons, text="Przypisz kartę…", state="disabled")
    btn_assign_card.pack(side="left", padx=(6, 0))
    btn_open_card = ttk.Button(upcoming_buttons, text="Otwórz kartę", state="disabled")
    btn_open_card.pack(side="left", padx=(6, 0))

    history_section = ttk.LabelFrame(details, text="Historia")
    history_section.pack(fill="both", expand=True, padx=8, pady=(0, 8))
    history_tree = ttk.Treeview(history_section, columns=columns_details, show="headings", height=5)
    for cid, (label, width, anchor) in column_setup.items():
        history_tree.heading(cid, text=label)
        history_tree.column(cid, width=width, anchor=anchor)
    history_tree.pack(fill="both", expand=True, padx=6, pady=(0, 4))

    history_buttons = ttk.Frame(history_section)
    history_buttons.pack(fill="x", padx=6, pady=(0, 4))
    btn_restore = ttk.Button(history_buttons, text="Przywróć jako planowany", state="disabled")
    btn_restore.pack(side="left")

    def _selected_upcoming_entry() -> Optional[Dict[str, Any]]:
        sel = upcoming_tree.selection()
        if not sel:
            return None
        return upcoming_items.get(sel[0])

    def _selected_history_entry() -> Optional[Dict[str, Any]]:
        sel = history_tree.selection()
        if not sel:
            return None
        return history_items.get(sel[0])

    def _selected_schedule_entry() -> Optional[Dict[str, Any]]:
        entry = _selected_upcoming_entry()
        if entry is not None:
            return entry
        return _selected_history_entry()

    def _update_detail_buttons() -> None:
        if _selected_upcoming_entry() is not None:
            btn_mark_done.state(["!disabled"])
        else:
            btn_mark_done.state(["disabled"])
        if _selected_history_entry() is not None:
            btn_restore.state(["!disabled"])
        else:
            btn_restore.state(["disabled"])
        entry = _selected_schedule_entry()
        if entry is not None:
            btn_assign_card.state(["!disabled"])
            if entry.get("card"):
                btn_open_card.state(["!disabled"])
            else:
                btn_open_card.state(["disabled"])
        else:
            btn_assign_card.state(["disabled"])
            btn_open_card.state(["disabled"])

    def _populate_details(machine: Optional[Dict]) -> None:
        upcoming_items.clear()
        history_items.clear()
        for tree_view in (upcoming_tree, history_tree):
            for item in tree_view.get_children():
                tree_view.delete(item)
        if not machine:
            summary_var.set("Wybierz maszynę, aby zobaczyć harmonogram.")
            _update_detail_buttons()
            return
        summary = machine.get("__schedule_summary") or {}
        summary_var.set(summary.get("status_text") or "Brak danych harmonogramu")
        for entry in summary.get("upcoming", []):
            date_text = entry.get("date") or "—"
            typ = entry.get("type") or ""
            status_text, _status_key = _describe_entry_status(entry)
            notes = entry.get("notes") or ""
            iid = upcoming_tree.insert("", "end", values=(date_text, typ, status_text, notes))
            upcoming_items[iid] = entry
        for entry in summary.get("history", []):
            date_text = entry.get("date") or "—"
            typ = entry.get("type") or ""
            status_text, _status_key = _describe_entry_status(entry)
            notes = entry.get("notes") or ""
            iid = history_tree.insert("", "end", values=(date_text, typ, status_text, notes))
            history_items[iid] = entry
        _update_detail_buttons()

    def _set_selected_machine(machine_id: Optional[str]) -> None:
        nonlocal selected_machine_id
        selected_machine_id = machine_id
        if hall is not None:
            hall.select(machine_id)
        _populate_details(_find_machine(machine_id))

    def _mark_done() -> None:
        entry = _selected_upcoming_entry()
        if not entry:
            return
        entry["status"] = "wykonany"
        entry["completed_at"] = dt.datetime.now().isoformat()
        _on_schedule_changed()

    def _restore_plan() -> None:
        entry = _selected_history_entry()
        if not entry:
            return
        entry["status"] = "planowany"
        entry.pop("completed_at", None)
        _on_schedule_changed()

    def _assign_card() -> None:
        entry = _selected_schedule_entry()
        if not entry:
            return
        path = filedialog.askopenfilename(
            parent=root,
            title="Wybierz kartę przeglądu",
            filetypes=(("Dokumenty", "*.pdf;*.doc;*.docx;*.xlsx;*.xls;*.txt"), ("Wszystkie pliki", "*.*")),
        )
        if not path:
            return
        stored = _resolve_card_storage(path, cfg_manager)
        entry["card"] = stored
        _on_schedule_changed()

    def _open_selected_card() -> None:
        entry = _selected_schedule_entry()
        if not entry or not entry.get("card"):
            messagebox.showinfo("Karta przeglądu", "Brak przypisanej karty do tego wpisu.")
            return
        absolute = _resolve_card_absolute(str(entry.get("card")), cfg_manager)
        if not absolute or not os.path.exists(absolute):
            messagebox.showerror("Karta przeglądu", f"Plik nie istnieje: {absolute}")
            return
        if not _open_external(absolute):
            messagebox.showerror("Karta przeglądu", "Nie udało się otworzyć pliku.")

    def _on_upcoming_select(_event=None) -> None:
        if upcoming_tree.selection():
            history_tree.selection_remove(history_tree.selection())
        _update_detail_buttons()

    def _on_history_select(_event=None) -> None:
        if history_tree.selection():
            upcoming_tree.selection_remove(upcoming_tree.selection())
        _update_detail_buttons()

    def _selected_id() -> Optional[str]:
        sel = tree.selection()
        return str(sel[0]) if sel else None

    def _do_import() -> None:
        nonlocal schedule_entries, schedule_year, schedule_path
        path = filedialog.askopenfilename(
            parent=root,
            title="Wybierz plik harmonogramu",
            filetypes=(("Pliki Excel", "*.xlsx;*.xls;*.xlsm"), ("Wszystkie pliki", "*.*")),
        )
        if not path:
            return
        try:
            new_entries, meta = _import_schedule_from_excel(path)
        except RuntimeError as exc:
            messagebox.showerror("Import harmonogramu", str(exc))
            return
        except ValueError as exc:
            messagebox.showerror("Import harmonogramu", str(exc))
            return
        merged = _merge_schedule_status(new_entries, schedule_entries)
        schedule_entries = list(merged)
        schedule_meta["year"] = meta.get("year", schedule_meta.get("year", schedule_year))
        schedule_year = int(schedule_meta.get("year", schedule_year) or schedule_year)
        schedule_meta["source"] = meta.get("source", os.path.basename(path))
        schedule_meta["imported_at"] = meta.get("imported_at")
        schedule_path = resolve_schedule_path(schedule_year, cfg_manager)
        _on_schedule_changed()
        messagebox.showinfo("Import harmonogramu", f"Zaimportowano {len(schedule_entries)} wpisów.")

    class MachineEditDialog(tk.Toplevel):
        STATUSES = ("OK", "WARN", "ALERT")

        def __init__(self, master: tk.Misc, row: dict | None, on_ok):
            super().__init__(master)
            self.title("Edycja maszyny")
            self.resizable(False, False)
            self.transient(master)
            self.grab_set()
            self._row = dict(row or {})
            self._on_ok = on_ok
            frm = ttk.Frame(self)
            frm.pack(fill="both", expand=True, padx=12, pady=12)

            def row_entry(r, label, key):
                ttk.Label(frm, text=label, width=18, anchor="e").grid(
                    row=r, column=0, padx=6, pady=4, sticky="e"
                )
                ent = ttk.Entry(frm, width=36)
                ent.grid(row=r, column=1, padx=6, pady=4, sticky="w")
                ent.insert(0, str(self._row.get(key, "")))
                return ent

            self.e_id = row_entry(0, "ID / nr_ewid:", "id")
            self.e_nazwa = row_entry(1, "Nazwa:", "nazwa")
            self.e_typ = row_entry(2, "Typ:", "typ")
            self.e_lok = row_entry(3, "Lokalizacja:", "lokalizacja")

            ttk.Label(frm, text="Status:", width=18, anchor="e").grid(
                row=4, column=0, padx=6, pady=4, sticky="e"
            )
            self.cb_status = ttk.Combobox(frm, values=self.STATUSES, state="readonly", width=34)
            cur_status = str(self._row.get("status") or "").upper()
            self.cb_status.set(cur_status if cur_status in self.STATUSES else "OK")
            self.cb_status.grid(row=4, column=1, padx=6, pady=4, sticky="w")

            def int_or_none(value: str):
                try:
                    return int(value.strip())
                except Exception:
                    return None

            self.e_x = row_entry(5, "x (px):", "x")
            self.e_y = row_entry(6, "y (px):", "y")

            _build_edit_footer(frm, self._row, lambda: None)

            btns = ttk.Frame(frm)
            btns.grid(row=8, column=0, columnspan=2, pady=(10, 0))
            ttk.Button(btns, text="Anuluj", command=self.destroy).pack(side="right", padx=6)
            ttk.Button(
                btns,
                text="Zapisz",
                command=lambda: self._ok(int_or_none(self.e_x.get()), int_or_none(self.e_y.get())),
            ).pack(side="right", padx=6)
            self.bind(
                "<Return>",
                lambda *_: self._ok(int_or_none(self.e_x.get()), int_or_none(self.e_y.get())),
            )
            self.bind("<Escape>", lambda *_: self.destroy())

        def _ok(self, x, y):
            row = {
                "id": (
                    self.e_id.get().strip()
                    or self._row.get("id")
                    or self._row.get("nr_ewid")
                    or ""
                ),
                "nazwa": self.e_nazwa.get().strip(),
                "typ": self.e_typ.get().strip(),
                "lokalizacja": self.e_lok.get().strip(),
                "status": self.cb_status.get().strip() or "OK",
                "x": x,
                "y": y,
            }
            row.setdefault("nr_hali", "1")
            if isinstance(self._row.get("zadania"), list):
                row["zadania"] = self._row["zadania"]
            image_path = self._row.get("image") or self._row.get("obraz")
            if image_path:
                norm_path = os.path.normpath(image_path)
                row["image"] = norm_path
                row["obraz"] = norm_path
            if callable(self._on_ok):
                self._on_ok(row)
            self.destroy()

    def _on_add() -> None:
        def commit(new_row: Dict) -> None:
            nonlocal rows_cache
            if not new_row.get("id"):
                return
            new_row.setdefault("nr_hali", "1")
            new_row.setdefault("zadania", [])
            new_rows = upsert_machine(rows_cache, new_row)
            persisted = _save_rows(new_rows)
            rows_cache = list(persisted)
            _on_rows_changed()

        MachineEditDialog(container, row=None, on_ok=commit)

    def _on_edit() -> None:
        nonlocal rows_cache
        mid = _selected_id()
        if not mid:
            return
        current = _find_machine(mid)
        if not current:
            return

        def commit(upd: Dict) -> None:
            nonlocal rows_cache
            upd.setdefault("nr_hali", "1")
            if "zadania" not in upd and isinstance(current.get("zadania"), list):
                upd["zadania"] = current["zadania"]
            new_rows = upsert_machine(rows_cache, upd)
            persisted = _save_rows(new_rows)
            rows_cache = list(persisted)
            _on_rows_changed()

        MachineEditDialog(container, row=current, on_ok=commit)

    def _on_del() -> None:
        nonlocal rows_cache
        mid = _selected_id()
        if not mid:
            return
        if messagebox.askyesno("Usuń", f"Czy usunąć maszynę: {mid}?"):
            new_rows = delete_machine(rows_cache, mid)
            persisted = _save_rows(new_rows)
            rows_cache = list(persisted)
            _on_rows_changed()

    def _on_save() -> None:
        nonlocal rows_cache
        for row in rows_cache:
            if "nr_hali" not in row or row.get("nr_hali") in (None, ""):
                row["nr_hali"] = "1"
        persisted = _save_rows(rows_cache)
        rows_cache = list(persisted)
        _on_rows_changed()
        info.set(f"Zapisano {len(rows_cache)} maszyn.")

    def _on_tree_select(_event=None) -> None:
        _set_selected_machine(_selected_id())

    btn_import.configure(command=_do_import)
    btn_add.configure(command=_on_add)
    btn_edit.configure(command=_on_edit)
    btn_del.configure(command=_on_del)
    btn_save.configure(command=_on_save)
    btn_mark_done.configure(command=_mark_done)
    btn_restore.configure(command=_restore_plan)
    btn_assign_card.configure(command=_assign_card)
    btn_open_card.configure(command=_open_selected_card)

    tree.bind("<<TreeviewSelect>>", _on_tree_select)
    tree.bind("<Double-1>", lambda _e: _on_edit())
    upcoming_tree.bind("<<TreeviewSelect>>", _on_upcoming_select)
    history_tree.bind("<<TreeviewSelect>>", _on_history_select)
    filter_box.bind("<<ComboboxSelected>>", _apply_filter)

    _refresh_schedule_info()
    _recompute_visible_rows()
    _refresh_tree()
    _populate_details(None)

    logger.info("[Maszyny] Panel otwarty; rekordów: %d", len(rows_cache))
    return tree


def panel_maszyny(root, frame, login=None, rola=None):
    _open_maszyny = _open_machines_panel  # alias nazwy
    if open_dyspo_wizard is not None:
        toolbar = ttk.Frame(frame)
        toolbar.pack(fill="x", padx=6, pady=(6, 0))
        target = root
        if hasattr(root, "winfo_toplevel"):
            try:
                target = root.winfo_toplevel()
            except Exception:
                target = root
        ttk.Button(
            toolbar,
            text="Nowa dyspozycja…",
            command=lambda: _maybe_open_dyspo(
                target, {"module": "Maszyny"}
            ),
        ).pack(side=tk.RIGHT)
        bind_ctrl_d(target, context={"module": "Maszyny"})
    _open_maszyny(root, frame, Renderer=None)


def init_maszyny_view(
    parent: tk.Misc,
    lista_maszyn: Optional[List[Dict[str, object]]] = None,
    logout_cb: Optional[Callable[[], None]] = None,
    quit_cb: Optional[Callable[[], None]] = None,
    reset_cb: Optional[Callable[[], None]] = None,
) -> MachinesView:
    """Zainicjalizuj widok maszyn w trybie uproszczonym."""

    cfg = Settings(path="config.json", project_root=__file__)
    bg_path = cfg.path_assets("hala.png")
    view = MachinesView(
        parent,
        cfg,
        bg_path=bg_path,
        logout_cb=logout_cb,
        quit_cb=quit_cb,
        reset_cb=reset_cb,
    )
    view.set_records(lista_maszyn or [])
    return view


if __name__ == "__main__":
    root = tk.Tk()
    root.title("Warsztat Menager — Maszyny")
    ensure_theme_applied(root)
    main = tk.Frame(root)
    main.pack(fill="both", expand=True)
    _open_machines_panel(root, main, Renderer=None)
    root.mainloop()

